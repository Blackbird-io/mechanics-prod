# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2017
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: xl_parser
"""

Module provides tools for parsing an excel file into Financials values.

====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:

FUNCTIONS:
add_projections()          creates a EngineModel with projections values from xl

_build_fins_from_sheet()    builds LineItem structure
build_sheet_maps()     checks if upload xl projections is the right format
_populate_fins_from_sheet() writes projected line values to financials


CLASSES:
n/a
====================  =========================================================
"""



# Imports
import bb_settings
import bb_exceptions
import openpyxl as xlio
import json

from data_structures.modelling.model import Model
from data_structures.modelling.business_unit import BusinessUnit
from data_structures.modelling.line_item import LineItem
from data_structures.modelling.statement import Statement
from data_structures.modelling.time_line import TimeLine
from data_structures.modelling.time_period import TimePeriod

from formula_manager import local_catalog as FC
from datetime import datetime, date, timedelta
from openpyxl.formula import Tokenizer
from . import sheet_map
from . import parser_settings as ps




def add_projections(xl_serial, engine_model):
    """


    add_projections(xl_serial, engine_model) -> EngineModel()

    --``xl_serial``     serialized string of an Excel workbook (".xlsx")
    --``engine_model``  instance of Engine Model

    Function takes a serialized Excel workbook in a specific format
    and converts it to an EngineModel with LineItem values. Model input will
    determine which investment card we are adding projections to.

    Function delegates to:
        build_sheet_map()
        _build_fins_from_sheet()
        _populate_fins_from_sheet()

    """
    model = engine_model

    # 1) Extract xl_serial. Make sure it is in the right format
    wb = xlio.load_workbook(xl_serial, data_only=True)  # Includes Values Only
    wb_f = xlio.load_workbook(xl_serial, data_only=False)  # Includes Formulas

    # For testing local excel files:
    # filename = r"C:\Workbooks\Forecast_Rimini8.xlsx"
    # wb = xlio.load_workbook(filename=filename, data_only=True)

    # Look for the BB Metadata tab
    metadata_names = ps.BB_METADATA_NAMES

    bb_tabname = None
    for tab in wb:
        if tab.title.casefold() in metadata_names:
            bb_tabname = tab.title
            break
    if not bb_tabname:
        c = "No BB Metadata Tab!"
        raise bb_exceptions.BBAnalyticalError(c)

    sheet = wb[bb_tabname]
    sheet_f = wb_f[bb_tabname]

    # Make sure sheet is valid format
    sm = build_sheet_map(sheet)

    # 2) Align model.time_line to ref_date. Add additional periods as needed
    sheet_rows = [r for r in sheet.rows]
    header_row = sheet_rows[sm.rows["HEADER"]-1]
    xl_dates = []
    for cell in header_row[sm.cols["FIRST_PERIOD"]-1:]:
        if isinstance(cell.value, datetime):
            xl_dates.append(cell.value.date())

    first_date = xl_dates[0]
    if isinstance(first_date, str):
        first_date = datetime.strptime(first_date, "%Y-%m-%d").date()

    # set default timeline to only contain relevant dates
    new_timeline = TimeLine(model)
    model.set_timeline(new_timeline, overwrite=True)
    model.time_line.build(first_date, fwd=0, year_end=False)
    model.change_ref_date(first_date)

    company = model.get_company()
    if not company:
        company = BusinessUnit(model.name)
        model.set_company(company)
        model.target = company

    # 3) Determine Line Structure from Excel upload
    _build_fins_from_sheet(company, sheet, sm)

    # 4) Make sure actuals timeline has same structure
    actl_tl = model.get_timeline('monthly', name='actual')
    if not actl_tl:
        actl_tl = TimeLine(model)
        model.set_timeline(actl_tl, resolution='monthly', name='actual')

    # 5) Write values to both actuals and projected.
    _populate_fins_from_sheet(model, sheet, sheet_f, sm)

    return model


def revise_projections(xl_serial, old_model):
    """


    revise_projections(xl_serial, engine_model) -> EngineModel()

    --``xl_serial``     serialized string of an Excel workbook (".xlsx")
    --``old_model``     old instance of Engine Model

    Function revises projection values while keeping existing actuals values
    Function takes a serialized Excel workbook in a specific format
    and converts it to an EngineModel with LineItem values. 

    Function delegates to:
        build_sheet_maps()
        _build_fins_from_sheet()
        _populate_fins_from_sheet()

    """
    new_model = Model(bb_settings.DEFAULT_MODEL_NAME)
    new_model.start()
    new_model._ref_date = old_model._ref_date

    # 1) Extract xl_serial. Make sure it is in the right format
    wb = xlio.load_workbook(xl_serial, data_only=True)  # Includes Values Only
    wb_f = xlio.load_workbook(xl_serial, data_only=False)  # Includes Formulas

    # For testing local excel files:
    # filename = r"C:\Workbooks\Forecast_Rimini8.xlsx"
    # wb = xlio.load_workbook(filename=filename, data_only=True)

    # Look for the BB Metadata tab
    metadata_names = ps.BB_METADATA_NAMES

    bb_tabname = None
    for tab in wb:
        if tab.title.casefold() in metadata_names:
            bb_tabname = tab.title
            break
    if not bb_tabname:
        c = "No BB Metadata Tab!"
        raise bb_exceptions.BBAnalyticalError(c)

    sheet = wb[bb_tabname]
    sheet_f = wb_f[bb_tabname]

    # Make sure sheet is valid format
    sm = build_sheet_map(sheet)

    company = new_model.get_company()
    if not company:
        company = BusinessUnit(new_model.name)
        new_model.set_company(company)
        new_model.target = company

    # 2) Determine Line Structure from Excel upload
    _build_fins_from_sheet(company, sheet, sm)

    # 3) Add any missing lines from old model
    _combine_fins_structure(old_model, new_model)

    # 4) Make sure actuals timeline has same structure
    actl_tl = new_model.get_timeline('monthly', name='actual')
    if not actl_tl:
        actl_tl = TimeLine(new_model)
        new_model.set_timeline(actl_tl, resolution='monthly', name='actual')

    # 5) Write values to both actuals and projected.
    _populate_fins_from_sheet(new_model, sheet, sheet_f, sm)

    # 6) Retain any actual values from old model that are hardcoded
    _populate_old_actuals(old_model, new_model)

    return new_model


def build_sheet_map(sheet):
    """


    build_sheet_map(sheet) -> SheetMap()

    --``sheet`` is an instance of openpyxl.WorkSheets

    Function checks if the uploaded excel sheet is in the right format.
    Raises Error if format is violated. Function also finds and sets the Column
    numbers for each field.
    """
    # Initialize a SheetMap object
    sm = sheet_map.SheetMap()
    sm.rows["DATES"] = 1
    sm.rows["TIMELINE"] = 2  # "Actual" or "Forecast"
    sm.rows["FIRST_DATA"] = 3  # First row with LineItem data

    sheet_rows = [r for r in sheet.rows]
    header_row = sheet_rows[sm.rows["HEADER"]-1]

    # Next few columns can be in any order
    for cell in header_row:
        if isinstance(cell.value, datetime):
            sm.cols["FIRST_PERIOD"] = cell.col_idx
            break

        # Ensure backwards compatibility with old column names
        col_name = cell.value
        if col_name in ps.ALERT:
            col_name = ps.ALERT[0]
        elif col_name in ps.ADD_TO_PATH:
            col_name = ps.ADD_TO_PATH[0]

        if col_name in sm.cols.keys():
            sm.cols[col_name] = cell.col_idx
        else:
            c = "col_name not in sm.cols.keys()"
            raise bb_exceptions.ExcelPrepError(c)

    missing_columns = []
    for col in ps.REQUIRED_COLS:
        if sm.cols[col] is None:
            missing_columns.append(col)
    if missing_columns:
        c = "No header for %s" % missing_columns
        raise bb_exceptions.ExcelPrepError(c)

    # Make sure everything after sm.cols["FIRST_PERIOD"] is all in a date format
    for cell in header_row[sm.cols["FIRST_PERIOD"]-1:]:
        if cell.value is None and cell is header_row[sm.cols["FIRST_PERIOD"]-1:][-1]:
            # Sometimes the last blank column cell is read in
            continue
        if not isinstance(cell.value, datetime):
            # # print(cell.value, type(cell.value))
            c = "Header in %s is not a datetime obj!" % cell.coordinate
            raise bb_exceptions.ExcelPrepError(c)

    return sm


def _build_fins_from_sheet(bu, sheet, sm):
    """


    _build_fins_from_sheet(financials, sheet) -> None

    --``bu`` is the instance of Business Unit we want build Financials on
    --``sheet`` is an instance of openpyxl.WorkSheet
    --``sm`` is an instance of SheetMap

    Function extracts information from 'sheet' to build the LineItem structure
    in 'financials'. Functions creates new statements if necessary.
    """
    full_order = list()
    financials = bu.financials  # SSOT Fins
    line = None

    # Loop through each row that contains LineItem information
    # Start at the first row that has line data
    iter_rows = sheet.iter_rows(row_offset=sm.rows["FIRST_DATA"]-1,
                                column_offset=0)
    for row in iter_rows:

        if not row[sm.cols[ps.STATEMENT]-1].value:
            if line:
                line.xl_format.blank_row_after = True
            # Skip blank rows
            continue

        statement_name = row[sm.cols[ps.STATEMENT]-1].value.strip()

        if "parameter" in statement_name.casefold():
            continue  # Ignore parameters

        # put statements in order
        if statement_name.casefold() not in full_order:
            full_order.append(statement_name.casefold())

        statement = financials.get_statement(statement_name)

        if not statement and statement_name:
            statement = Statement(statement_name)
            financials.add_statement(name=statement_name, statement=statement)

        if statement is financials.valuation:
            statement.compute = True

        line_name = row[sm.cols[ps.LINE_NAME]-1].value
        if not line_name:
            continue  # Must have line name

        line_title = row[sm.cols[ps.LINE_TITLE]-1].value
        if not line_title:
            continue  # Must have line title

        parent_name = row[sm.cols[ps.PARENT_NAME]-1].value or ""
        # find_first always needs a str.

        num_parents = len(statement.find_all(parent_name))
        if num_parents > 1:
            c = "PARENT_NAME must be unique within a statement!"
            c += " There are %s occurances of %s" % (num_parents, parent_name)
            # Current design flaw is that if there are more than one
            # parent lines with the same name, we do not know which one to use
            raise bb_exceptions.BBAnalyticalError(c)

        parent = statement.find_first(parent_name)
        if not parent:
            if parent_name:
                # Add parent as a top level line, can be moved later
                parent = LineItem(parent_name)
                statement.append(parent)

        # Main logic below
        existing_line = statement.find_first(line_name)
        # existing_line could be the line we want or another line with the same
        # name but under a different parent

        if not existing_line:
            line = LineItem(line_name)
            if parent:
                parent.append(line)
            else:
                statement.append(line)

        else:
            # A line with the same name exists on statement
            # We want to check if in the right location
            if parent:
                if statement.find_first(parent_name, line_name):
                    # Existing_line is the line we want to edit
                    # It is already under the correct parent
                    line = existing_line
                else:
                    if existing_line.relationships.parent is statement:
                        # Existing_line was first created as a parent line,
                        # Move existing_line to the correct parent
                        # Existing_line is the line we want to edit
                        line = statement.find_first(line_name, remove=True)
                        parent.append(line)

                    else:
                        # Existing_line is NOT the line we want to edit
                        # Create a new line, same name, different parent
                        line = LineItem(line_name)
                        parent.append(line)
            else:
                # Existing_line is the line we want to edit
                # Its location is correct at the top level
                line = existing_line

        if not line:
            c = "Import Error: Line %s, Parent %s" % (line_name, parent_name)
            c += ". Check financials structure."
            raise bb_exceptions.BBAnalyticalError(c)

        line.set_title(line_title)
        _add_line_effects(line, bu, row, sm)

    financials.set_order(full_order)
    print(financials)


def _add_line_effects(line, bu, row, sm):
    """


    _add_line_effects(financials, sheet) -> None

    --``line`` is the target LineItem
    --``bu`` is the instance of Business Unit that the line belongs to
    --``row`` is list of openpyxl.Cells in the same row
    --``sm`` is an instance of SheetMap

    Function validates column values. This information and stores on the line
    so that the Engine knows what to do with line later.
    """

    # Add comparison ("<" or ">") as a tag for KPI and Covenant analysis
    if sm.cols[ps.COMPARISON]:
        comparison_str = row[sm.cols[ps.COMPARISON]-1].value
        if comparison_str in ('<', '<=', '>', '>='):
            # Only tag valid comparisons
            line.tags.add(comparison_str)

    # Add sum_details attribute if FALSE (TRUE is default for blank cells)
    if sm.cols[ps.SUM_DETAILS]:
        sum_details = row[sm.cols[ps.SUM_DETAILS]-1].value
        if not (_check_truthy(sum_details) or sum_details is None):
            line.sum_details = False

    # Tag line with which summary report we want to display it on.
    if sm.cols[ps.REPORT]:
        report_str = row[sm.cols[ps.REPORT]-1].value
        if _check_truthy(report_str, others=ps.VALID_REPORTS):
            line.usage.show_on_report = True

    if sm.cols[ps.MONITOR]:
        monitor_bool = row[sm.cols[ps.MONITOR]-1].value
        if _check_truthy(monitor_bool):
            line.usage.monitor = True

    if sm.cols[ps.PARSE_FORMULA]:
        parse_formula_bool = row[sm.cols[ps.PARSE_FORMULA]-1].value
        if _check_truthy(parse_formula_bool):
            line.tags.add('parse formula')

    if sm.cols[ps.ADD_TO_PATH[0]]:
        topic_formula_bool = row[sm.cols[ps.ADD_TO_PATH[0]] - 1].value
        if _check_truthy(topic_formula_bool):
            new_path_line = LineItem(line.name)
            bu.stage.path.append(new_path_line)
            line.tags.add('topic formula')

    if sm.cols[ps.ALERT[0]]:
        alert_val = row[sm.cols[ps.ALERT[0]] - 1].value
        if alert_val is not None:
            new_path_line = LineItem(line.name + " alert")
            new_path_line.tags.add('alert commentary')
            bu.stage.path.append(new_path_line)
            line.tags.add('alert commentary')

            # Backwards compatibility when ALERT was a bool column
            if _check_truthy(alert_val):
                alert_val = '{"comparison":"=","limit":"Needs Review"}'

            try:
                conditions_dict = json.loads(alert_val)
                line.usage.alert_commentary = conditions_dict
            except ValueError:
                c = "Invalid JSON String: " + alert_val
                raise bb_exceptions.BBAnalyticalError(c)

    if sm.cols[ps.ON_CARD]:
        on_card_bool = row[sm.cols[ps.ON_CARD] - 1].value
        if _check_truthy(on_card_bool):
            line.usage.show_on_card = True

    # Tag line with one or more tags.
    if sm.cols[ps.TAGS]:
        tags_str = row[sm.cols[ps.TAGS]-1].value
        if tags_str:
            tags_list = tags_str.split(",")
            for t in tags_list:
                new_tag = t.strip()  # Remove white space from both sides
                line.tags.add(new_tag)


def _check_truthy(var, others=list()):
    """


    _build_fins_from_sheet(financials, sheet) -> None

    --``old_model``     old instance of Engine Model
    --``new_model``     new instance of Engine Model

    Function combines line structure from the old and new models
    """
    truths = ["true", "yes"]

    others = [o.casefold() for o in others if isinstance(o, str)]
    truths.extend(others)

    val = False
    if isinstance(var, str):
        var = var.casefold()
        if var in truths:
            val = True
    elif isinstance(var, bool):
        val = var
    elif var in truths:
        val = True

    return val


def _combine_fins_structure(old_model, new_model):
    """


    _build_fins_from_sheet(financials, sheet) -> None

    --``old_model``     old instance of Engine Model
    --``new_model``     new instance of Engine Model

    Function combines line structure from the old and new models
    """
    old_bu = old_model.get_company()
    new_bu = new_model.get_company()

    old_fins = old_bu.financials
    new_fins = new_bu.financials

    for old_stmt in old_fins.full_ordered:
        if not old_stmt:
            # Valuation is empty
            continue

        stmt_name = old_stmt.name.casefold().strip()
        new_stmt = new_fins.get_statement(stmt_name)
        if not new_stmt:
            c = 'Old %s doesn exist!' % old_stmt.name
            raise bb_exceptions.BBAnalyticalError(c)

        last_position = 0
        for old_line in old_stmt.get_full_ordered():
            old_parent = old_line.relationships.parent  # Can be stmt or line
            if isinstance(old_parent, LineItem):
                new_line = new_stmt.find_first(old_parent.name, old_line.name)
            else:
                new_line = new_stmt.find_first(old_line.name)

            if not new_line:
                new_line = old_line.copy()
                if isinstance(old_parent, LineItem):
                    new_parent = new_stmt.find_first(old_parent.name)
                    new_parent.add_line(new_line, last_position + 1)
                else:
                    new_stmt.add_line(new_line, last_position + 1)

            last_position = new_line.position


def _populate_old_actuals(old_model, new_model):
    """


    _populate_old_actuals() -> None

    --``old_model``     old instance of Engine Model
    --``new_model``     new instance of Engine Model

    Function looks for hardcoded actuals in old model and copies thier value 
    onto the the new model.
    """
    old_bu = old_model.get_company()
    new_bu = new_model.get_company()

    old_actl_tl = old_model.get_timeline(resolution='monthly', name='actual')
    new_actl_tl = new_model.get_timeline(resolution='monthly', name='actual')
    new_proj_tl = new_model.get_timeline(resolution='monthly', name='default')

    for old_pd in old_actl_tl.iter_ordered():

        # If there are actuals, we want to populate both actl and proj timelines
        for tl in [new_actl_tl, new_proj_tl]:
            new_pd = tl.find_period(old_pd.end)
            if not new_pd:
                # new_pd = old_pd.copy()
                new_pd = TimePeriod(start_date=old_pd.start,
                                    end_date=old_pd.end,
                                    model=new_model)
                tl.add_period(new_pd)

            old_fins = old_bu.get_financials(old_pd)
            new_fins = new_bu.get_financials(new_pd)

            for old_stmt in old_fins.full_ordered:
                if not old_stmt:
                    # Valuation is empty
                    continue

                stmt_name = old_stmt.name.casefold().strip()
                new_stmt = new_fins.get_statement(stmt_name)
                if not new_stmt:
                    c = 'Old %s doesn exist!' % old_stmt.name
                    raise bb_exceptions.BBAnalyticalError(c)

                for old_line in old_stmt.get_full_ordered():
                    old_parent = old_line.relationships.parent
                    if isinstance(old_parent, LineItem):
                        new_line = new_stmt.find_first(old_parent.name,
                                                       old_line.name)
                    else:
                        new_line = new_stmt.find_first(old_line.name)

                    if old_line.hardcoded:
                        new_line.set_value(old_line.value,
                                           signature='old hardcoded value',
                                           override=True)
                        new_line.set_hardcoded(True)


def _populate_fins_from_sheet(engine_model, sheet, sheet_f, sm):
    """


    _populate_fins_from_sheet() -> None

    --``engine_model`` is the instance of EngineModel
    --``sheet`` is an instance of openpyxl.WorkSheet with Values
    --``sheet_f`` is an instance of openpyxl.WorkSheet with Formulas as str
    --``sm`` is an instance of SheetMap

    Function extracts LineItem values from 'sheet' and writes them to
    'financials' for multiple periods. Function assumes that the structure
    for financials is already in place for every period.
    """
    model = engine_model
    bu = model.get_company()
    proj_tl = model.get_timeline('monthly', name='default')
    actl_tl = model.get_timeline('monthly', name='actual')
    ssot_fins = bu.financials

    # Loop across periods (Left to Right on Excel)
    sheet_columns = [c for c in sheet.columns]
    for col in sheet_columns[sm.cols["FIRST_PERIOD"]-1:]:
        dt = col[0].value.date()
        timeline_name = col[sm.rows["TIMELINE"]-1].value

        start_dt = date(dt.year, dt.month, 1)
        end_dt = date(dt.year, dt.month, 28)
        while end_dt.month == start_dt.month:
            end_dt += timedelta(1)
        end_dt -= timedelta(1)

        # Always add a proj_pd regardless of if something is forecast or actual
        proj_pd = proj_tl.find_period(dt)
        if not proj_pd:
            proj_pd = TimePeriod(start_date=start_dt,
                                 end_date=end_dt,
                                 model=model)
            proj_tl.add_period(proj_pd)

        actl_pd = actl_tl.find_period(dt)

        if _check_truthy(timeline_name, others=["actual"]):
            if not actl_pd:
                actl_pd = TimePeriod(start_date=start_dt,
                                     end_date=end_dt,
                                     model=model)
                actl_tl.add_period(actl_pd)

        # Loop across Lines (Top to Down on Excel)
        for cell in col[sm.rows["FIRST_DATA"]-1:]:
            # Skip blank cells
            if cell.value in (None, ""):
                behavior_str = ''
                if sm.cols[ps.BEHAVIOR]:
                    behavior_str = sheet.cell(row=cell.row,
                                               column=sm.cols[ps.BEHAVIOR]).value

                if not behavior_str:
                    continue

            row_num = cell.row
            col_num = cell.col_idx

            statement_str = sheet.cell(row=row_num, column=sm.cols[ps.STATEMENT]).value
            if not statement_str:
                continue

            statement_name = statement_str.casefold().strip()
            line_name = sheet.cell(row=row_num, column=sm.cols[ps.LINE_NAME]).value
            parent_name = sheet.cell(row=row_num, column=sm.cols[ps.PARENT_NAME]).value

            # Handle rows where we just want to add a parameter value
            if statement_name in ("parameters", "parameter"):
                if parent_name in ("Timeline", "timeline"):
                    if cell.value is not None:
                        actl_tl.parameters.add({line_name: cell.value})
                        proj_tl.parameters.add({line_name: cell.value})
                else:
                    if actl_pd:
                        actl_pd.parameters.add({line_name: cell.value})
                    if proj_pd:
                        proj_pd.parameters.add({line_name: cell.value})
                # # print("Param", cell.value)
                continue

            # # print(statement_name, line_name, parent_name)
            if actl_pd:
                actl_fins = bu.get_financials(actl_pd)
                actl_stmt = actl_fins.get_statement(statement_name)
            proj_fins = bu.get_financials(proj_pd)
            proj_stmt = proj_fins.get_statement(statement_name)

            # Always match number formats
            ssot_stmt = ssot_fins.get_statement(statement_name)
            ssot_line = ssot_stmt.find_first(line_name)

            if cell.number_format and not ssot_line.xl_format.number_format:
                ssot_line.xl_format.number_format = cell.number_format

            # Status column
            if sm.cols[ps.STATUS]:
                status_cell = sheet.cell(row=row_num, column=sm.cols[ps.STATUS])
                status_str = status_cell.value

                if status_str:
                    try:
                        status_dict = json.loads(status_str)
                    except ValueError:
                        c = "Invalid JSON String: " + status_str
                        raise bb_exceptions.BBAnalyticalError(c)

                    required_keys = {"style"}
                    missing_keys = required_keys - status_dict.keys()
                    if len(missing_keys) > 0:
                        c = "Missing Keys: " + missing_keys
                        raise bb_exceptions.BBAnalyticalError(c)

                    # If a key is "0.8", turn it into 0.8
                    for key in status_dict:
                        try:
                            float(key)
                        except ValueError:
                            continue
                        else:
                            status_value = status_dict.pop(key)
                            status_dict[float(key)] = status_value

                    ssot_line.usage.status_rules = status_dict

                if ssot_line.usage.status_rules and ssot_stmt.display_type == \
                        ssot_stmt.REGULAR_TYPE:
                    ssot_stmt.display_type = ssot_stmt.KPI_TYPE

            # Behaviour column
            if sm.cols[ps.BEHAVIOR]:
                behavior_cell = sheet.cell(row=row_num,
                                           column=sm.cols[ps.BEHAVIOR])
                behavior_str = behavior_cell.value

                limits_str = None
                if sm.cols[ps.LIMITS]:
                    limits_cell = sheet.cell(row=row_num,
                                             column=sm.cols[ps.LIMITS])
                    limits_str = limits_cell.value

                if behavior_str:
                    _set_behavior(behavior_str, limits_str, ssot_stmt,
                                  ssot_line, model)
                    continue  # Don't parse or hardcode line if behavior exists

            # Look to see if Formula should be automatically imported
            cell_f = sheet_f.cell(row=row_num, column=col_num)

            can_parse = _parse_formula(sheet, cell_f, bu, sm)

            # Otherwise, set the lowest level lines to a hardcoded value
            if can_parse is False:
                # Always populate projected statement. (Include Historical Actuals)
                _populate_line_from_cell(cell, line_name, parent_name, proj_stmt)

                if _check_truthy(timeline_name, others=["actual"]):
                    _populate_line_from_cell(cell, line_name, parent_name, actl_stmt)

            if ssot_line.usage.status_rules and ssot_stmt.display_type == \
                    ssot_stmt.REGULAR_TYPE:
                ssot_stmt.display_type = ssot_stmt.KPI_TYPE


def _set_behavior(behavior_str, limits_str, statement, line, model):
    """


    _set_behavior(sheet, cell_f, bu) -> bool

    --``behavior_str`` is string representing a JSON dict
    --``limits_str`` is string representing of a list of 2 item lists
    --``line`` is the instance of LineItem which we want to add behavior
    --``model`` is an instance of EngineModel    

    Function takes a behavior and and limits string and figures out a formula
    and driver to attach to the line.
    """

    try:
        behavior_dict = json.loads(behavior_str)
    except ValueError:
        c = "Invalid JSON String: " + behavior_str
        raise bb_exceptions.BBAnalyticalError(c)

    line.usage.behavior = behavior_dict

    if not limits_str:
        # set default value
        limits_str = '[[1.3,"Overperforming"],[1.1,"Performing"]]'

    try:
        limits_list = json.loads(limits_str)
    except ValueError:
        c = "Invalid JSON String: " + limits_str
        raise bb_exceptions.BBAnalyticalError(c)

    line.usage.limits = limits_list

    if behavior_dict.get('action'):
        action_name = behavior_dict.pop('action')
    else:
        action_name = "rolling sum over time"

    if action_name.casefold().strip() == "custom status":
        statement.display_type = statement.COVENANT_TYPE

    formula_name = ps.ACTION_TO_FORMULA_MAP.get(action_name)
    if not formula_name:
        c = "No formula mapped to the action name of: " + action_name
        raise bb_exceptions.BBAnalyticalError(c)

    f_id = FC.by_name[formula_name]
    formula = FC.issue(f_id)
    if not formula:
        c = "No formula found by name of: " + formula_name
        raise bb_exceptions.BBAnalyticalError(c)

    required_keys = formula.required_data

    if len(required_keys - behavior_dict.keys()) == 0:
        if not line.get_driver():
            data = dict()
            data.update(behavior_dict)  # Same key names

            parent_name = line.relationships.parent.name
            dr_name = (parent_name or "") + ">" + line.name
            driver = model.drivers.get_or_create(dr_name, data, formula)
            line.assign_driver(driver.id.bbid)


def _parse_formula(sheet, cell_f, bu, sm):
    """


    _parse_formula(sheet, cell_f, bu) -> bool

    --``sheet`` is an instance of openpyxl.WorkSheet with Values
    --``cell_f`` is an instance of openpyxl.Cell with Formulas as str
    --``bu`` is the instance of EngineModel.BusinessUnit
    --``sm`` is an instance of SheetMap    

    Function takes a formula string from excel and creates a driver that will
    provide the same value in the Blackbird Engine. Function returns False if
    it was not able to parse the formula and True if the driver was added.
    """
    f_id = FC.by_name["custom formula from tokens."]
    formula = FC.issue(f_id)

    row = cell_f.row
    col = cell_f.col_idx

    cell = sheet.cell(row=row, column=col)  # data only cell
    if cell.value is None:
        return False

    parse_formula_bool_cell = sheet.cell(row=row, column=sm.cols[ps.PARSE_FORMULA])
    if not _check_truthy(parse_formula_bool_cell.value):
        return False

    line_name = sheet.cell(row=row, column=sm.cols[ps.LINE_NAME]).value
    parent_name = sheet.cell(row=row, column=sm.cols[ps.PARENT_NAME]).value
    stmt_name = sheet.cell(row=row, column=sm.cols[ps.STATEMENT]).value

    stmt_str = stmt_name.casefold().strip()
    statement = bu.financials.get_statement(stmt_str)

    ancestors = [line_name]
    if parent_name:
        ancestors.insert(0, parent_name)  # Insert in 1st position
    line = statement.find_first(*ancestors)

    if not line:
        print("No line found!", ancestors)
        return False

    formula_str = cell_f.value

    if not isinstance(formula_str, str):
        return False
    elif "!" in formula_str:
        return False  # Don't include other tabs "=Sheet2!A1"
    else:
        tok = Tokenizer(formula_str)

        for t in tok.items:
            if t.type == "FUNC" and t.subtype in ("OPEN", "CLOSE"):
                # # print(t.value)
                if t.value in ("IF(", "MAX(", "MIN(", "ROUND(", ")"):
                    # Allow simple functions like "=MAX(A1,B1)"
                    pass
                else:
                    # Don't include all other functions like "=SUM(A1)"
                    return False
            if ":" in t.value and t.subtype == "RANGE":
                return False  # Don't include ranged sources "A1:A8"

    # if col == sm.cols["FIRST_PERIOD"]:  # Only insert drivers in first column
    if not line.get_driver():
        data = dict()

        i = 0
        for t in tok.items:
            i += 1

            t_name = "token%02d" % i  # token01, token02, ... token99
            t_type = t_name + "_type"
            t_fixed_dt = t_name + "_fixed_date"
            t_rel_pd = t_name + "_relative_pd"

            # # print("Token Attr:", t_name, t.value, t.type, t.subtype)

            if t.type == "OPERAND" and t.subtype == "NUMBER":  # "100"
                data[t_name] = t.value
                data[t_type] = "constant"

            elif t.type == "OPERAND" and t.subtype == "LOGICAL":  # "TRUE"
                if t.value == "TRUE":
                    data[t_name] = True
                elif t.value == "FALSE":
                    data[t_name] = False
                # data[t_name] = t.value.title()  # "FALSE" -> "False"
                data[t_type] = "constant"

            elif t.type == "OPERAND" and t.subtype == "TEXT":  # "EBITDA < 0"
                # t.value might be '"EBITDA<="', we just want "EBITDA<="
                if t.value[1:-1] in ps.ALLOWABLE_XL_TEXT:
                    data[t_name] = t.value
                else:
                    data[t_name] = ""

                data[t_type] = "constant"

            elif t.type == "OPERAND" and t.subtype == "RANGE":  # "A1"
                source_addr = t.value
                source_cell = sheet.cell(source_addr)
                source_row = source_cell.row
                source_col = source_cell.col_idx

                source_line_name = sheet.cell(row=source_row,
                                              column=sm.cols[ps.LINE_NAME]).value
                source_statement = sheet.cell(row=source_row,
                                              column=sm.cols[ps.STATEMENT]).value
                source_statement = source_statement.casefold().strip()
                data[t_name] = source_line_name
                data[t_type] = "source"
                data[t_name + "_statement"] = source_statement
                if source_addr[0] == "$":
                    # Fixed Dates (source is always same column)
                    data[t_fixed_dt] = sheet.cell(row=sm.rows["DATES"],
                                                  column=source_col).value
                elif source_col != cell_f.col_idx:
                    # Relative Periods (n periods past or future)
                    data[t_rel_pd] = source_col - cell_f.col_idx

            elif t.type == "OPERATOR-INFIX":  # +, -, *, /, =
                if t.value == "=":
                    # Equality comparisons are "=" in Excel and "==" in Python
                    data[t_name] = t.value = "=="
                else:
                    data[t_name] = t.value
                data[t_type] = "operator"

            elif t.type == "OPERATOR-PREFIX":  # 1st character +, -, *, /, =
                data[t_name] = t.value
                data[t_type] = "operator"

            elif t.type == "PAREN":  # "(" or ")"
                data[t_name] = t.value
                data[t_type] = "operator"

            elif t.type == "FUNC":
                if t.value in ("MAX(", "MIN(", "ROUND("):
                    # Excel functions same as lower case Python functions
                    data[t_name] = t.value.casefold()
                else:
                    # Excel functions require a custom Python function ie "IF()"
                    data[t_name] = t.value
                data[t_type] = "operator"

            elif t.type == "SEP" and t.subtype == "ARG":
                data[t_name] = t.value.casefold()
                data[t_type] = "operator"

        model = bu.relationships.model
        dr_name = stmt_name + '>' + (parent_name or "") + ">" + line_name
        driver = model.drivers.get_or_create(
            dr_name,
            data,
            formula
        )
        line.assign_driver(driver.id.bbid)

    return True


def _populate_line_from_cell(cell, line_name, parent_name, statement):
    """


    _populate_fins_from_sheet() -> None

    --``cell``        Cell() instance from Openpyxl
    --``line_name``   string, name of LineItem we want to write to
    --``parent_name`` string or None, name of
    --``statement``   Statement() instance

    Function writes a hardcoded line value to a statement
    """
    value = cell.value
    if value is None:
        return

    parent = statement.find_first(parent_name or "")
    if parent:
        line = parent.find_first(line_name)
    else:
        line = statement.find_first(line_name)

    if line is None:
        print(line_name, parent_name)

    if line._details and line.sum_details:
        pass
    else:
        line.set_value(value, "uploaded projections", override=True)
        line.set_hardcoded(True)

        # If the line is a hardcoded value, set a direct reference in excel
        sheet = cell.parent
        full_address = "'" + sheet.title + "'!" + cell.coordinate
        line.xl_data.set_ref_direct_source(full_address)
