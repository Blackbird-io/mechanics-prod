#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2015
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: data_structures.serializers.eggcellent.bb_workbook
"""

Module defines workbook with custom native-Python data storage on each sheet.
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
BB_Workbook           workbook where each sheet has a SheetData record set
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .data_management import SheetData



# Constants
# n/a

# Module Globals
# n/a

# Classes
class BB_Workbook(excel_interface.Workbook):
    """

    Class modifies standard workbook to include a SheetData record set on each
    sheet. As a result, we can do data lookups and reference builds faster and
    more explicitly.     
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    n/a

    FUNCTIONS:
    create_sheet()        returns sheet with a SheetData instance at sheet.bb
    ====================  ======================================================
    """ 
    def __init__(self, *pargs, **kwargs):
        excel_interface.Workbook.__init__(self, *pargs, **kwargs)

    def create_sheet(self, name):
        """


        BB_WorkBook.create_sheet(name) -> Worksheet


        Return worksheet with a SheetData record set at instance.bb.
        """
        sheet = excel_interface.Workbook.create_sheet(self, name)
        sheet.bb = SheetData()

        return sheet

        
