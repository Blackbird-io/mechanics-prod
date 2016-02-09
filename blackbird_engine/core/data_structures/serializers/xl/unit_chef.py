# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.xl.unit_chef
"""

Module defines a class that represents arbitrarily rich BusinessUnit instances
as a collection of linked Excel worksheets. 
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
UnitChef              chop BusinessUnit into dynamic Excel structure
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .data_types import TypeCodes, NumberFormats
from .field_names import FieldNames
from .formulas import FormulaTemplates
from .tab_names import TabNames

from .line_chef import LineChef

from ...modelling import common_events




# Constants
# n/a

# Module Globals
_INVALID_CHARS = r"\/*[]:?"
# Excel forbids the use of these in sheet names.
REPLACEMENT_CHAR = None
# When the replacement character is None, UnitChef will remove all bad chars
# from sheet titles.

bad_char_table = {ord(c):REPLACEMENT_CHAR for c in _INVALID_CHARS}
# May be this should be on the UnitChef class itself

field_names = FieldNames()
formula_templates = FormulaTemplates()
number_formats = NumberFormats()
tab_names = TabNames()
type_codes = TypeCodes()

get_column_letter = excel_interface.utils.get_column_letter

line_chef = LineChef()

# Classes
class UnitChef:
    """

    [Add doc string ]
    One tab per unit
    Children first
    Arbitrarily recursive (though should max out at sheet limit; alternatively,
    book should prohibit new sheets after that? or can have 2 limits: a soft
    limit where ModelChopper shifts into different representation mode, and a
    hard limit, where you just cant create any more sheets.)

    Most non-public methods force keyword-based arg entry to avoid potentially
    confusing erros (switching rows for columns, etc.)

    Methods generally leave current row pointing to their last completed (filled)
    row. 
    
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    n/a

    FUNCTIONS:
    chop_unit()           returns sheet with a SheetData instance at sheet.bb
    ====================  ======================================================
    """
    MAX_TITLE_CHARACTERS = 30
    SHOW_GRID_LINES = False
    ZOOM_SCALE = 80
    
    def chop_unit(self, *pargs, book, unit):
        """


        -> Worksheet


        Children should be spread before parent

        # Single-period only for now
        """
        # First, chop up the kids so we can link to their results.
        before_kids = len(book.worksheets)
        children = unit.components.get_ordered()
        # Spread children in order to ensure stable results across run times.
        
        for child in children:           
            self.chop_unit(book=book, unit=child)

        # Second, chop the parent
        sheet = self._create_unit_sheet(book=book, unit=unit, index=before_kids)        
        sheet = self._add_unit_life(sheet=sheet, unit=unit)
        
        current = sheet.bb.time_line.columns.get_position(unit.period.end)
        self._add_financials(sheet=sheet, unit=unit, column=current)

        # Third, return result
        return sheet


    def chop_multi(self, *pargs, book, unit):
        """

        Breadth-first analysis helps with vertical alignment at the cost
        of some performance (go through loop 3x instead of once, but do
        less at each step). 
        """

        # 1.   Chop the children
        before_kids = len(book.worksheets)
        children = unit.components.get_ordered()

        for child in children:
            self.chop_unit(book=book, unit=child)


        # 2.   Chop the parent
        # 2.1.   set up the unit sheet and spread params
        sheet = self._create_unit_sheet(book=book, unit=unit, index=before_kids)
##        for future_snapshot in unit:
##            self._add_unit_params(set_labels=False)
        

        # 2.2.   spread life
        sheet = self._add_unit_life(sheet=sheet, unit=unit)
        for future_snapshot in unit:
            
            sheet.bb.current_row = sheet.bb.parameters.ending
            # Have to reindex row up for every period, otherwise sheet will look
            # like a staircase.
            self._add_unit_life(set_labels=False)
        

        # 3.3.  spread fins
        current = sheet.bb.time_line.columns.get_position(unit.period.end)
        self._add_financials(sheet=sheet, unit=unit, column=current)

        for future_snapshot in unit:

            sheet.bb.current_row = sheet.bb.events.ending
            column = sheet.bb.time_line.columns.get_position(snapshot.period.end)
            # Load balance from prior column
            self._add_financials(sheet=sheet, unit=unit, column=column, set_labels=False)

            # Should make sure rows align here from one period to the next.
            # Main problem lies in consolidation logic.

        return sheet
    
    def _add_financials(self, *pargs, sheet, unit, column, set_labels=True):
        """

        -> Worksheet

        """
        unit.fill_out()
        # Make sure the unit contains all relevant calculations by filling it
        # out. If BB already performed this action, call will be a no-op.
        
        for statement in unit.financials.ordered:

            if statement:
            
                line_chef.chop_statement(
                    sheet=sheet,
                    statement=statement,
                    column=column,
                    set_labels=set_labels
                    )

    # Have to manage book depth (ie max sheets) #<--------------------------------------------------!!

    def _add_param_rows(self, sheet, params, active_column, label_column=None, master_column=None):
        """


        -> Worksheet


        "active_column" : integer, 1-based column index
        "label_column" : integer,
        "master_column" : integer

        if label_column or master_column are left blank, method will find
        indeces using .get_position(). to improve performance in loops, caller
        can supply these indeces as arguments.
        """

        # The parameter is specific to the period; we don't have it
        # on the page yet. Add a row and write the value there.

        parameters = sheet.bb.parameters

        if label_column is None:
            label_column = sheet.bb.parameters.columns.get_position(field_names.LABELS)
        if master_column is None:
            master_column = sheet.bb.parameters.columns.get_position(field_names.MASTER)

        new_row = parameters.rows.ending + 1
        # TO DO: Could also use Workbook.max_row()
        
        for param_name in sorted(params):

            param_value = params[param_name]
            
            # Register the new row
            parameters.rows.by_name[param_name] = new_row
            
            # Add the label
            label_cell = sheet.cell(column=label_column, row=new_row)
            label_cell.value = param_name

            # Add the master value (from this period)
            master_cell = sheet.cell(column=master_column, row=new_row)
            master_cell.value = param_value

            # Link the period to the master
            param_cell = sheet.cell(column=active_column, row=new_row)
            link = formula_templates.ADD_COORDINATES
            link = link.format(coordinates=master_cell.coordinate)
            param_cell.set_explicit_value(link, data_type=type_codes.FORMULA)

            sheet.bb.current_row = new_row
            new_row +=1
            
        return sheet

    def _add_items_to_area(self, *pargs, sheet, area, items, active_column, set_labels=True):
        """

        -> Worksheet

        Adds names in sorted order
        Expects sheet to come with parameters unless you specify the label and master column
        """

        parameters = sheet.bb.parameters
        label_column = parameters.columns.get_position(field_names.LABELS)
        master_column = parameters.columns.get_position(field_names.MASTER)
            # Master and labels always align to parameters.
            # May be these should be in the general area?
            # Upgrade-S: Can pass col positions in for speed; downside is that
            #            signature gets ugly. Can move to kwargs to protect. 

        new_row = sheet.bb.current_row + 1
        # TO DO: Could also use Workbook.max_row()
        
        for name in sorted(items):

            value = items[name]
            
            # Register the new row
            area.rows.by_name[name] = new_row
            
            # Add the label
            if set_labels:
                label_cell = sheet.cell(column=label_column, row=new_row)
                label_cell.value = name
                # Or run through labels routine

            # Add the master value (from this period)
            master_cell = sheet.cell(column=master_column, row=new_row)
            master_cell.value = value

            # Link the period to the master
            current_cell = sheet.cell(column=active_column, row=new_row)
            link = formula_templates.ADD_COORDINATES
            link = link.format(coordinates=master_cell.coordinate)
            current_cell.set_explicit_value(link, data_type=type_codes.FORMULA)

            sheet.bb.current_row = new_row
            new_row +=1
            
        return sheet
        
    def _add_unit_life(self, *pargs, sheet, unit, column=None, set_labels=True):
        """

        -> Worksheet
        
        Expects to get sheet with current row pointing to a blank
        Will start writing on current row

        """
        sheet_data = sheet.bb
        active_column = column

        if not active_column:
            active_column = sheet_data.time_line.columns.get_position(unit.period.end)
            
        if not getattr(sheet_data, "life", None):
            sheet.bb.add_area("life")

        if not getattr(sheet_data, "events", None):
            sheet.bb.add_area("events")
        
        first_life_row = sheet.bb.current_row + 2
        first_event_row = first_life_row + 9
        # Leave nine rows for basic life layout

        sheet.bb.current_row = first_event_row
        sheet = self._add_life_events(
            sheet=sheet,
            unit=unit,
            active_column=active_column,
            set_labels=set_labels
            )
        # For events, can keep link if they are the same as the master value,
        # or write the specified value if its different.

        sheet.bb.current_row = first_life_row
        sheet = self._add_life_analysis(
            sheet=sheet,
            unit=unit,
            active_column=active_column,
            set_labels=set_labels
            )

        sheet.bb.current_row = sheet.bb.events.rows.ending
        # Move current row down to the bottom (max_row() probably best here). 
        return sheet

        # To Do:
        # - for multiperiod generally, want to have period n+1 inherit life and
        #   params from period n if they are the same. so establish links.

##    def _add_life_events_old(self, *pargs, sheet, unit, active_column):
##        """
##
##
##        -> Worksheet
##
##
##        Expects to get sheet with current row pointing to first place you want to write
##        Returns sheet with current row pointing to last filled event
##        """
##        active_row = sheet.bb.current_row
##        parameters = sheet.bb.parameters
##        
##        label_column = parameters.columns.get_position(field_names.LABELS)
##        master_column = parameters.columns.get_position(field_names.MASTER)
##
##        events = sheet.bb.events
##
##        # For natural presentation order, sort events by date
##        sorted_events = sorted(unit.life.events.items(), key=lambda x:x[1])
##
##        for event_name, event_date in sorted_events:
##            events.rows.by_name[event_name] = active_row
##
##            label_cell = sheet.cell(column=label_column, row=active_row)
##            label_cell.value = event_name
##
##            master_cell = sheet.cell(column=master_column, row=active_row)
##            master_cell.value = event_date
##
##            event_cell = sheet.cell(column=active_column, row=active_row)
##            link_template = formula_templates.LINK_TO_COORDINATES
##            link = link_template.format(coordinates=master_cell.coordinate)
##            
##            event_cell.set_explicit_value(link, data_type=type_codes.FORMULA)
##            event_cell.number_format = master_cell.number_format
##
##            sheet.bb.current_row = active_row
##            active_row += 1
##
##        return sheet

    def _add_life_events(self, *pargs, sheet, unit, active_column, set_labels=True):
        """

        -> Worksheet

        Runs through 
        """
        active_row = sheet.bb.current_row
        events = sheet.bb.events

        existing_names = unit.life.events.keys() & events.rows.by_name.keys()
        new_names = unit.life.events.keys() - existing_names

        # Write values for existing events
        for name in existing_names:

            existing_row = events.rows.get_position(name)
            cell = sheet.cell(column=active_column, row=existing_row)

            cell.value = unit.life.events[name]
            # Upgrade: link to master value if it's the same as our live one

        # Now add 
        new_events = dict()
        for name in new_names:
            new_events[name] = unit.life.events[name]

        self._add_items_to_area(
            sheet=sheet,
            area=events,
            items=new_events,
            active_column=active_column,
            set_labels=set_labels
            )
        # _add_items() will update current row to the last filled position
        
        return sheet

    def _add_life_analysis(self, sheet, unit, active_column, set_labels=True):
        """

        -> Worksheet

        Add unit life for a single period

        Assumes events are filled out
        """

        active_row=sheet.bb.current_row + 1 #<------------- Think about whether this belongs
        
        label_column = sheet.bb.parameters.columns.get_position(field_names.LABELS)
        time_line_row = sheet.bb.time_line.rows.get_position(field_names.LABELS)
        
        fs = formula_templates
        set_label = line_chef._set_label

        events = sheet.bb.events
        life = sheet.bb.life

        birth = sheet.cell(
            column=active_column,
            row=events.rows.get_position(common_events.KEY_BIRTH)
            )
        death = sheet.cell(
            column=active_column,
            row=events.rows.get_position(common_events.KEY_DEATH)
            )
        conception = sheet.cell(
            column=active_column,
            row=events.rows.get_position(common_events.KEY_CONCEPTION)
            )

        cells = dict()
        cells["birth"] = birth
        cells["death"] = death
        cells["conception"] = conception       
        
        # 1. Add ref_date
        sheet.bb.life.rows.by_name[field_names.REF_DATE]=active_row
        if set_labels:
            set_label(
                label=field_names.REF_DATE,
                sheet=sheet,
                row=active_row,
                column=label_column
                )        

        ref_date = sheet.cell(column=active_column, row=active_row)
        
        time_line = sheet.cell(column=active_column, row=time_line_row)

        cells["ref_date"] = ref_date
        formula = fs.LINK_TO_COORDINATES.format(
            coordinates=time_line.coordinate
            )

        ref_date.value = formula
        ref_date.number_format = number_formats.DATETIME
        del formula
        # Make sure each cell gets its own formula by deleting F after use.
        

        # Move down two rows (to leave one blank)
        active_row += 2
        

        # 2. Add age
        sheet.bb.life.rows.by_name[field_names.AGE] = active_row
        if set_labels:
            set_label(
                label=field_names.AGE,
                sheet=sheet,
                row=active_row,
                column=label_column
                )

        age = sheet.cell(column=active_column, row=active_row)

        cells["age"] = age
        cos = {k:v.coordinate for k,v in cells.items()}
        formula = fs.COMPUTE_AGE_IN_DAYS.format(**cos)

        age.set_explicit_value(formula, data_type=type_codes.FORMULA)
        del formula


        # Move row down 
        active_row += 1
        

        # 3. Add alive
        life.rows.by_name[field_names.ALIVE]=active_row
        if set_labels:
            set_label(
                label=field_names.ALIVE,
                sheet=sheet,
                row=active_row,
                column=label_column
                )
        
        alive = sheet.cell(column=active_column, row=active_row)

        cells["alive"] = alive
        cos["alive"] = alive.coordinate
        formula = fs.IS_ALIVE.format(**cos)
        
        alive.set_explicit_value(formula, data_type=type_codes.FORMULA)
        del formula


        # Move row down
        active_row += 1


        # 4. Add span (so we can use it as the denominator in our percent
        #    computation below). 
        life.rows.by_name[field_names.SPAN] = active_row
        if set_labels:
            set_label(
                label=field_names.SPAN,
                sheet=sheet,
                row=active_row,
                column=label_column
                )

        span = sheet.cell(column=active_column, row=active_row)

        cells[field_names.SPAN] = span
        cos[field_names.SPAN] = span.coordinate
        formula = fs.COMPUTE_SPAN_IN_DAYS.format(**cos)

        span.set_explicit_value(formula, data_type=type_codes.FORMULA)
        del formula


        # Move row down
        active_row += 1
        sheet.bb.current_row = active_row


        # 5. Add percent
        life.rows.by_name[field_names.PERCENT] = active_row
        if set_labels:
            set_label(
                label=field_names.PERCENT,
                sheet=sheet,
                row=active_row,
                column=label_column
                )

        percent = sheet.cell(column=active_column, row=active_row)

        formula = fs.COMPUTE_AGE_IN_PERCENT.format(**cos)

        percent.set_explicit_value(formula, data_type=type_codes.FORMULA)

        # Return sheet
        return sheet

        # To Do:
        # - should make this multiperiod:
        #   -- cover multiple columns in range
        #   -- pull cells out first
        #   -- add labels once if at all
        
    def _add_unit_params(self, sheet, unit):
        """


        -> Worksheet
        

        Adds unit parameters for a single period
        Returns sheet with current row pointing to final param row
        """
        parameters = sheet.bb.parameters
        
        period_column = sheet.bb.time_line.columns.get_position(unit.period.end)
        label_column = parameters.columns.get_position(field_names.LABELS)

        existing_param_names = unit.parameters.keys() & parameters.rows.by_name.keys()
        new_param_names = unit.parameters.keys() - existing_param_names
        
        for param_name in existing_param_names:
            param_value = unit.parameters[param_name]

            existing_row = sheet.bb.parameters.rows.get_position(param_name)
            data_cell = sheet.cell(column=period_column, row=existing_row)
            data_cell.value = param_value
            # should format the cell in blue or whatever for hardcoded

        new_params = dict()
        for k in new_param_names:
            new_params[k] = unit.parameters[k]

        self._add_param_rows(sheet, new_params, period_column)

        sheet.bb.current_row = parameters.rows.ending
        return sheet    
        
        # To Do:
        # - add formatting

    def _create_unit_sheet(self, *pargs, book, unit, index):
        """


        -> Worksheet


        Returns sheet with current row pointing to last parameter row        
        """

        name = unit.name
        if name in book:
            rev_name = name + " ..." + str(unit.id.bbid)[-8: ]
            name = rev_name

            if name in book:
                name = str(unit.id.bbid)

        name = name.translate(bad_char_table)
        name = name[:self.MAX_TITLE_CHARACTERS]
        # Replace forbidden characters, make sure name is within length limits
        
        sheet = book.create_sheet(name, index)
        try:
            sheet.sheet_view.showGridLines = self.SHOW_GRID_LINES
            sheet.sheet_view.zoomScale = self.ZOOM_SCALE

            # Try semi-experimental formatting. Unclear whether openpyxl fully
            # supports these features.
            #
            # Re gred, see https://bitbucket.org/openpyxl/openpyxl/issues/199
            # Re zoom, see https://bitbucket.org/openpyxl/openpyxl/issues/262
            
        except Exception:
            pass      
        
        unit.xl.set_sheet(sheet)

        # Should auto-hide components below level 2 or smtg to make the book cleaner; sheet_state="hidden"
        # Could have a relationships.depth number on each unit. Then if relationships.depth >= x, you
        # do something. Ask Erika to start working on refactoring .relationships out of Tags. 

        self._link_to_time_line(book=book, sheet=sheet)
        self._add_unit_params(sheet, unit)
        # At this point, sheet.bb.current_row will point to the last parameter.

        # Freeze panes:
        corner_row = sheet.bb.time_line.rows.ending
        corner_row +=1
        
        corner_column = sheet.bb.parameters.columns.get_position(field_names.MASTER)
        corner_column +=1
        
        corner_cell = sheet.cell(column=corner_column, row=corner_row)
        sheet.freeze_panes = corner_cell

        # Return sheet
        return sheet

        # To Do:
        # - add naming with better semantic relationships. For example, could specify name and
        #   name of parent object, or something like that. May be add the type if its specified too.
        

    def _link_to_area(self, source_sheet, local_sheet, area_name, group=False, keep_format=True):
        """


        -> Worksheet


        "area_name" : string 
        """
        source_area = getattr(source_sheet.bb, area_name)
        local_area = getattr(local_sheet.bb, area_name, None)

        if local_area is None:
            local_area = local_sheet.bb.add_area(area_name)

        local_area.update(source_area)
        coordinates = {"sheet" : source_sheet.title}
        
        for row in source_area.rows.by_name.values():

            source_row = (source_area.rows.starting or 0) + row
            local_row = (local_area.rows.starting or 0) + row

            for column in source_area.columns.by_name.values():

                source_column = (source_area.columns.starting or 0) + column
                local_column = (local_area.columns.starting or 0) + column

                local_cell = local_sheet.cell(column=local_column, row=local_row)

                cos = coordinates.copy()
                cos["row"] = source_row
                cos["alpha_column"] = get_column_letter(source_column)

                link = formula_templates.LINK_TO_CELL_ON_SHEET.format(**cos)
                local_cell.set_explicit_value(link, data_type=type_codes.FORMULA)

                if keep_format:
                    source_cell = source_sheet.cell(column=source_column, row=source_row)
                    local_cell.number_format = source_cell.number_format

            local_sheet.bb.current_row = local_row
        
        # if group:
        #   ##group cells

        return local_sheet

        # To do:
        # - if group: group the range
        # - copy styles? the result should inherit styling from source

    def _link_to_time_line(self, *pargs, book, sheet):
        """


        -> Worksheet
        

        Link sheet to book's time_line
        Force keyword-entry for book and sheet to make sure we feed in the right
        arguments. 
        """
        source = book.get_sheet_by_name(tab_names.TIME_LINE)
        
        self._link_to_area(source, sheet, "time_line")
        self._link_to_area(source, sheet, "parameters")

    
