# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.xl.line_chef
"""

Module defines a class that represents arbitrarily rich BusinessUnit instances
as a collection of linked Excel worksheets. 
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
LineChef              chop BB statements into dynamic Excel structures
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .data_types import TypeCodes
from .field_names import FieldNames
from .formulas import FormulaTemplates

from ...modelling.line_item import LineItem




# Constants
# n/a

# Module Globals
field_names = FieldNames()
formula_templates = FormulaTemplates()
type_codes = TypeCodes()

get_column_letter = excel_interface.utils.get_column_letter

# Classes
class LineChef:
    """

    Objects that export models into Excel

    """
    def chop_line(self, *pargs, sheet, line, set_labels=True, indent=0):
        """
        by the end of this routine, the line and all its details should have a
        sheet assignment. that way, we can gather 

        Expects sheet to have current_column and current row SET UP

        Routines deliver sheet with the current_row pointing to the last filled in cell.

        Direct cell references make a lot of this logic way more transparent, at the expense
        of making it a lot more dependent on the nature of the xlio Cell object. 
        """
        self._add_consolidation_logic(
            sheet=sheet,
            line=line,
            set_labels=set_labels,
            indent=indent
            )

        self._add_derivation_logic(
            sheet=sheet,
            line=line,
            set_labels=set_labels,
            indent=indent
            )

        details = line.get_ordered()
        if details:

            sheet.bb.current_row += 1
            
            sub_indent = indent + LineItem.TAB_WIDTH
            detail_summation = ""

            for detail in details:

                sheet.bb.current_row += 1
                
                self.chop_line(
                    sheet=sheet,
                    line=detail,
                    set_labels=set_labels,
                    indent=sub_indent
                    )

                link_template = formula_templates.ADD_COORDINATES
                link = link_template.format(coordinates=detail.xl.get_coordinates())
                detail_summation += link
                
            else:
                # Should group all the details here 
                sheet.bb.current_row +=1
                
                subtotal_cell = sheet.cell(column=sheet.bb.current_column, row=sheet.bb.current_row)
                subtotal_cell.set_explicit_value(detail_summation, data_type=type_codes.FORMULA)

                line.xl.detailed.ending = sheet.bb.current_row
                line.xl.detailed.cell = subtotal_cell
                
                if set_labels:
                    label = line.name + ": details"
                    self._set_label(
                        sheet=sheet,
                        label=label,
                        row=sheet.bb.current_row
                        )

        self._combine_segments(sheet=sheet, line=line, set_labels=set_labels)
        # Could also group here

        return sheet

    
        # If I toggle this logic a bit, I should be able to put all of the consolidation
        # cells in one part of the sheet, and then the derivation logic elsewhere.

        # This relies on:
            # line.increment() tracking sources
            # driver.workOnThis() collating data #<------------------------------------------------------------------------------------------------------CHECK WHAT THIS DOES!
        
    # formula.to_excel()-> string, dict;
        # should move formulas to return row data, so the xl formula is less
        # complicated

    # could add to verification protocol: a map of coordinates to compute values!
    # have to run verification later, because formulas NEVER eval in memory. You actually
    # need to save the excel to get the result. Can put these things on the sheet. keyed by coordinates.
    # coordinate object should hash into (x,y,z)

    # Need to add basic / hardcoded value logic: <------------------------------------------------------------------------------------------------------------MUST!
    # if not derived and not consolidated:
    #  add_hard_coded() # but have to watch out for details?
    #
    # can be part of combine_segments()

    # To Do:
    # - group all details
    # -

    def chop_statement(self, *pargs, sheet, statement):

        # Relies on current column!

        # Add a nice header row
        # Add some white space

        for line in statement.get_ordered():

            sheet.bb.current_row += 1
            self.chop_line(sheet=sheet, line=line)
            
        return sheet

    def _rows_to_coordinates(self, *pargs, lookup, column):
        result = dict()
        alpha_column = get_column_letter(column)
        
        for k in lookup.by_name:
            row = lookup.get_position(k)
            result[k] = alpha_column + str(row)

        return result

    def _add_consolidation_logic(self, *pargs, sheet, line, set_labels=True):
        """

        -> Worksheet
        
        Expects line to include a full range of pointers (?) or formulas?
        
        """
        # Add feature to control how many refs per cell (children_per_cell)

        if not line.xl.consolidated.sources:
            pass

        else:
            sheet.bb.current_row += 2
            line.xl.consolidated.starts = sheet.bb.current_row

            for source_pointer in line.xl.consolidated.sources:
                
                link_cell = sheet.cell(sheet.bb.current_column, sheet.bb.current_row)
                source_coordinates = source.xl.get_coordinates() #<------------------------------this won't work unless i add a .cell attribute to the line
                # which may be i should. then can return .cell.parent.title + "!" + .cell.coordinate
                
                link_formula = self.formulas.LINK_TO_COORDINATES.format(coordinates=source_coordinates)
                link_cell.value = link_formula

                line.xl.consolidated.ends = sheet.current_row
                sheet.current_row += 1

            # Group the cells!! <--------------------------------------------------------------------------

            summation_params = {
                "starting_row" : line.xl.consolidated.starts, #<-------------------------------------------------- CHECK WHETHER THIS PIECE IS VALID
                "ending_row" : line.xl.consolidated.ends,
                "alpha_column" : sheet.bb.current_column
                }
            #<------------------------------------------------------------------------------------------------------------------------need to make sure column is alphabetical
            
            summation_formula = self.formulas.SUM_RANGE.format(**summation_params)
            summation_cell = sheet.cell(sheet.current_column, sheet.current_row)
            summation_cell.value = summation_formula

            if set_labels:
                # Add the "x : consolidated value" label #<---------------------------------------------------------------------------------------fix
                pass

            line.xl.consolidated.ending = sheet.bb.current_row

        return sheet

        # To Do:
        # - group the cells
        # 
    
    def _add_derivation_logic(self, *pargs, sheet, line, set_labels=True, indent=0):
        """

        -> Worksheet

        Expects sheet current row to point to last filled cell. 
        
        """
        if not line.xl.derived.calculations:
            pass

        else:
            for data_cluster in line.xl.derived.calculations:

                sheet.bb.current_row += 2
                # Leave a blank row between each calculation
                
                self._add_driver_calculation(
                    sheet=sheet,
                    line=line,
                    driver_data=data_cluster, 
                    set_labels=set_labels,
                    indent=indent
                    )

            else:
                pass

                # NOTE: No summation at the end of the derive process. Final
                # derived value may OR MAY NOT be the sum of priors. Up to
                # driver to decide.

        return sheet

    def _add_driver_calculation(self, *pargs, sheet, line, driver_data, set_labels=True, indent=0):
        """


        -> Worksheet

        Will write to sheet current row
        """
        private_data = sheet.bb.parameters.copy()
        # Set up a private range that's going to include both "shared" period & unit
        # parameters from the column and "private" driver parameters. 

        label_column = sheet.bb.parameters.columns.get_position(field_names.LABELS)
        period_column = sheet.bb.current_column #<---------------------------------------------------- NEED EXPLICIT COLUMN PASSING
        
        for row_data in driver_data.rows:

            private_label = row_data[field_names.LABELS]
            private_value = row_data[field_names.VALUES]
            
            if private_label and set_labels:
                self._set_label(
                    sheet=sheet,
                    label=private_label,
                    row=sheet.bb.current_row,
                    column=label_column
                    )

            param_cell = sheet.cell(column=period_column, row=sheet.bb.current_row)
            param_cell.value = private_value
            # Assume that private_value is a number or string, NOT a formula or
            # container. Can change this later. 

            relative_position = sheet.bb.current_row - (private_data.rows.starting or 0)
            private_data.rows.by_name[private_label] = relative_position
            # ... In this particular case, we could map a specific cell
            # (in memory) to the parameter. Unclear whether that's useful though,
            # because we generally look up locations for lines, not parameters.
            # And lines continue to span several rows. 

            sheet.bb.current_row += 1
            # Will add a blank row after all the columns

        # Transform the range values from rows to coordinates
        param_coordinates = self._rows_to_coordinates(
            lookup=private_data.rows,
            column=sheet.bb.current_column
            )
            
        # Apply param:var conversions so formula can find its expected inputs
        for param_name, var_name in driver_data.conversion_map.items():
            param_coordinates[var_name] = param_coordinates[param_name]

        # Finally, format the formula as necessary
        # (if references are a dict of objects, could map each obj to its coordinates)
        template = driver_data.formula
        
        line_coordinates = dict()
        for k, obj in driver_data.references.items():
            line_coordinates[k] = obj.xl.get_coordinates()
            
        materials = dict()
        materials["lines"] = line_coordinates
        materials["parameters"] = param_coordinates

        life_coordinates = self._rows_to_coordinates(
            lookup=sheet.bb.life.rows,
            column=sheet.bb.current_column
            )
        materials["life"] = life_coordinates

        event_coordinates = self._rows_to_coordinates(
            lookup=sheet.bb.events.rows,
            column=sheet.bb.current_column
            )
        materials["events"] = event_coordinates
        
        formula = template.format(**materials)
        # Formulas should deliver templates with the {lines} key.
        
        calc_cell = sheet.cell(column=sheet.bb.current_column, row=sheet.bb.current_row)
        calc_cell.set_explicit_value(formula, data_type=type_codes.FORMULA)

        # If formula included a reference to the prior value of the line itself, it
        # picked up here. Can now change line.xl.derived.final <--------------------------------------------incrementation is discouraged, right?
        line.xl.derived.ending = sheet.bb.current_row
        line.xl.derived.cell = calc_cell

        if set_label:
            label = (indent * " ") + driver_data.name
            self._set_label(sheet=sheet, label=label, row=sheet.bb.current_row)

        return sheet
        #<--------------------------------------------------------------------------------------formula should be able to deliver multiple rows
        #that build on each other (for complex calculations): could have something like last_step as the variable in params that we keep
        # changing

        #<-------- should group each calculation like this
        # TO DO:
        # -- group each calculation
        # -- add a nice header for the calculation (may be?)
        # 

    def _combine_segments(self, *pargs, sheet, line, set_labels=True):
        """


        LineChef._combine_segments() -> Worksheet


        Adds the combination to the current row.
        """

        ends = [
            line.xl.consolidated.ending,
            line.xl.derived.ending,
            line.xl.detailed.ending
                     ]

        segment_summation = self._sum_endpoints(rows=ends, column=sheet.bb.current_column)
        if segment_summation:

            sheet.bb.current_row += 2
            
            total_cell = sheet.cell(column=sheet.bb.current_column, row=sheet.bb.current_row)
            total_cell.set_explicit_value(segment_summation, data_type=type_codes.FORMULA)

            line.xl.ending = sheet.bb.current_row
            line.xl.cell = total_cell

            if set_labels:
                label = indent*" " + LineItem.SUMMARY_PREFIX + line.name
                self._set_label(label=label, sheet=sheet, row=sheet.bb.current_row)
            
        return sheet

    def _set_label(self, *pargs, label, sheet, row, column=None, overwrite=False):
        """


        LineChef._set_label() -> Worksheet


        Set (column, row) cell value to label. Throw exception if cell already
        has a different label, unless ``overwrite`` is True.

        If ``column`` is None, method attempts to locate the labels column in
        the sheet.bb.parameters area.
        """
        if column is None:
            column=sheet.bb.parameters.columns.get_position(field_names.LABELS)

        label_cell = sheet.cell(column=column, row=row)
        existing_label = label_cell.value

        if overwrite or not existing_label:
            label_cell.value = label
        else:
            if existing_label != label:
                        
                c = """
                Something is wrong with our alignment. We are trying to
                write a parameter to an existing row with a different label.""" 

                raise Error(c)
                    
                # Check to make sure we are writing to the right row; if the label 
                # doesn't match, we are in trouble.

        return sheet

    def _sum_endpoints(self, *pargs, rows, column):
        """


        LineChef._sum_endpoints() -> string
        

        Return a summation of each of the rows in column. Expects rows to be a
        collection of absolute row indeces.
        """
        summation = ""

        coordinates = dict()
        coordinates["alpha_column"] = get_column_letter(column)

        for row in rows:
            if row is not None:
                coordinates["row"] = row
                link = formula_templates.ADD_CELL.format(**coordinates)
                summation += link

        return summation
                

    

    



    
               
        
        

        
        
    


    
    
    
        
        
    
        
        

        

    
    
            

            

            

            


            

            

            
            

                

                
            

            

                
        

    
