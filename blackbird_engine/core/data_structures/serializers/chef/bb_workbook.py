#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2016
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: data_structures.serializers.eggcellent.bb_workbook
"""

Module defines workbook with custom native-Python data storage on each sheet.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
BB_Workbook           workbook where each sheet has a SheetData record set
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from ._chef_tools import test_book
from .data_management import SheetData




# Constants
# n/a

# Module Globals
# n/a

# Classes
class BB_Workbook(xlio.Workbook):
    """

    Class modifies standard workbook to include a SheetData record set on each
    sheet. As a result, we can do data lookups and reference builds faster and
    more explicitly.     
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    n/a

    FUNCTIONS:
    create_sheet()        returns sheet with a SheetData instance at sheet.bb
    save()                saves workbook to file and runs test on contents
    ====================  ======================================================
    """ 
    def __init__(self, *pargs, **kwargs):
        xlio.Workbook.__init__(self, *pargs, **kwargs)

    def create_sheet(self, name, index=None):
        """


        BB_WorkBook.create_sheet(name) -> Worksheet

        --``name`` must be string name for new worksheet
        --``index`` is the desired index at which to place the new worksheet

        Return worksheet with a SheetData record set at instance.bb.

        """
        sheet = xlio.Workbook.create_sheet(self, name, index=index)
        sheet.bb = SheetData()

        return sheet

    def save(self, filename):
        """


        BB_WorkBook.save(filename) -> None

        --``filename`` must be string path at which to save workbook

        Method saves workbook to specified file and initiates test of file
        contents.
        """
        # save workbook
        xlio.Workbook.save(self, filename)

        # test the workbook against engine values
        test_book(filename)
