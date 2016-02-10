#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2016
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: serializers.chef.model_chef
"""

Class for creating dynamic Excel representations of Blackbird Engine models.

Eggcellent modules write formulas to cells explicitly, using the .set_explicit_value()
method, to make sure Excel interprets the strings correctly. 
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
field_names           commonly used field names
formula_templates     string templates for commonly used formulas
tab_names             standard tab names

FUNCTIONS:
n/a

CLASSES:
Chef                  chop Blackbird Engine model into a dynamic Excel workbook
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .bb_workbook import BB_Workbook as Workbook

from .data_management import Area
from .field_names import FieldNames
from .formulas import FormulaTemplates
from .tab_names import TabNames
from .data_types import TypeCodes

from .unit_chef import UnitChef




# Constants
# n/a

# Module Globals
field_names = FieldNames()
formula_templates = FormulaTemplates()
tab_names = TabNames()
type_codes = TypeCodes()

unit_chef = UnitChef()

get_column_letter = excel_interface.utils.get_column_letter

# Classes
class ModelChef:
    """

    Class packages an Engine model into an Excel Workbook with dynamic links.
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    n/a

    FUNCTIONS:
    chop_model()          returns sheet with a SheetData instance at sheet.bb
    ====================  ======================================================
    """
    
    def chop_model(self, model):
        """

        -> Workbook

        """
        # self._import_historical_data(model)
        
        book = self._build_foundation(model)
        now = model.time_line.current_period
        company = now.content
        
        company_sheet = unit_chef.chop_multi(book=book, unit=company)

        # book = self._create_annual_summary_tab(model, book)
        # book = self._create_dcf_tab(model, book)
        # book = self._create_ev_tab(model, book)
                
        return book

    #*************************************************************************#
    #                          NON-PUBLIC METHODS                             #
    #*************************************************************************#

    def _build_foundation(self, model):
        """


        ModelChopper._build_foundation() -> Workbook
        

        Return a workbook with:
           cover [not implemented yet]
           scenarios
           timeline
           
        """       
        book = Workbook()
        
        # self._create_cover_tab(book, model)
        self._create_scenarios_tab(book, model)
        self._create_time_line_tab(book, model)

        return book

    def _create_scenarios_tab(self, book, model):
        """


        ModelChopper._create_scenarios_tab() -> Worksheet

        
        Return a worksheet that lays out the assumptions used by the model in
        various scenarios. 
        """
        my_tab = book.create_sheet(tab_names.SCENARIOS)

        # Sheet map:
        # A          |  B      | C             | D     | E
        # param names|  blank  | active values | blank | blackbird values

        starting_row = 1

        label_column = 1
        in_effect_column = 3
        base_case_column = 5

        area = my_tab.bb.general

        area.columns.by_name[field_names.LABELS] = label_column
        area.columns.by_name[field_names.VALUES] = in_effect_column
        area.columns.by_name[field_names.BASE_CASE] = base_case_column

        current_row = starting_row
        
        for param_name in sorted(model.time_line.parameters):
            # Sort to make sure we display the parameters in stable order,
            # otherwise order could vary from chop to chop on the same model.

            label_cell = my_tab.cell(column=label_column, row=current_row)
            area.rows.by_name[param_name] = current_row

            in_effect_cell = my_tab.cell(column=in_effect_column, row=current_row)
            base_case_cell = my_tab.cell(column=base_case_column, row=current_row)
                        
            label_cell.value = param_name
            base_case_cell.value = model.time_line.parameters[param_name]

            link = formula_templates.ADD_COORDINATES
            link = link.format(coordinates=base_case_cell.coordinate)
            in_effect_cell.set_explicit_value(link, data_type=type_codes.FORMULA)
            # Since our formulas start with a "+" instead of "=" to allow easy
            # nesting, we use the explicit call to tell Excel to read them as
            # formulas instead of strings

            current_row += 1
    
        return my_tab

        # TO DO:
        # - print every scenario in model.scenarios ("base", "bad", "good", etc.)
        # - add headers to the columns (to specify what they are)
        # - hide grid lines
        # - add formatting
        # - add the selector function for user input so can see own assumptions
        #   and bb side by side, with only the active choice feeding into
        #   in_effect column.
        # - potentially add a widget on each page that allows you to toggle the
        #   scenarios (so you can keep looking wherever you were looking and
        #   see how it plays out).
        
    def _create_time_line_tab(self, book, model):
        """


        ModelChopper._create_time_line_tab() -> Worksheet


        Add a sheet that spreads out the time periods and shows each period's
        parameters.

        By default, each period inherits all parameters from the active scenario
        (on the scenarios sheet) by default. If a period specifies its own value
        for a parameter, routine will overwrite the link with a hard-coded
        value. 
        
        Method expects book to include a completed scenarios sheet. 
        """        
        scenarios = book[tab_names.SCENARIOS]
        scenarios_area = scenarios.bb.general
        
        my_tab = book.create_sheet(tab_names.TIME_LINE)

        parameters = my_tab.bb.add_area("parameters")
        time_line = my_tab.bb.add_area("time_line")

        # Pick starting positions
        local_labels_column = 1
        local_master_column = 3

        alpha_master_column = get_column_letter(local_master_column)
        
        parameters.columns.by_name[field_names.LABELS] = local_labels_column
        parameters.columns.by_name[field_names.MASTER] = local_master_column
        
        header_row = 3
        time_line.rows.by_name[field_names.LABELS] = header_row
        my_tab.bb.current_row = header_row



        # First, pull the parameter names and active values from the scenarios
        # tab into a local "label" and "master" column, respectively.
        external_coordinates = dict()
        external_coordinates["sheet"] = scenarios.title

        source_label_column = scenarios_area.columns.get_position(field_names.LABELS)
        source_value_column = scenarios_area.columns.get_position(field_names.VALUES)

        for param_name in scenarios.bb.general.rows.by_name:
            # We can build the page in any order here

            row_number = scenarios.bb.general.rows.get_position(param_name)
            # Get the correct relative position
            
            source_coordinates = external_coordinates.copy()
            source_coordinates["row"] = row_number

            active_row = header_row + row_number 
            my_tab.bb.parameters.rows.by_name[param_name] = active_row
            # TO DO: Think about whether we should be keeping track of the row
            # number differently here.

            # Label cell should link to the parameter name on the scenarios
            # sheet
            label_cell = my_tab.cell(column=local_labels_column, row=active_row)
            
            cos = source_coordinates.copy()
            cos["alpha_column"] = get_column_letter(source_label_column)
            link = formula_templates.ADD_CELL_FROM_SHEET.format(**cos)
            label_cell.set_explicit_value(link, data_type=type_codes.FORMULA)

            # Master cell should link to the active value 
            master_cell = my_tab.cell(column=local_master_column, row=active_row)

            cos = source_coordinates.copy()
            cos["alpha_column"] = get_column_letter(source_value_column)
            link = formula_templates.ADD_CELL_FROM_SHEET.format(**cos)
            master_cell.set_explicit_value(link, data_type=type_codes.FORMULA)



        # Second, build a column for each period. Pull values from our local
        # master column and overwrite them if necessary with direct values.        
        
        starting_column = local_master_column + 2
        active_column = starting_column

        past, present, future = model.time_line.get_segments()
        covered_dates = present + future
        for end_date in covered_dates:

            period = model.time_line[end_date]

            my_tab.bb.time_line.columns.by_name[period.end] = active_column
            parameters.columns.by_name[period.end] = active_column
            # Need this to make sure the parameters Area looks as wide as the
            # timeline. Otherwise, other routines may think that the params area
            # is only one column wide. 
            
            header_cell = my_tab.cell(column=active_column, row=header_row)
            header_cell.value = period.end

            # 1. Pulling the master values for each parameter.

            for param_row in my_tab.bb.parameters.rows.by_name.values():
                # May write the column in undefined order

                param_cell = my_tab.cell(column=active_column, row=param_row)

                link_template = formula_templates.ADD_CELL
                cos = dict(alpha_column=alpha_master_column, row=param_row)
                link = link_template.format(**cos)
                
                param_cell.set_explicit_value(link, data_type=type_codes.FORMULA)

            # 2. Overwrite links with hard-coded values where the period
            #    specifies them. Add period-specific parameters.

            existing_param_names = period.parameters.keys() & parameters.rows.by_name.keys()
            new_param_names = period.parameters.keys() - existing_param_names
            # New parameters are specific to the period. We don't have a row for
            # them on the sheet yet, so we'll add them later.

            for spec_name in existing_param_names:                

                spec_value = period.parameters[spec_name]
                    
                param_row = parameters.rows.get_position(spec_name)                
                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = spec_value
##                spec_cell.format = blue_font_color

            new_params = dict()
            for k in new_param_names:
                new_params[k] = period.parameters[k]

            unit_chef.add_items_to_area(
                sheet=my_tab,
                area=my_tab.bb.parameters,
                items=new_params,
                active_column=active_column
                )
            # Upgrade-S: For speed, can supply master and label column indeces
            # to the add_items() routine.
            
            active_column += 1

        return my_tab
  
        # To do:
        # - Group
        # - Add formatting for hard-coded numbers (blue font)
        
        
