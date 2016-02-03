import openpyxl as excel_interface

class BB_Workbook(excel_interface.Workbook):
    def __init__(self, *pargs, **kwargs):

        excel_interface.Workbook.__init__(*pargs, **kwargs)

    def create_sheet(self, name):
        sheet = excel_interface.Workbook.create_sheet(self, name)
        sheet.bb = SheetData()

        return sheet

        
