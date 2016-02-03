#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2016
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: data_structures.serializers.eggcellent.model_chopper
"""

Class for creating dynamic Excel representations of Blackbird Engine models.
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
ModelChopper          chop Blackbird Engine model into a dynamic Excel workbook
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .bb_workbook import BB_Workbook as Workbook

from .field_names import FieldNames
from .formulas import FormulaTemplates
from .tab_names import TabNames

from .unit_chopper import UnitChopper




# Constants
# n/a

# Module Globals
get_column_letter = excel_interface.utils.get_column_letter

# Classes
class ModelChopper:
    """

    Class packages an Engine model into an Excel Workbook with dynamic links.
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    field_names           commonly used field names
    formula_templates     string templates for commonly used formulas
    tab_names             standard tab names

    FUNCTIONS:
    chop_model()          returns sheet with a SheetData instance at sheet.bb
    ====================  ======================================================
    """
    field_names = FieldNames()
    formula_templates = FormulaTemplates()
    tab_names = TabNames()
    
    def chop_model(self, model):
        """

        -> Workbook

        """
        # self._import_historical_data(model)
        
        book = self._spread_foundation(model)
        now = model.time_line.current_period
        company = now.content
        
        UnitChopper.chop_unit(company, book)
        # <-------------------------------------------------------------------------------- May need to specify period, column, etc.

        # book = self._create_annual_summary_tab(model, book)
        # book = self._create_dcf_tab(model, book)
        # book = self._create_ev_tab(model, book)
                
        return book

    #*************************************************************************#
    #                          NON-PUBLIC METHODS                             #
    #*************************************************************************#

    def _build_foundation(self, model):
        """


        ModelChopper._build_foundation() -> Workbook
        

        Return a workbook with:
           cover [not implemented yet]
           scenarios
           timeline
           
        """       
        book = Workbook()
        
        # self._create_cover_tab(model, book)
        self._create_scenarios_tab(book)
        self._create_time_line_tab(book)

        return book

    def _create_scenarios_tab(self, book, model):
        """


        ModelChopper._create_scenarios_tab() -> Worksheet

        
        Return a worksheet that lays out the assumptions used by the model in
        various scenarios. 
        """
        scenarios = book.create_sheet(self.tab_names.SCENARIOS)

        # Sheet map:
        # A          |  B      | C             | D     | E
        # param names|  blank  | active values | blank | blackbird values

        starting_row = 1

        label_column = 1
        in_effect_column = 3
        base_case_column = 5
        
        scenarios.bb.general.columns.by_name[self.field_names.LABELS] = label_column
        scenarios.bb.general.columns.by_name[self.field_names.VALUES] = in_effect_column
        scenarios.bb.general.columns.by_name[self.field_names.BASE_CASE] = base_case_column

        current_row = starting_row
        
        for param_name in sorted(model.time_line.parameters):
            # Sort to make sure we display the parameters in stable order,
            # otherwise order could vary from chop to chop on the same model.

            label_cell = scenarios.cell(column=label_column, row=current_row)
            scenarios.bb.general.rows.by_name[param_name] = current_row

            in_effect_cell = scenarios.cell(column=in_effect_column, row=current_row)
            base_case_cell = scenarios.cell(column=base_case_column, row=current_row)
                        
            label_cell.value = param_name
            base_case_cell.value = model.time_line.parameters[param_name]

            link = self.formulas.ADD_COORDINATES
            link = link.format(coordinates=base_case_cell.coordinate)
            in_effect_cell.value = link

            current_row += 1
        
        return scenarios

        # TO DO:
        # - print every scenario in model.scenarios ("base", "bad", "good", etc.)
        # - add headers to the columns (to specify what they are)
        # - hide grid lines
        # - add formatting
        # - add the selector function for user input so can see own assumptions
        #   and bb side by side, with only the active choice feeding into
        #   in_effect column.
        # - potentially add a widget on each page that allows you to toggle the
        #   scenarios (so you can keep looking wherever you were looking and
        #   see how it plays out).
        
    def _create_time_line_tab(self, book, model):
        """


        ModelChopper._create_time_line_tab() -> Worksheet


        Add a sheet that spreads out the time periods and shows each period's
        parameters.

        By default, each period inherits all parameters from the active scenario
        (on the scenarios sheet) by default. If a period specifies its own value
        for a parameter, routine will overwrite the link with a hard-coded
        value. 
        
        Method expects book to include a completed scenarios sheet. 
        """        
        scenarios = book[self.tab_names.SCENARIOS]
        
        my_tab = book.create_sheet(self.tab_names.TIME_LINE)
        parameters = Area("Parameters")
        parameters.parent = my_tab #<------------------------------------------------------------------------- this should be a sheet-level routine
        # Sheet.add_area("parameters") -> Area with name and parent relationship
        
        my_tab.bb.parameters = parameters

        # Pick starting positions
        local_labels_column = 1
        local_master_column = 3

        alpha_master_column = get_column_letter(local_master_column)
        
        parameters.columns.by_name[self.field_names.LABELS] = local_labels_column
        parameters.columns.by_name[self.field_names.MASTER] = local_master_column
        
        header_row = 3
        my_tab.bb.time_line.rows.by_name[self.field_names.LABELS] = header_row



        # First, pull the parameter names and active values from the scenarios
        # tab into a local "label" and "master" column, respectively.
        
        external_link = self.formula_templates.ADD_CELL_FROM_SHEET

        external_coordinates = dict()
        external_coordinates["sheet"] = scenarios.title

        source_label_column = scenarios.bb.general.columns.get_position(self.column_names.LABELS)
        source_value_column = scenarios.bb.general.columns.get_position(self.column_names.ACTIVE_SCENARIO)
        
        for param_name in scenarios.bb.general.rows.by_name:
            # We can build the page in any order here

            row_number = scenarios.bb.general.rows.get_position(param_name)
            # Get the correct relative position
            
            source_coordinates = external_coordinates.copy()
            source_coordinates["row"] = row_number

            active_row = header_row + row_number 
            my_tab.bb.parameters.rows.by_name[param_name] = active_row
            # TO DO: Think about whether we should be keeping the row number differently here.

            # Label cell should link to the parameter name on the scenarios sheet
            label_cell = my_tab.cell(column=label_column, row=active_row)
            
            cos = source_coordinates.copy()
            cos["column"] = source_label_column
            link = master_link.format(**cos)
            label_cell.value = link

            # Master cell should link to the active value 
            master_cell = my_tab.cell(column=local_master_column, row=active_row)

            cos = source_coordinates.copy()
            cos["column"] = source_value_column
            link = master_link.format(**cos)
            master_cell.value = link



        # Second, build a column for each period. Pull values from our local
        # master column and overwrite them if necessary with direct values.        
        
        starting_column = local_master_column + 2
        active_column = starting_column
        
        for period in model.time_line.get_ordered():

            my_tab.bb.time_line.columns.by_name[period.end] = active_column
            parameters.columns.by_name[period_end] = active_column
            # Need this to make sure the parameters Area looks as wide as the
            # timeline. Otherwise, other routines may think that the params area
            # is only one column wide. 
            
            header_cell = my_tab.cell(column=active_column, row=header_row)
            header_cell.value = period.end

            # 1. Pulling the master values for each parameter.

            for param_row in my_tab.bb.parameters.rows.by_name.values():
                # May write the column in undefined order

                param_cell = my_tab.cell(column=active_column, row=param_row)
                link = local_link.format(alpha_column=alpha_master_column, row=param_row)
                param_cell.value = link

            # 2. Overwrite links with hard-coded values where the period
            #    specifies them. Add period-specific parameters.

            existing_param_names = period.parameters.keys() & parameters.rows.by_name.keys()
            new_param_names = period.parameters.keys() - existing_param_names
            # New parameters are specific to the period. We don't have a row for
            # them on the sheet yet, so we'll add them later.

            for spec_name in existing_param_names:                

                spec_value = period.parameters[spec_name]
                    
                param_row = parameters.rows.get_position(spec_name)                
                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = spec_value
##                spec_cell.format = blue_font_color

            new_params = dict()
            for k in new_param_names:
                new_params[k] = period.parameters[k]

            UnitChopper._add_param_rows(my_tab, new_params, active_column,
                                        label_column=local_labels_column,
                                        master_column=local_master_column)
            # Supply column indeces for speed, otherwise routine would look
            # them up on every call.
            
            active_column += 1

        return my_tab
  
        # To do:
        # - Group
        # - Add formatting for hard-coded numbers (blue font)
        # - Improve efficiency by splitting period params into uniques and
        #   specifics first. That way, don't have to overwrite anything.

        
        
