#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2015
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: data_structures.eggscellent.bb_workbook
"""

Module defines workbook with custom native-Python data storage on each sheet.
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
BB_Workbook           workbook where each sheet has a SheetData record set
====================  ==========================================================
"""




# imports
import openpyxl as excel_interface




# classes
class BB_Workbook(excel_interface.Workbook):
    def __init__(self, *pargs, **kwargs):

        excel_interface.Workbook.__init__(*pargs, **kwargs)

    def create_sheet(self, name):
        sheet = excel_interface.Workbook.create_sheet(self, name)
        sheet.bb = SheetData()

        return sheet

        
