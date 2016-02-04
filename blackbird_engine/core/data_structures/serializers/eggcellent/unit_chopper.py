#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2016
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

#Blackbird Environment
#Module: data_structures.serializers.eggcellent.unit_chopper
"""

Module defines a class that represents arbitrarily rich BusinessUnit instances
as a collection of linked Excel worksheets. 
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
UnitChopper           chop BusinessUnit into dynamic Excel structure
====================  ==========================================================
"""




# Imports
import openpyxl as excel_interface

from .field_names import FieldNames
from .tab_names import TabNames




# Constants
# n/a

# Module Globals
get_column_letter = excel_interface.utils.get_column_letter
field_names = FieldNames()
tab_names = TabNames()

# Classes
class UnitChopper:
    """

    [Add doc string ]
    One tab per unit
    Children first
    Arbitrarily recursive (though should max out at sheet limit; alternatively,
    book should prohibit new sheets after that? or can have 2 limits: a soft
    limit where ModelChopper shifts into different representation mode, and a
    hard limit, where you just cant create any more sheets.)
    
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    n/a

    FUNCTIONS:
    chop_unit()           returns sheet with a SheetData instance at sheet.bb
    ====================  ======================================================
    """
    MAX_TITLE_CHARACTERS = 8

    
 
    def chop_unit(self, book, unit):
        """


        -> Worksheet


        Children should be spread before parent
        """
        children = unit.components.get_ordered()

        for child in children:
            
            self.chop_unit(book, child)

        sheet = self._create_unit_sheet(unit)
        unit.xl.set_sheet(sheet)

        self._add_params(sheet, unit)
##        self._add_life(sheet, unit)
        self._add_financials(sheet, unit)

        return sheet

        # Premise 1: by the time you run this routine, all children should already be in book
        # Premise 2: a unit without any children should be easy to spread on a sheet
        # Premise 3:

    def _add_financials(self, sheet, unit):
        unit.fill_out()
        # fill out populates the unit with all of the information
        
        for statement in unit.financials:
            for line in statement.get_ordered():

                self._spread_line(sheet, line)

    # Have to manage book depth (ie max sheets) #<--------------------------------------------------!!

    def _add_param_rows(self, sheet, params, active_column, label_column=None, master_column=None):
        """


        -> Worksheet


        "active_column" : integer, 1-based column index
        "label_column" : integer,
        "master_column" : integer

        if label_column or master_column are left blank, method will find
        indeces using .get_position(). to improve performance in loops, caller
        can supply these indeces as arguments.
        """

        # The parameter is specific to the period; we don't have it
        # on the page yet. Add a row and write the value there.

        parameters = sheet.bb.parameters

        if label_column is None:
            label_column = sheet.bb.parameters.columns.get_position(field_names.LABELS)
        if master_column is None:
            master_column = sheet.bb.parameters.columns.get_position(field_names.MASTER)
            
        for param_name in sorted(params):

            param_value = params[param_name]
            
            # Register the new row
            param_row = parameters.rows.ending + 1
            parameters.rows.by_name[spec_name] = param_row
            # TO DO: Could also use max_row() or bb.current_row

            # Add the label
            label_cell = sheet.cell(column=label_column, row=param_row)
            label_cell.value = param_name

            # Add the master value (from this period)
            master_cell = my_tab.cell(column=master_column, row=param_row)
            master_cell.value = param_value

            # Link the period to the master
            param_cell = my_tab.cell(column=active_column, row=param_row)
            link = self.formula_templates.ADD_COORDINATES
            link = link.format(coordinates=master_cell.coordinate)
            param_cell.value = link

        return sheet
            
    def _add_unit_params(self, sheet, unit):
        """


        -> Worksheet
        

        """

        period_column = sheet.bb.time_line.columns.get_position(unit.period.ends)
        label_column = sheet.bb.parameters.columns.get_position(field_names.LABELS)

        existing_param_names = unit.parameters.keys() & parameters.rows.by_name.keys()
        new_param_names = unit.parameters.keys() - existing_param_names
        
        for param_name in existing_param_names:
            param_value = unit.parameters[param_name]

            existing_row = sheet.bb.parameters.rows.get_position(param_name)
            data_cell = sheet.cell(column=period_column, row=existing_row)
            data_cell.value = param_value
            # should format the cell in blue or whatever for hardcoded

        new_params = dict()
        for k in new_param_names:
            new_params[k] = unit.parameters[k]

        self._add_param_rows(sheet, new_params, period_column)

        return sheet    
        
        # To Do:
        # - add formatting

    def _create_unit_sheet(self, book, unit):
        """


        -> Worksheet


        
        """
        name = str(unit.id.bbid)[-self.MAX_TITLE_CHARACTERS: ]
        sheet = book.create_sheet(name)
        # Could also check if the actual unit name is in book, then switch to id
        # if it is (or 4 chars of name + 8 chars of id); would make more sense
        
        unit.xl.set_sheet(sheet) #<--------------------------------------------------------------------------------------- IMPLEMENT ROUTINE

        # Should auto-hide components below level 2 or smtg to make the book cleaner; sheet_state="hidden"
        # Could have a relationships.depth number on each unit. Then if relationships.depth >= x, you
        # do something. Ask Erika to start working on refactoring .relationships out of Tags. 

        self._link_to_time_line(sheet, book)
        self._add_unit_params(sheet, unit)

        return sheet   

    def _link_to_area(self, source_sheet, local_sheet, area_name, group=False):
        """


        -> Worksheet


        "area_name" : string 
        """
        source_area = getattr(source.bb, area_name)
        local_area = getattr(local.bb, area_name, None)

        if local_area is None:
            local_area = local_sheet.bb.add_area(area_name)

        local_area.update(source_area)
        coordinates = {"sheet" : source_sheet.title}
        
        for row in source_area.rows.by_name.values():

            source_row = source_area.rows.starting + row
            local_row = local_area.rows.starting + row

            for column in source_area.columns.by_name.values():

                source_column = source_area.columns.starting + column
                local_column = source_area.columns.starting + column

                local_cell = local_sheet.cell(column=local_column, row=local_row)

                cos = coordinates.copy()
                cos["row"] = source_row
                cos["alpha_column"] = get_column_letter(source_column)

                link = self.formulas.ADD_CELL_FROM_SHEET.format(**cos)
                local_cell.value = link
                
        # if group:
        #   ##group cells
        
        return local_sheet

        # To do:
        # - if group: group the range

    def _link_to_time_line(self, book, sheet):
        """


        -> Worksheet
        

        Link sheet to book's time_line
        """
        source = book[tab_names.TIME_LINE]
        
        self._link_to_area(source, sheet, "time_line")
        self._link_to_area(source, sheet, "parameters")

    