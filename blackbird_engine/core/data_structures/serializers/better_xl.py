import openpyxl as excel_interface

class Bbook(excel_interface.Workbook):
    def __init__(self, *pargs, **kwargs):

        excel_interface.Workbook.__init__(*pargs, **kwargs)

    def create_sheet(self, name):
        sheet = excel_interface.Workbook.create_sheet(self, name)
        sheet.bb = SheetData()

        return sheet
    
class Formulas:
    SUM_RANGE = "+SUM({sheet}!{column}{starting_row}:{sheet}!{column}{ending_row}"
    ADD_CELL = "+{column}{row}"
    ADD_FOREIGN_CELL = "+{sheet}!{column}{row}"

class Eggcellent:
    """

    Objects that export models into Excel

    """
    formulas = dict()
    
    def rows_to_coordinates(self, table, column, sheet_name=None):
        """
        -> dict

        Returns dict with table keys, coordinate values
        """
        result = dict()
        alpha_column = get_alpha(column)
        
        for k, v in table.items():

            result[k] = alpha_column + str(v)

        return result

    def line_to_coordinates(self, line, column):
        result = line.xl.get_coordinates(column)
        return result
            
    def _spread_line_derivation(self, sheet, line, set_labels=True):
        """
        """
        for data_cluster in line.xl.derived.data:
            self._spread_driver_data(sheet, line, data_cluster, set_labels=set_labels)

    def _spread_driver_data(self, sheet, line, driver_data, set_labels=True):
        """
        -> sheet
        """
        label_column = sheet.time_line.cols.get_position("label") #<-----------------------------this can be sheet.labels if we
        private_data = sheet.parameters.copy()
        # Set up a private range that's going to include both "shared"
        # unit parameters from the column and "private" driver parameters
        
        for row_data in driver_data.rows:

            label_cell = sheet.cell(label_column, sheet.current_row)
            
            if set_labels:
                if label_cell.value is None:
                    label_cell.value = row_data["label"]
                else:
                    if label_cell.value != row_data["label"]:
                        raise Error
            # Check to make sure we are writing to the right row; if the label
            # doesn't match, we are in trouble.
            
            param_cell = sheet.cell(sheet.current_column, sheet.current_row)
            param_cell.value = row_data["value"]

            relative_position = sheet.current_row - private_data.rows.starting
            private_data.rows.by_name[label_cell.value] = relative_position #<-----------should be a relative position? relative to what?
            # should just keep private_data in absolute terms for simplicity ! <------------------------------------------------------------------------ 

            sheet.current_row += 1

        # Now, transform the range values from rows to coordinates
        param_coordinates = self.to_coordinates(param_rows, sheet.current_column)
        
        # Now, apply the conversions as necessary to make sure formula
        # sees all the params it wants
        conversion_map = driver_data.conversion_map
        for param_name, var_name in conversion_map.items():
            param_coordinates[var_name] = param_coordinates[param_name]

        # Finally, format the formula as necessary
        # (if references are a dict of objects, could map each obj to its coordinates)\
        template = driver_data.formula_string
        
        references = dict()
        for k, obj in driver_data.line_references.items():
            references[k] = self.obj_to_coordinates(obj)

        formula = template.format(references=references, parameters=param_coordinates)
        
        formula_cell = sheet.cell(sheet.current_column, sheet.current_row)
        formula_cell.value = formula

        # If formula included a reference to the prior value of the line itself, it
        # picked up here. Can now change line.xl.derived.final <--------------------------------------------incrementation is discouraged, right?
        line.xl.rows.derived.ending = sheet.current_row

        sheet.current_row += 1

        return sheet
        #<--------------------------------------------------------------------------------------formula should be able to deliver multiple rows
        #that build on each other (for complex calculations): could have something like last_step as the variable in params that we keep
        # changing

    def _spread_line_consolidation(self, sheet, line, set_labels=True):
        """

        Expects line to include a full range of pointers (?) or formulas?
        
        """
        # Add feature to control how many refs per cell (children_per_cell)

        line.xl.consolidated.starts = sheet.current_row
        
        for source_pointer in line.xl.consolidated.data:
            
            link_cell = sheet.cell(sheet.current_column, sheet.current_row)
            source_coordinates = source.xl.get_coordinates()
            
            link_formula = self.formulas.LINK_TO_CELL.format(**source_coordinates)
            link_cell.value = link_formula

            line.xl.consolidated.ends = sheet.current_row
            sheet.current_row += 1

        # Group the cells!! <--------------------------------------------------------------------------

        summation_params = {
            "starting_row" : line.xl.consolidated.starts,
            "ending_row" : line.xl.consolidated.ends,
            "column" : sheet.current_column
            }
        #<------------------------------------------------------------------------------------------------------------------------need to make sure column is alphabetical
        
        summation_formula = self.formulas.SUM_RANGE.format(**summation_params)
        summation_cell = sheet.cell(sheet.current_column, sheet.current_row)
        summation_cell.value = summation_formula

        if set_labels:
            # Add the "x : consolidated value" label #<---------------------------------------------------------------------------------------fix
            pass

        line.xl.consolidated.ending = sheet.current_row

        sheet.current_row +=1 #<---------------------------------------should move this to the manager function so can explictly see how we move up and down book
        return sheet
    
    def _combine_line_segments(self, sheet, line, set_labels=True):
        row_summation = ""
        coordinates = dict()
        coordinates["column"] = get_alpha(sheet.current_column) #<-------------------------------------------------sheet.bb should point to the interface i care about; cols and rows should be i_row, i_cols or even 
        
        if line.xl.consolidated.ending:
            cos = coordinates.copy()
            cos["row"] = line.xl.consolidated.ending
            link = self.formulas.ADD_CELL.format(**cos)
            row_summation += link
            # <------------------------------------------------------this should be a utility function

        if line.xl.derived.ending:
            cos = coordinates.copy()
            cos["row"] = line.xl.derived.ending
            link = self.formulas.ADD_CELL.format(**cos)
            row_summation += link

        if line.xl.detailed.ending:
            cos = coordinates.copy()
            cos["row"] = line.xl.detailed.ending
            link = self.formulas.ADD_CELL.format(**cos)
            row_summation += link

        total_cell = sheet.cell(sheet.bb.x, sheet.bb.y)
        total_cell.value = row_summation

        if set_labels:
            pass
            # This should also be a utility function: set_label() 

        line.xl.ending = sheet.bb.y
        return sheet

    def _spread_line(self, line, sheet):
        """
        by the end of this routine, the line and all its details should have a sheet assignment
        """

        line.xl.set_sheet(sheet)
        # Line can now deliver coordinates
        
        self._spread_consolidation_data(sheet, line, set_labels=set_labels)
        sheet.bb.y += 2

        self._spread_derivation_data(sheet, line, set_labels=set_labels)
        sheet.bb.y += 2

        details = line.get_ordered()
        if details:

            sub_indent = indent + 2
            summation = ""
            coordinates = dict()
            coordinates["column"] = sheet.bb.current_column

            for detail in details:
                self._spread_line(sheet, detail, set_labels=set_labels, indent=sub_indent)

                cos = coordinates.copy()
                cos["row"] = detail.xl.ending
                link = self.formulas.ADD_CELL.format(**cos)
                
                summation += link
        
                sheet.bb.current_row += 1
                
            else:
                # Should group all the details here
                
                subtotal_cell = sheet.cell(sheet.bb.i_col, sheet.bb.i_row)
                subtotal_cell.value = detail_summation

                line.xl.detailed.ending = sheet.bb.current_row
                
                if set_labels:
                    pass
                
                sheet.bb.current_row +=1

        self._combine_line_segments(sheet, line, set_labels=set_labels)
        # Could also group here

        return sheet
        # If I toggle this logic a bit, I should be able to put all of the consolidation
        # cells in one part of the sheet, and then the derivation logic elsewhere.

        # This relies on:
            # line.increment() tracking sources
            # driver.workOnThis() collating data
        
    # formula.to_excel()-> string, dict;
        # should

    def _spread_unit(self, book, unit):
        """

        Children should be spread before parent
        """
        children = unit.components.get_ordered()

        for child in children:
            
            self._spread_unit(book, child)

        sheet = book.create_sheet(unit.id.bbid)
        unit.xl.set_sheet(sheet)

        self._add_params(sheet, unit)
##        self._add_life(sheet, unit)
        self._add_financials(sheet, unit)

        return sheet

        # Premise 1: by the time you run this routine, all children should already be in book
        # Premise 2: a unit without any children should be easy to spread on a sheet
        # Premise 3: 

    def _add_financials(self, sheet, unit):
        unit.fill_out()
        # fill out populates the unit with all of the information
        
        for statement in unit.financials:
            for line in statement.get_ordered():

                self._spread_line(sheet, line)

    # Have to manage book depth (ie max sheets)

    def _create_unit_sheet(self, book, unit):

        name = str(unit.id.bbid)[-8:] #<-----------------------------------------------------------------should be a class var
        sheet = book.create_sheet(name)
        unit.xl.set_sheet(sheet)

        self._link_to_time_line(sheet, book)
        self._add_unit_params(sheet, unit)

        return sheet

    def to_excel(self, model):
        """

        -> book

        # save operation should be external probably
        
        """
        pass

    def _spread_scenarios(self, book, model):
        """
        -> should return sheet!
        """
        scenarios = book.create_sheet("Scenarios") #<--------------------------------------------------------------key should be standard

        # Sheet map:
        # A          |  B      | C             | D     | E
        # param names|  blank  | active values | blank | blackbird values

        starting_row = 1
        starting_column = 1
        # should make this a lookup table too: column name to index
        
        scenarios.bb.general.columns.by_name["params"] = 1
        scenarios.bb.general.columns.by_name["active"] = 3
        scenarios.bb.general.columns.by_name["base"] = 5
        # these keys should be standard
        # all other scenarios should be a function of selection
        # can do something like bb_case cells are an index of bb scenarios
        # and then active just points to those

        active_row = starting_row

        f_pull_from_cell = "=%s"
        
        for param_name in sorted(model.time_line.parameters):
            # Use sorted to make sure order is consistent

            label_cell = scenarios.cell(row=active_row, column=starting_column)
            scenarios.bb.general.rows.by_name[param_name] = active_row

            active_cell = scenarios.cell(row=active_row, column=(starting_column+2)) #<--------------------------------fix the column indexing
            bb_cell = scenarios.cell(row=active_row, column=(starting_column+4))
            # ideally should be active_cell.column+2 <---------------------------------------------------------------- fix the coordinates
                        
            label_cell.value = param_name
            bb_cell.value = model.time_line.parameters[param_name]

            active_cell.value = f_pull_from_cell % bb_cell.coordinate

            active_row += 1
            
        #Can expand this logic to print every scenario
        
        return scenarios

    def _spread_time_line(self, book, model):
        """

        -> sheet ?

        book should have a "scenarios" tab by now
        """
        
        scenarios = book["Scenarios"] #<-------------------------------------------------------------------------------- search by standard key
        
        my_tab = book.create_sheet("Timeline") #<------------------------------------------------------------------------add standard key
        my_tab.bb.parameters = Area()

        get_column_letter = excel_interface.utils.get_column_letter

        # first column will be params
        # second column will be blank
        # third column will be master
        
        col_params = 1
        col_master = 3
        # have to record column <--------------------------------------------------------------------------------------------------------------

        header_row = 3
        
        f_active_scenario = "=Scenarios!C%s"
        # first, write the param colum and the master column: "=Scenarios! $A1"

        # Make the param name and master value columns
        for param_name, row_number in scenarios.bb.general.rows.by_name.items(): #<-----------------------------------need to make sure we get the right position here
            # Order doesn't matter here

            active_row = header_row + row_number
            my_tab.bb.parameters.rows.by_name[param_name] = active_row
            #<--------------------------------------------------------------------------------have to think about whether these are absolute or relative
            
            name_cell = my_tab.cell(column=col_params, row=active_row)
            name_cell.value = param_name

            master_cell = my_tab.cell(column=col_master, row=active_row)
            master_cell.value = f_active_scenario % row_number

        f_pull_master = "=" + "$" + get_column_letter(col_master) + "%s"
        # "=$A1"
        
        starting_column = col_master + 2
        active_column = starting_column

        my_tab.bb.time_line.rows.by_name["labels"] = header_row
        
        for period in model.time_line.get_ordered():

            my_tab.bb.time_line.columns.by_name[period.end] = active_column
            my_tab.bb.parameters.columns.by_name[period_end] = active_column
            
            header_cell = my_tab.cell(column=active_column, row=header_row)
            header_cell.value = period.end

            # first, add period-level parameters
            # first, drop the pull value for all cells through the params thing
            for param_row in my_tab.bb.parameters.rows.by_name.values():
                
                # Note: will probably write cells in non-sequential order

                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = f_pull_master % param_row

                # Or could add links to the prior period's parameters <--------------------------

            # then, overwrite them with explicitly specified params
            for spec_name, spec_value in period.parameters.items():
                param_row = my_tab.bb_row_lookup[spec_name]
                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = spec_value
##                spec_cell.format = blue_font_color
            
            active_column += 1
            # Move on to the next period

        return my_tab

    def _link_to_time_line(self, book, sheet):
        """

        -> sheet

        Link sheet to book's time_line
        """
        source = book[self.names.TIME_LINE]
        
        self._link_to_area(source, sheet, "time_line")
        self._link_to_area(source, sheet, "parameters")

    def _add_unit_params(self, sheet, unit):
        """

        -> sheet
        
        """

        period_column = sheet.bb.time_line.columns.get_position(unit.period.ends)
        label_column = sheet.bb.parameters.columns.get_position("label")
        
        for param_name in sorted(unit.parameters):

            param_value = unit.parameters[param_name]

            if param_name in sheet.bb.parameters.rows.by_name:

                existing_row = sheet.bb.parameters.rows.get_position(param_name)
                data_cell = sheet.cell(column=period_column, row=existing_row)
                data_cell.value = param_value
                # should format the cell in blue or whatever for hardcoded

            else:

                new_row = sheet.bb.parameters.ending + 1
                data_cell = sheet.cell(column=period_column, row=new_row)
                data_cell.value = param_value

                label_cell = sheet.cell(column=label_column, row=new_row)
                label_cell.value = param_name

        return sheet

    def _link_to_area(self, source_sheet, local_sheet, area_name):
        """
        """

        source_area = getattr(source.bb, area_name)
        local_area = getattr(local.bb, area_name, None)
        if local_area is None:
            local_area = Area()
            setattr(local.bb, range_name, local_area)

        local_area.update(source_area)
        
        coordinates = {
            "sheet" : source_sheet.name
            }
        
        for row in source_area.rows.by_name.values():

            source_row = source_area.rows.starting + row
            local_row = local_area.rows.starting + row

            for column in source_area.columns.by_name.values():

                source_col = source_area.columns.starting + column
                local_col = source_area.columns.starting + column

                local_cell = local_sheet.cell(column=local_column, row=local_row)

                cos = coordinates.copy()
                cos["row"] = source_row
                cos["column"] = source_column

                link = self.formulas.ADD_FOREIGN_CELL.format(**cos)
                local_cell.value = link
                
        # if group, group starting and ending
        return local_sheet

    
               
        
        

        
        
    


    
    
    
        
        
    
        
        

        

    
    
            

            

            

            


            

            

            
            

                

                
            

            

                
        

    
