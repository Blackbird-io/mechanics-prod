#PROPRIETARY AND CONFIDENTIAL
#Property of Blackbird Logical Applications, LLC
#Copyright Blackbird Logical Applications, LLC 2015
#NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL OF ILYA PODOLYAKO

#Blackbird Environment
#Module: data_structures.modelling.model
"""

Module defines Model class, a container that fully describes a company's past,
present, and future.
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
Model                 structured snapshots of a company across time periods
====================  ==========================================================
"""




# imports
import time
import dill

import bb_exceptions
import bb_settings
import openpyxl as excel_interface

from data_structures.guidance.interview_tracker import InterviewTracker
from data_structures.system.bbid import ID
from data_structures.system.tags import Tags

from .time_line import TimeLine




# constants
# n/a

# classes
class Model(Tags):
    """

    This class provides a form of a standard time model of business performance.
    A Model object revolves around a Timeline, which consists of a list of
    snapshots of the business at different points in wall time. Each such
    snapshot is an instance of a TimePeriod object.

    Each Model instance has its own namespace_id derived from the origin
    Blackbird UUID. The Model's namespace_id is the source of truth for business
    units within that model. In other words, business units have ids that are
    unique within the Model. If a business unit has an id that's equal to that
    of another business unit, they represent the same real life referent within
    a given model.

    A Model can and should contain multiple instances of the same business unit,
    as determined by the business unit's bbid. Each TimePeriod in the TimeLine
    should contain its instance for the BusinessUnit to show how it evolves over
    time.

    On the other hand, each BusinessUnit should only appear once per TimePeriod,
    because within a given Model, the BusinessUnit can only have single state
    at a given point in time.

    To enforce the above requirement, each TimePeriod has a directory of all
    business units that exist within it. The directory is a dictionary of
    pointers to BUs, keyed by bbid, available at TimePeriod.bu_directory. The
    directory cuts across all levels of the BusinessUnit hierarchy within the
    TimePeriod: it includes keys for the top-level BU (TimePeriod.content) and
    any children, grandchildren, etc that BU may have.

    All BusinessUnits within a TimePeriod use the directory to check whether
    they can add a new business unit as a component. If the business unit's bbid
    is already in the directory, they cannot. If it is new to the TimePeriod,
    they can.     
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    id                    instance of ID object, carries bbid for model 
    interview             instance of an InterviewTracker object 
    portal_data           dict; stores data from Portal related to the instance
    stage                 P; pointer to either interview or defined _stage
    started               bool; property, tracks whether engine has begun work
    summary               P; pointer to current period summary
    taxonomy              dict with tree of unit templates
    time_line             list of TimePeriod objects
    transcript            list of entries that tracks Engine processing
    used                  set of bbids for used topics
    valuation             P; pointer to current period valuation

    FUNCTIONS:
    from_portal()         class method, extracts model out of API-format
    start()               sets _started and started to True
    transcribe()          append message and timestamp to transcript
    ====================  ======================================================
    
    ``P`` indicates attributes decorated as properties. See attribute-level doc
    string for more information.
    """    
    def __init__(self, name):
        Tags.__init__(self,name)
        self._stage = None
        self._started = False
        #
        self.id = ID()
        self.id.assign(name)
        # Models carry uuids in the origin namespace.
        self.interview = InterviewTracker()
        self.portal_data = dict()
        self.taxonomy = dict()
        self.transcript = []
        self.time_line = TimeLine()
        self.used = set()
        #
        self.time_line.id.set_namespace(self.id.bbid)

    #DYNAMIC ATTRIBUTES
    @property
    def stage(self):
        """


        **property**


        When instance._stage points to a True object, property returns the
        object. Otherwise property returns model.interview.

        Since the default value for instance._path is None, property starts out
        with a ``pass-through``, backwards-compatible value. 
        
        Setter sets _stage to value.

        Deleter sets _stage to None to restore default pass-through state.
        """
        result = self._stage
        if not result:
            result = self.interview
        return result

    @stage.setter
    def stage(self, value):
        self._stage = value

    @stage.deleter
    def stage(self):
        self._stage = None
   
    @property
    def started(self):
        """


        **read-only property**


        Once True, difficult to undo (a toggle that sticks).
        """
        return self._started

    @property
    def summary(self):
        """


        **read-only property**


        Pointer to company summary on current period. If current period has no
        content, returns None.
        """
        result = None
        #
        company = self.time_line.current_period.content
        if company:
            #catch periods with empty content
            result = company.summary
        #
        return result

    @summary.setter
    def summary(self, value):
        c = "Assignment prohibited. ``model.summary`` serves only as a pointer"
        c += " to the current period company summary."
        raise bb_exceptions.ManagedAttributeError(c)

    @property
    def valuation(self):
        """


        **read-only property**


        Pointer to company valuation on current period. If current period has no
        content, returns None.
        """
        result = None
        #
        company = self.time_line.current_period.content
        if company:
            #catch periods with empty content
            result = company.valuation
        #
        return result

    @valuation.setter
    def valuation(self, value):
        c = "Assignment prohibited. ``model.valuation`` serves only as a pointer"
        c += " to the current period company valuation."
        raise bb_exceptions.ManagedAttributeError(c)

    #METHODS
    @classmethod
    def from_portal(cls, portal_model):
        """

    
        Model.from_portal(portal_model) -> Model


        **CLASS METHOD**

        Method extracts a Model from portal_model.

        Method expects ``portal_model`` to be a string serialized by dill (or
        pickle).

        If portal_model does not specify a Model object, method creates a new
        instance. Method stores all portal data other than the Model in the
        output's .portal_data dictionary.         
        """
        M = None
        flat_model = portal_model["e_model"]
        #
        if flat_model:
            M = dill.loads(flat_model)
        else:
            business_name = portal_model["business_name"]
            if not business_name:
                business_name = bb_settings.DEFAULT_MODEL_NAME
            M = cls(business_name)
        M.portal_data.update(portal_model)
        del M.portal_data["e_model"]
        return M
    
    def start(self):
        """


        Model.start() -> None


        Method sets instance._started (and ``started`` property) to True. 
        """
        self._started = True

    def transcribe(self, message):
        """


        Model.transcribe(message) -> None


        Appends a tuple of (message ,time of call) to instance.transcript.
        """
        time_stamp = time.time()
        record = (message,time_stamp)
        self.transcript.append(record)

    def spread(self, path):
        """
        Return an Excel model
        """
        pass

    def _spread_foundation(self):
        """
        Return a workbook with:
           cover
           scenarios
           timeline
           
        """       
        wb = excel_interface.Workbook()
        self._create_scenarios_tab(wb)
        self._create_time_line_tab(wb)
        return wb

    def _create_scenarios_tab(self, book):
        scenarios = book.create_sheet("Scenarios")
        scenarios.bb_row_lookup = dict() #<----------------------------should be Statement so we can manage insertions
        scenarios.bb_col_lookup = dict()
        # Should just create an object with both of these things

        # Sheet map:
        # A          |  B      | C             | D     | E
        # param names|  blank  | active values | blank | blackbird values

        starting_row = 1
        starting_column = 1
        # should make this a lookup table too: column name to index
        
        scenarios.bb_col_lookup["params"] = 1
        scenarios.bb_col_lookup["active"] = 3
        scenarios.bb_col_lookup["base"] = 5
        # these keys should be standard
        # all other scenarios should be a function of selection
        # can do something like bb_case cells are an index of bb scenarios
        # and then active just points to those

        active_row = starting_row

        f_pull_from_cell = "=%s"
        
        for param_name in sorted(self.time_line.parameters):

            param_cell = scenarios.cell(row=active_row, column=starting_column)
            scenarios.bb_row_lookup[param_name] = active_row

            active_cell = scenarios.cell(row=active_row, column=(starting_column+2))
            bb_cell = scenarios.cell(row=active_row, column=(starting_column+4))
            # ideally should be active_cell.column+2
                        
            param_cell.value = param_name
            bb_cell.value = self.time_line.parameters[param_name]

            active_cell.value = f_pull_from_cell % bb_cell.coordinate

            active_row += 1
            
        #Can expand this logic to print every scenario
        
        return book
            
    def _create_time_line_tab(self, book):
        
        scenarios = book["Scenarios"]
        
        my_tab = book.create_sheet("Timeline")
        my_tab.bb_row_lookup = dict()
        my_tab.bb_col_lookup = dict()

        get_column_letter = excel_interface.utils.get_column_letter

        # first column will be params
        # second column will be blank
        # third column will be master
        
        col_params = 1
        col_master = 3
        # have to record column <---------------------------------

        header_row = 3

##        f_active_scenario = "=" + str(scenarios.title) + "!" + "$" + str(scenarios.bb_col_lookup["active"]) + "%s"
        f_active_scenario = "=Scenarios!C%s"
        # first, write the param colum and the master column: "=Scenarios! $A1"

        # Make the param name and master value columns
        for param_name, row_number in scenarios.bb_row_lookup.items():
            # Order doesn't matter here

            active_row = header_row + row_number
            my_tab.bb_row_lookup[param_name] = active_row
            
            name_cell = my_tab.cell(column=col_params, row=active_row)
            name_cell.value = param_name

            master_cell = my_tab.cell(column=col_master, row=active_row)
            master_cell.value = f_active_scenario % row_number

        f_pull_master = "=" + "$" + get_column_letter(col_master) + "%s"
        # "=$A1"
        
        starting_column = col_master + 2
        active_column = starting_column
        
        for period in self.time_line.get_ordered():

            # record column in lookup table so future routines can coordinate
            my_tab.bb_col_lookup[period.end] = active_column
            
            header_cell = my_tab.cell(column=active_column, row=header_row)
            header_cell.value = period.end

            # first, add period-level parameters
            # first, drop the pull value for all cells through the params thing
            for param_row in my_tab.bb_row_lookup.values():
                
                # Note: will probably write cells in non-sequential order

                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = f_pull_master % param_row

                # Or could add links to the prior period's parameters <--------------------------
                # Can probably also make this a named range or something? Cause these formulas
                # are all the same.

            # then, overwrite them with explicitly specified params
            for spec_name, spec_value in period.parameters.items():
                param_row = my_tab.bb_row_lookup[spec_name]
                param_cell = my_tab.cell(column=active_column, row=param_row)
                param_cell.value = spec_value
##                spec_cell.format = blue_font_color
            
            active_column += 1
            # Move on to the next period

        return book

    

        #//

        # later routines:
            # in business unit, should have ._add_time_line(tab)
            # 

        # should have written the full timeline
        # should enter all of these into bb_col_lookup
            #by end date
            #then unit-specific sheets can pick those up        
    
        # but better logic is:
            # write the date
            # write the period      

        # first, carry over the master
        ##for every row that's filled out in params, should also copy here
        ##need to know what column params start in
        ##so let's establish the params as a named range

        # ranges i set up here:
        # end_dates
        ## need to somehow be able to access the params on every unit
        ## could just have them copied

        # for cell in tl_param_names:
            # param names in th

        # for cell in tl_param_values:
            # put in the master column

        # then for every period, ...
            # period params should point to master by default
                #populate these first
                #starting_row
                #period_

            # if time_line.period has a certain value,
                # we should overwrite it and hardcode it (a particular color)
                # find the row, write there?

        # can probably delegate this down to the timeline itself
        # time_line should raise error if there is no scenarios tab
            #would then be easier to ask each period to do its thing?
            #or alternatively would pick every unit in the current period and go from there
                #yeah, should start with the current period, and then step through the company hierarchy
        
    
        # Loren's point: create an intermediate representation in memory
            # then can write that quickly
            # so would probably set up sheets (in memory), formulas, etcetera

        #for unit:
            #take the active tab
            #for each line, add consolidation functionality
            #have to connect old balance sheet as starting
            
        #for driver:
            # need to map ``data`` to cell coordinates
                #as in, take each key and connect it to a cell
                    #populate the cell with the driver's value
                    #remember the coordinates

                #
            
            # need to translate formula to one that works on cell coordinates
            # formula should have an _xl equivalent
                # which should take the actual formula and turn it into excel compatible string
                # use string formatting to convert var names into coordinates

            # for something like ebitda calc, would need to:
                # know the column
                # figure out the relevant rows for each input cell
                ## could have a general conversion routine for .find_first(), which looks up name in lookup table
                ## may also want to organize lookup tables by section to avoid name collisions
                ## so have a .income or .cash section where names have to be unique

        # should put all of this stuff into a single class called ExcelConverter
        # which could then have interfaces and other stuff

        

    
        
            
