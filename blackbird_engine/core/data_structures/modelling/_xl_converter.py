#Excel converter class

import openpyxl as xlio

class BBWorkBook(xlio.WorkBook):

    def __init__(self):        
        xlio.WorkBook.__init__(self)
        # Should probably use composition here: ._native_book = xlio.WorkBook()?

    def create_sheet_with_lookups(self, *pargs, **kwargs):

        sheet = self.create_sheet(*pargs, **kwargs)
        sheet.row_lookup = dict()
        sheet.col_lookup = dict()
    
        return sheet

    # Add range-management features (ie, make sure no ranges overlap)
    # can add a bb_ranges attribute that's just a dictionary. Then
    # would add ranges there.

    def validate(self):
        pass
        # check that none of the ranges overlap
    
class AttributeRange:
    """
    A range of cells that describe a particular attribute. 

    All values should be relative to starting points, 0-indexed. For example,
    row_lookup = {"revenue":4} means that the revenue figure is in the 5th row
    from the start. If instance.starting_row == 3, revenue will be in row 7 (8th
    row from the top).

    Ranges cannot overlap...? 
    """
    #<---------------------- Replace this guy with a real range object?
    
    def __init__(self, starting_col=0, starting_row=0):

        self.starting_col = starting_col
        self.ending_col = None

        self.starting_row = starting_row
        self.ending_row = None
        
        self.row_lookup = dict()
        self.col_lookup = dict()
        # These should be Statement objects to support nesting

    def get_row(self, label):
        relative_position = self.row_lookup[label] 
        result = self.starting_row + relative_position
        return result

    def get_column(self, label):
        relative_position = self.col_lookup[label]
        result = self.starting_col + relative_position
        return result
        
    def update(self, source):
        self.row_lookup.update(source.row_lookup)
        self.col_lookup.update(source.col_lookup)

        # Update ending values
        if self.col_lookup:
            highest = max(self.col_lookup.values())
            self.ending_col = self.starting_col + highest

        if self.row_lookup:
            highest = max(self.row_lookup.values())
            self.ending_row = self.starting_row + highest        

    def validate(self):
        """
        Return True or error if endpoints dont match up
        """
        pass


class RowTracker:
    """
    Used for tracking line spreads
    """
    def __init__(self):
        self.consolidated = None
        self.derived = None
        self.final = None

class CellRange:
    """

    Most basic
    """
    def __init__(self):
        self.starting
        self.ending

class ExcelConverter:
    def convert(self, model):
        """
        Return excel model
        """
        pass
    
    def save(self, model, path):
        """
        Save model to path
        """
        pass

    def _make_cover_tab(self, model, book):
        pass

    def _make_scenarios_tab(self, model, book):
        pass

    def _make_time_line_tab(self, model, book):
        pass

    # move to a general ledger
    # then balance sheet consolidation would be easier:
        # start with starting balance sheet
        # increment ending balance sheet by all of the child entries

    def _spread_unit(self, unit, book):
        children = unit.components.get_ordered()
        if children:
            # complex routine
            pass
        else:
            # simple routine

    def _spread_childless_unit(self, unit, book):
        """
        # create a tab for the unit
            # tab should be named same as unit? or same as unit id?
            # let's say id, that's guaranteed to be unique

        # add timeline info to that tab
            # link to timeline dates
            # link to timeline params

        # add unit-level params
            # overwrite timeline parameters
            # [should be in different color]

        # [add life]

        # [add relationships]
        
        # add financials

        housekeeping:
          return book
          active tab should be where you made the unit
          
        """
        sheet = book.create_sheet(unit.id.bbid)

        sheet.row_sections = Statement()
        # or could have segments 
        sheet.col_sections = Statement()

        

        # Can add a profile section here: name (or at the top)

        # Should configure each unit sheet with sections: timeline, parameters, life, financials

            # Could even add components: a list of all component ids in that period
            # Could then do consolidation through an actual routine in VBA:

                # for each unit id, find the tab, find the line, etc.
                # works better with named ranges <--------------------------------------------------------------------------

                # in the meantime, can just have a row_sections and col_sections dictionary
                # can make this a statement so they nest
                
        self._link_to_timeline(sheet)
        # <---------------------------------------------- implement

        self._add_unit_params(sheet, unit.params, column)
        # <---------------------------------------------- implement

    def _build_unit_sheet(self, book, unit):
        """
        -> sheet
        """

        sheet = book.create_sheet(unit.id.bbid)
        sheet.time_line = Range(self.TL_STARTING_ROW, self.TL_STARTING_COL)
        self._link_to_time_line(sheet, book)

        sheet.parameters = Range()
        # Populate the unit params

        sheet.life = Range()
        # Populate unit life (could be events?)        
                
        return sheet
           
            # Each section should have a starting_, ending_, and elements
                # lo and hi?

            # elements should be a Statement() and should nest (?)
                # nesting may not be useful, because then cant see top-down
                # but you do want to be able to find cells in a given range:
                    # search for assets in ending balance sheet
                    # search for cash in starting balance sheet

            # point is, you need structure
                # some structure is driven by sheet-level setup
                # some structure is driven by object characteristics
                    # for now, i can make the unit-tracking sheet manually
                    # ideally though, it should be a function of the unit itself
                        # so unit could pick out the attributes that get their own sections
                        # and each attribute could write itself, sort of

            # so its almost like:
                # create_sheet:
                    # sheet has certain attributes, like sections
                    # could have "general" key

                # add_timeline:
                    # adds other attributes to the sheet; new sections

                # add_unit_info:
                    # adds a whole swath of other data to the sheet

                    # what if each object had it's own xl_map attribute
                    # would contain rows
                    # would need to make sure rows appear only once
                    # would need to make sure it doesnt overlap with others
        # \\
        # Could have individual classes for different kinds of tabs
        # UnitSheet(unit) -> sheet with all of the stuff, standard offsets, etc.
        # but not that great; similar to a function, and may need multiple sheets (ie a book)
        
    def _link_to_time_line(self, sheet, book):
        """

        -> None

        Link sheet to the book's timeline.
        
        """

        source = book[self.SHEET_NAME_TIME_LINE]
        # Timeline has columns with info. The first column is the index. Others
        # are periods. We want to copy the index and the periods. We can't iterate
        # over cells because that creates them in memory.

        # This is an element we should definitely fix in pandas <--------------------------------------------------------------------------------
        linking_formula = "{sheet_name}!{column}{row}"
        linking_formula.format(
            {
                "sheet_name" : source.name
             }
            )

        # Use 3.x string formatting for more explicit syntax
        # formulas can be constant at class or module level    <-------------------------------------------------------------------------------------
        local = sheet
        local.time_line.update(source.time_line)

        for row in source.time_line.row_lookup.values():

            source_row = source.time_line.starting_row + row
            local_row = local.time_line.starting_row + row

            for column in source.time_line.column_lookup.values():

                source_col = source.time_line.starting_col + column
                local_col = local.time_line.starting_col + column 

                cell = local.get_cell(local_col, local_row)
                cell.value = linking_formula.format(source_col, source_row) 

            # May be we can automate this whole thing? Call it _link_to_range() <--------------------------------------------------------------------
            # _link_to_range(local, source, range_name)

    def _link_to_period_parameters(self, sheet, book):

        source = book[TL_NAME]

        sheet.parameters.update(source.parameters)

        for row in source.parameters.row_lookup.values():
            # can be sorted if necessary

            source_row = source.parameters.starting_row + row
            local_row = sheet.parameters.starting_row + row

            for column in sheet.time_line.col_lookup.values():

                source_col = source.time_line.starting + column
                local_col = local.time_line.starting + column

                cell = local.get_cell(local_col, local_row)
                formula = link_to_cell.format(source_col, source_row)
                cell.value = formula

        # may want to group this stuff

    def _link_to_range(self, local, source, range_name, group=True):
        """

        -> sheet

        Updates local range with source data. 

        ``local`` : sheet that receives data
        ``source``: sheet that provides data
        ``range_name``: name of BB-style range

        Expects sheets to be in the same book, otherwise formulas won't work
        (unless you switch the formula)
        
        """
        local_range = getattr(local, range_name)
        source_range = getattr(source, range_name)

        local_range.update(source_range)

        # Since linking copies all existing data from source, the cell-wise
        # order of operation should not matter. You can copy the cellsf in 3x4
        # array starting anywhere and get the same result every time, especially
        # since we never compute anything during the export runtime (so
        # Excel will evaluate cell dependencies only when we have long since
        # finished writing the formulas.

        for row in source_range.row_lookup.values():
            # Can add a sorted parameter here

            source_row = source_range.starting_row + row
            local_row = local_range.starting_row + row

            for column in source_range.col_lookup.values():
                # Expects source range to contain proper column data <-------------------------------------------------!
                # Later, we can introduce ability to link row range across specific columns

                source_col = source_range.starting_col + column
                local_col = local_range.starting_col + column

                local_cell = local.get_cell(local_col, local_row) #<-------------------------------------------------------check syntax
                custom_link = linking_formula.format(             #<-------------------------------------------------------add var, check syntax
                    {
                        "sheet" : source.name,
                        "column" : source_column,
                        "row" : source_row
                        }
                    )
                local_cell.value = custom_link

        #if group: group starting to ending
                
        return local
                    
    # general unit routine:
        # (single period)
        #
        # set starting point on local tl
        # link timeline
        #
        # set starting point on local params (end of tl + 2)
        # link params
        # add unit params
        #
        # [if components]
        # spread components
        # add the consolidation range
        #
        # add the derivation range
        #
        # add the clean financials range
        #
        # that's it.
        

    def _add_unit_parameters(self, local, unit, column):
        """
        add to active sheet?
        """
        #<-------------------------------------should have a current_period_column attribute on my sheets

        if active_column is None:
            active_column = sheet.time_line.current_period_column
            # or could do get_current() and mathc unit.period.ends to known
            # columns in timeline. probably better <----------------------------------------------
                        
        ordered_params = sorted(unit.parameters.items(), key = lambda x: x[0])
        # Sort parameters so they appear in constant order

        for parameter_name, parameter_value in ordered_params:

            parameter_row = None
            existing_row = local.parameters.row_lookup.get(parameter_name)

            if existing_row:
                parameter_row = local.parameters.get_row(parameter_name)               

            else:
                parameter_row = sheet.parameters.ending_row + 1
                # write the param name to the index column 
            
            cell = local.get_cell(current_column, parameter_row)
            cell.value = parameter_value
            # add formatting
                # cell.format(local_data_in_blue) <------------------------------------------------------------------------- add formatting

        return local

    def _add_unit_life(self, sheet, unit, column):

        sheet.life = AttributeRange()
        # should start where Params end + 2
        # walk through events in order of keys
        # add each event to the section
        # check if event is already in section, if not, add

        # later, this can link to templates

        # <------------------------------------------------------------------------------------------------------------ finish

    # Both of these should be in _add_financials(), which will also add the clean
    # section
    
    def _add_consolidating(self, sheet, unit):
        # Figure out single-column or multicolumn? Multicolumn is more efficient
        pass

    def _add_derived(self, sheet, unit):
        # go through lines recursively
        # add drivers
        # when you finish details, add a summation line



    def _spread_statement(self, unit, statement, column):
        """

        ? default column is current_period
          or active column?
        """
        # look up line in consolidation section
        #   # if it's there, add a first subtotal with a link to the consolidated result
        # 
        # if line.details:
            # for detail in line.get_ordered():
                # _spread_line(detail)
            # else:
                # _add_summation()
        # else:
            # get drivers
            # for each driver, _spread_driver(starting_point)

        for line in statement.get_ordered():
            if line._details:
                for detail in line.get_ordered():
                    self._spread_statement(detail, unit, column=column)
                    # Recursion
            else:
                drivers = unit.drivers.get_drivers(line)
                for driver in drivers:
                    self._spread_driver(column)
        else:
            # add summation
            self._add_summation()
            # should cover private range:
            # so line starts at x, ends at y
            # sum should be of all the things
            #

        # Have to keep writing the line names in the params column!!
        # Have to return the coordinates of the cell where the total is
            # Or may be return the range!
            # Right back to question of whether we should have nested ranges
                #Answer: probably
                #Let's try to piggyback on existing statement
                    # need:
                        # relative position (in range)
                        # starting position
                        # cumulative range for details
                        # find abilities

    # let's say:
        #statement.xl = AttributeRange()
            # statement.add_detail()
                #detail.xl.starting_row = something
                #statement.xl.row_lookup[detail.name] = x (something to do with rel positioning)

                # that means that every detail will come with its own row reference baked in!

            # statement.spread():
                # can do the work?            
            

    # line.xl.rows:
        # starting = int
            ## when you add this line as a detail, its xl.starting value changes
            ## to reflect relative position in the parent and parent's thing

            ## issue: how to deal with manual changes in position?
            ## presumably want to make xl position dynamic? so that it's always
            ## starting plus relative?
            ## or can assign based on order in get_ordered()
        
        # ending = int
        # lookup = ?
            ## for lookup, you have to do the usual line interface
            ## will go as deep as you want
            ## then you can get the line position

    # line.xl.columns:
        ## undefined
    

    ## each detail has its own range?
        
    
    #\\ 
    # can integrate the fill out and spreading routines
    # so you'd build the model but not extrapolate it, and then spread it

    # can build the model in parallel, one sheet at a time

    #line.xl.rows:
        #.consolidated
        #.derived
        #.cumulative
        #.clean

    #can walk one tree multiple times
        # [for simple line, without details]:
        
        #first build out the consolidated section of the spreadsheet
            #add links to all of the source units
            #sum them
            #record the subtotal row in .consolidated

        #second build out the derived section
            # [should be blank if consolidated]
            # spread each driver. after each, record location of in .derived (or .value)
        
        #third:
            # cumulative ops:
                # add all the details
                # make a formula that adds the .final value of all the details
                    # literally, for each detail, formula += "+{detail coordinates}"
                    
    # recap: [at line level]

        # 1. consolidate
            # [separate section?]
        # 2. derive (check if this should be last)
        # 3. recur for details
        # 4. total

    # so what should line.xl.rows look like?
        # .consolidated
        # .derived
        # .final
        
    # 
    #
    #
    def _fill_out(self, column):
        pass
        
    def _consolidate(self, column):
        # fill out each unit in components

        # walk down lines in financials
            # for each line:
                ## have to figure out what to increment
                ## otherwise will just keep adding pointers to sums
                ## want to have see-through financials at the top
                ##
                ## should just track existing logic:
                    ## if don't have detail, add it:
                        #i.e., add link
                        #needs to be recursive
                    ## if you do have the detail, increment it:
                        #add a line with a link to value

                ## basically, if you are working with a spreadsheet, consolidate() should automatically fill out
                ## the column
                    ## can keep passing the sheet down to the lines during increment()
                    ## when increment adds a detail, it will add a cell to the spreadsheet
                    ##
                    ## since you are going detail by detail, can add details to the bottom
                    ## of the current open item (will never close an item before completing all its details)
                    ##
                    ## but you do need to work on the details in order (can expedite this in c eventually) to
                    ## make sure they display in the same order every time
                    ##
                    ## can do this recursively
                    ##
                    ## if you do this during consolidate() run-time, you may add a bunch of lines from one
                    ## unit and then have to add a bunch of lines from another unit
                    ## <-----------need to be able to insert rows!!
                        ## alternative to inserting rows:
                            ## could just write the cell coordinate and value on the line
                            ## would leverage existing position management
                            ## but conceptually complicated

            
        pass

        ## let's say each line has a range for each of the steps
            ## goal is to maintain relative position and then assemble everything later
    
        ## challenge is how to maintain links
    
            ## upside:
                ## when you are doing the consolidation linking, the sources (children)
                ## already have stable sheets
                ## so can actually just specify the value as a formula: "Child!$A$3"
                ## you just need to know where to put this formula, and that's where dynamic ordering comes in
                ##

                ## basically, build a column (ie 1-dim array) of formulas during consolidation
                ## put that column segment in on .spread() or something like that (when actually constructing
                ## the sheet)

                ## the other stuff i should be able to build in order, so not as crazy
                    ## data should be in columns (a list should be column) to make appends easy!
                ## so:
                    # line.xl.consolidating.cells := []
                    # line.xl.consolidating.value := formula that adds all the cells
                    #
                    ## when you are assembling the column, write it in order
                    ## then add a formula that tabulates it

            ## then when you go on to derive, you can
                ## put the consolidated cells in,
                ## add the drivers,
                ## add the details
                ## add the total (formula that sums total for every detail)

            ## have to track starting rows anyways to make dynamic positioning work
                ## basically, starting + index in ordered

            ## 
                    
            ## 

            ## 
     

    def _derive(self, column):
        pass
        
    def _fill_out(self, column):
        # run the consolidate operation, pick up the cells
        # during derive:
            # already going through lines (switch to recursive)?
            # for each line:
                # get drivers
                # spread drivers in the column

                # move on to details
                # 

    def _spread_driver(self, range):
        # set up a private range
        # add a line for every datapoint in the driver's formula
        # link cells for the datapoints to existing params
        # if the driver contains its own data, overwrite that data where appropriate
        # map the formula's other inputs somehow
        # add the formula, as adjusted for this spreadsheet
            # include comments where necessary
            # comment can include code for complex formulas
            # that's the subtotal
        # update reference for the line
            # what does this mean
                # originally, i meant that you would be changing the row value associated
                # with the line name in the general lookup. the result would be that the
                # line name would point to the updated row value, which could potentially
                # be non-consecutive

            # but now...
                # ...unclear
                # goal: locate cell where driver finished work
                # ideally, this cell should be in the line range
                # we will need to be able to track the line down later

            # could potentially add .derive, .consolidate, and .clean sections to line.xl
                # these could be noncontiguous
                # 
                
        # group
        pass

    def _spread_driver_revised(self):
        # get a line
        # get a unit
            # unit should have sheet on it)
            # unit.xl.sheet
            # or book? and find sheet by name?
                # pointer to sheet should stay constant during extrapolation?
                # 
                
            # sheet should have current column highlighted or somehow active

        # find the parameters range on the sheet
        # populate

        ## should run once, when driver goes to work

        ## Basically, this whole routine of saving to Excel is predicated on
        ## building it once, then not building it again (ie setting .xl.book to None)

        ## so top-level control would:
            # reset financials on company
            # run fill out with the book
            # take a pointer to the book
            # clear the .xl storage
            # return the book

        ## now, just decide how we are going to access the book
        ## for purposes of derive just need a sheet really

        ## unit.xl.book = point to workbook
        ## unit.xl.sheet = point to unit sheet
        ## line.xl.rows.consolidating.cells =
        ## line.xl.rows.derived
        ## line.xl.rows.starting
        ## line.xl.rows.ending
    
    def _spread_formula(self):
        # let's assume this routine runs in the formula
        # give the formula a sheet with all of the info it expects in the standard format
        # 
        # give it an active cell
            # or, more specifically, an active column
            #
        # output should be a xl formula in a cell that points to all the right places
        #

        ## assume that you cannot add data at the formula level.
        ## formula can write intermediate steps in its own discretiojn
        ## should return sheet and active cell coordinates
        ## has to be part of the definition module
    
    def _add_financials(self, sheet, unit, column):

        # go down every line
        # go through
            # formulas should deliver xl_formula inputs as range, something, something
            # basically need to map attributes to ranges
            #
            # if you have a very complex calculation (ie, sort units by some feature and tag, then do x and y to them),
            # may want to just enter the answer to that with a memo about the code (logic)
            #
            # or could just enter the value as fixed with an explanation (and comment like "Use blackbird!")
            #
            #
            # or could require a local runtime for some of these things, in which case can actually route through
            # the python routine?
            #
            #
            # could i have the formulas map themselves?
                # driver.spread()
                    #spreads the parameters
                    #then 
                # give them a sheet
                # 
            # 
            
        
        # Step 2: we link the periodic values.

        # To link:
            # Need a formula        

        # go by column in tl_sheet
        # 
        # copy each period heading
        # copy each period params
        #
        # on sheet, need to populate lookup tables too
        
        ## may be i could do this with a named range?? but doesnt do too much good
        ##

    # also need to connect balance sheets (time-based ops)
    def _driver_spread(self):
        # spread the parameters
            # ! already have the params_to_inputs dictionary
            # basically go through that
                # add the input line
                # link the input line to a parameter
                    # first to shared ones
                    # second overlay local ones

                    # should look like:
                        # cows:     "=d14"
                        # bulls:    "=d18"
                        # chickens: "8" {source then overwrite}
                        # etc:      "8"
                        #
                        # subtotal: "A1 * A2 + A4"

            # hold its own map in memory?
                # can use a Range() but just not put it on the sheet
            # 
                    
        # add the formula
            # need to quickly map the attribute structure

        # group the lines
        
        # change the row value and run _update_endpoints()
        
    # notes on consolidate:
        # for the purposes of excel, can set parent to sum of children
            #and not worry so much about showing child-level detail

        # Alternatively:

        # for every line in child spread:
            ## since we spread the child, its flat, not recursive
            ## 
        
            # if line is already in parent
                # add to line

            # else:
                # copy line
            
        # for child in children:
            # for line in child.financials:
                ## child.financials is now flat

                # if line_name in parent.financials:
                    #

        ## conceptually:
            # for every line, make a list of units where it exists and the locations of the value
            # then go through that list and add a formula to the consolidated cell of parent that adds all of the children
                # try it on 5 stores

            # may be able to assume that lines dont change from period to period...

            # may be can replace ugly formula with cleaner Excel logic
                # or could spread vertically!!!!!! <--------------------------------------------------------------------------------EUREKA!
                # for every child, take their value
                    # copy it to a cell
                    # then sum all of those cells
                    # to get the consolidated value!!!!

                # will make for very long spreadsheets, but whatever
                    # problem is that you have to repeat for every line
                    # have to make sure you group

                    # can toggle format (consolidate_in_line = False)
                        # if consolidate_in_line=True, use long formulas
                        # otherwise, go by row
                        # can even toggle number of references: ie consolidate every 5 units into a cell, every 10
                            # number of units per cell is the control                        

                # can simplify the spreadsheets by adding templates and explicit links between the templates and everything else
                
                # and can then do the derive elements

        ##

        # consolidate: one range
        # derive: a second range
        # finals: a third range

        # add columns 1,2,3... that show ranges
            # specify them
    
        
        # 
        ##         

    def _spread_parent_unit(self, unit, book):
        pass

        # should populate relationships
    
    def _spread_driver_outline(self):
        """

        -> None

        column: point to column where we are doing this
        row: point to row where we starting writing

        options: can write to full tab, from start to end?
          probably should
          all in the same row

        let's start by writing column by column

        steps:
        1. if driver data:
          put driver data into rows, in order
            question: do we label these rows on the labels column?
            probably should

            #assemble driver-level row index that includes the driver-specific
            data, but DO NOT enter that data into the main sheet-level look-up
            table. that data is for the driver only. 

        2. convert formula to run on rows
            formula needs to return (i) a string with formula, and (ii) a
            list of all the fields

            if conversion logic cannot locate a field, should raise
            error

        3.  write the formula into a subtotal row
            #if the formula increments or changes a prior subtotal row, need to
            find it
            #can keep overwriting the subtotal for the line
              #so by default subtotal is zero
              #then we add subtotals, but "subtotal" always points to the last
              #row only

              #
        4. group the drivers into their own little thing
            #

        5. [Line level]:
            # header should be smtg like "{name} analysis:". header should be
            # grouped with all of the drivers and hidden.

        6. [Line level]: a little bit of a pain with details
            # can change the derive() algo to go through lines detail by detail, recursively
            # then can add the lines incrementally?
            # challenge: it is hard to insert rows and columns in openpyxl
               #check if that's the case
               #create methods if necessary: insert_rows(), insert_columns()

            # need to start with consolidate() though
              so line algo would be:
                 # go through line by line
                 # for each line:
                    # add a consolidation row
                    # then add driver details
                    # if at the conclusion of work row val doesnt equal bb val, raise error
                      # (but hard to do that, because won't evaluate formulas)
                      # (check how that works)
                 

        """
        pass

    
        

    
    
