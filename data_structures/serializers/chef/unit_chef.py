# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.unit_chef
"""

Module defines a class that represents arbitrarily rich BusinessUnit instances
as a collection of linked Excel worksheets.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
UnitChef              class containing methods to chop BusinessUnits into
                      dynamic Excel structures
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from .cell_styles import CellStyles
from .line_chef import LineChef
from .sheet_style import SheetStyle
from .unit_fins_chef import UnitFinsChef
from .unit_info_chef import UnitInfoChef

from chef_settings import (
    APPLY_COLOR_TO_DECEMBER, DECEMBER_COLOR, VALUATION_TAB_COLOR
)
from openpyxl.styles import PatternFill




# Constants
# n/a

# Module Globals
_INVALID_CHARS = r"\/*[]:?"
# Excel forbids the use of these in sheet names.
REPLACEMENT_CHAR = None
# When the replacement character is None, UnitChef will remove all bad chars
# from sheet titles.
bad_char_table = {ord(c): REPLACEMENT_CHAR for c in _INVALID_CHARS}
get_column_letter = xlio.utils.get_column_letter


# Classes
class UnitChef:
    """

    One tab per unit
    Children first
    Arbitrarily recursive (though should max out at sheet limit; alternatively,
    book should prohibit new sheets after that? or can have 2 limits: a soft
    limit where ModelChopper shifts into different representation mode, and a
    hard limit, where you just cant create any more sheets.)

    Most non-public methods force keyword-based arg entry to avoid potentially
    confusing erros (switching rows for columns, etc.)

    Methods generally leave current row pointing to their last completed
    (filled) row.

    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    model                 obj; instance of Blackbird model
    timeline              obj; instance of Timeline from which to pull financials

    FUNCTIONS:
    chop_company()        returns sheet with a SheetData instance at sheet.bb,
                          also spreads financials, life, parameters of units
    chop_multi()          recursor for chop_company to chop child units
    chop_multi_valuation() returns sheet with a SheetData instance at sheet.bb,
                          makes and fills Valuation tab
    ====================  =====================================================
    """
    def __init__(self, model, timeline=None, include_ids=False):
        self.model = model
        self.include_ids = include_ids

        if timeline is not None:
            self.timeline = timeline
        else:
            self.timeline = model.get_timeline(resolution='monthly')

    def chop_multi(self, book, unit=None, values_only=False, tab_name='',
                   tab_color='', index=None):
        """


        UnitChef.chop_multi() -> Worksheet

        --``book`` must be a Workbook
        --``unit`` is optionally an instance of BusinessUnit
        --``timeline`` is optionally the timeline from which to pull financials
        --``values_only`` must be a bool, whether all values should be written
                          as hardcoded values (don't print drivers and life)

        Method recursively walks through ``unit`` and components and chops them
        into Excel format.  Method also spreads financials, parameters, and
        life of the unit.

        If `values_only` = True, only chop company unit (no recursion) and
        write all financial values as hardcoded.  This option is used for
        reports.
        """
        model = self.model
        timeline = self.timeline

        # top level entry: get the company
        if not unit:
            unit = model.get_company()

        # 1.   Chop the children
        if index is None:
            if values_only:
                index = len(book.worksheets)-1
            else:
                index = len(book.worksheets) - book.original_tab_count

        children = unit.components.get_ordered()

        for child in children:
            self.chop_multi(book, child, values_only=values_only,
                            tab_color=tab_color)

        # 2.   Chop the parent
        #
        # In this block, we have to set the current row to the place where the
        # last area ends before running the breadth-wise routines. Otherwise,
        # they would treat the ending row of the last period as their own
        # starting row, and the spreadsheet would look like a staircase.

        # 2.1.   set up the unit sheet and spread params
        info_chef = UnitInfoChef(model, timeline)
        sheet = info_chef.create_unit_sheet(
            book=book, unit=unit, index=index, values_only=values_only,
            name=tab_name, tab_color=tab_color,
        )

        if not values_only:
            # 2.2.   spread life
            info_chef.unit_life(sheet, unit)

            # 2.3. add unit size
            info_chef.add_unit_size(sheet, unit)

        # 2.4.  spread fins
        fins_chef = UnitFinsChef(model, timeline)
        fins_chef.chop_financials(sheet, unit, values_only=values_only,
                                  include_ids=self.include_ids)

        # 2.5 add area and statement labels and sheet formatting
        SheetStyle.style_sheet(sheet)

        if not values_only:
            # 2.6 add selector cell
            #   Make scenario label cells
            info_chef.add_scenario_selector_logic(book, sheet)

        # Color the December columns
        if APPLY_COLOR_TO_DECEMBER and not values_only:
            for date, column in sheet.bb.time_line.columns.by_name.items():
                if date.month == 12:
                    for row in range(1, sheet.max_row + 1):
                        cell = sheet.cell(column=column, row=row)
                        cell.fill = PatternFill(start_color=DECEMBER_COLOR,
                                                end_color=DECEMBER_COLOR,
                                                fill_type='solid')

        body_rows = sheet.bb.row_axis.get_group('body')
        label_col = sheet.bb.col_axis.get_group('head')
        self.add_labels(sheet, body_rows.groups, label_col)

        return sheet

    def chop_multi_valuation(
        self, model, book, unit=None, index=2, recur=False
    ):
        """


        UnitChef.chop_multi_valuation() -> Worksheet

        --``book`` must be a Workbook
        --``unit`` must be an instance of BusinessUnit
        --``index`` is optionally the index at which to insert the tab
        --``recur`` bool; whether to chop valuation for lower units

        Method recursively walks through ``unit`` and components and will chop
        their valuation if any exists.
        """
        # top level entry: get unit from model
        if not unit:
            unit = model.get_company()

        # 1.   Chop the children
        if recur:
            children = unit.components.get_ordered()

            for child in children:
                index = book.get_index(child.xl_data.sheet) - 1
                self.chop_multi_valuation(
                    model, book, unit=child, index=index, recur=recur
                )

        # 2.6 add valuation tab, if any exists for unit
        time_line = model.get_timeline()
        now = time_line.current_period
        financials = model.get_financials(unit.id.bbid, now)
        if financials.has_valuation:
            # fins_chef = UnitFinsChef(model)
            self.add_valuation_tab(book, unit, index=index)

    def add_valuation_tab(self, book, unit, index=None):
        """


        UnitChef._add_valuation_tab() -> Worksheet

        --``book`` must be a Workbook
        --``unit`` must be an instance of BusinessUnit
        --``index`` is optionally the index at which to insert the tab

        Method creates a valuation tab and chops unit valuation statement.
        """

        # 1.0   set up the unit sheet and spread params
        if not index:
            index = len(book.worksheets)

        if index == 2:
            name = "Valuation"
        else:
            name = unit.name + ' val'

        info_chef = UnitInfoChef(self.model, self.timeline)
        sheet = info_chef.create_unit_sheet(
            book=book, unit=unit, index=index, name=name, current_only=True
        )
        sheet.bb.outline_level += 1

        # 1.1   set-up life
        info_chef.unit_life(sheet, unit, current_only=True)

        # 1.2  Add Valuation statement
        time_line = self.model.get_timeline()
        now = time_line.current_period
        current = sheet.bb.time_line.columns.get_position(now.end)
        SheetStyle.set_column_width(sheet, current, width=22)

        fins_chef = UnitFinsChef(self.model, self.timeline)
        financials = self.model.get_financials(unit.id.bbid, now)
        statement = financials.valuation
        statement_rows = fins_chef.add_statement_container(sheet, statement)

        line_chef = LineChef(include_ids=self.include_ids)
        line_chef.chop_statement(
            sheet=sheet,
            column=current,
            statement=statement,
            row_container=statement_rows
        )

        # 1.5 add area and statement labels and sheet formatting
        SheetStyle.style_sheet(sheet)

        # 1.6 add selector cell
        #   Make scenario label cells
        info_chef = UnitInfoChef(self.model, self.timeline)
        info_chef.add_scenario_selector_logic(book, sheet)

        sheet.sheet_properties.tabColor = VALUATION_TAB_COLOR

        body_rows = sheet.bb.row_axis.get_group('body')
        label_col = sheet.bb.col_axis.get_group('head')
        self.add_labels(sheet, body_rows.groups, label_col)

        return sheet

    def add_labels(self, sheet, groups, label_col):
        """


        UnitChef.add_labels() -> None

        Writes row labels on sheet. To show up on the axis, a group
        1. should have no subgroups
        2. should have a label
        To add a title row for a group with subgroups, create a one-row
        'title' subgroup.
        """
        for group in groups:
            if group.groups:
                self.add_labels(sheet, group.groups, label_col)
            elif group.size:
                label = group.extra.get('label')
                if label:
                    row = group.number()
                    col = label_col.number()
                    rank = group.extra.get('rank')
                    if group.name == 'title' and rank == 1:
                        formatter = CellStyles.format_area_label
                        formatter(sheet, label, row, col_num=col)
                    else:
                        label_cell = sheet.cell(row=row, column=col + 1)
                        label_cell.value = label
                        formatter = group.extra.get('formatter')
                        if formatter:
                            formatter(label_cell)
                        if group.outline:
                            r = sheet.row_dimensions[row]
                            r.outline_level = group.outline

                        if self.include_ids:
                            label_cell = sheet.cell(row=row, column=col + 2)
                            label_cell.value = group.extra.get('bbid')

            if group.extra.get('hidden'):
                for rownum in range(
                    group.number(),
                    group.number() + group.size + 1
                ):
                    row = sheet.row_dimensions[rownum]
                    row.hidden = True
