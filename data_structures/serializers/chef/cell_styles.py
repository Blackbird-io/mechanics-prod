# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.cell_styles
"""

Module defines a class that stores standard row and column names.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
CellStyles            standard styles for worksheet cells
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.styles.colors import WHITE, BLACK

from chef_settings import AREA_BORDER

from .data_types import NumberFormats
from .data_types import TypeCodes
from .field_names import FieldNames





# Constants
HARDCODED_COLOR = '0070c0'  # '1f5993'
CALCULATION_COLOR = '707070'
SUBHEADER_COLOR = '808080'
LOWHEADER_COLOR = 'CCCCCC'

# Module Globals
get_column_letter = xlio.utils.get_column_letter

# Classes
class CellStyles:
    """

    Class stores standard keys for row and column lookups within SheetData.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    format_area_label()   format label cell and row for sheet Areas
    format_border_group()  add thin border around group of cells
    format_calculation()  format text in cells containing calculations
    format_consolidated_label() consolidated rows in italic
    format_date()         format number in cells containing dates
    format_hardcoded()    format font in hardcoded cells
    format_header_label() white on black
    format_integer()      format number as integer
    format_line()         format text in line item cells
    format_line_borders() adds borders to rows where lineitems request them
    format_parameter()    format number in parameter cells
    format_scenario_label()  format scenario label cells (black with white text)
    format_scenario_selector_cell() format scenario selector cells
    format_sub_header_label() formats sub-header cells
    ====================  =====================================================
    """
    @staticmethod
    def format_area_label(sheet, label, row_num, col_num=None):
        """


        CellStyles.format_area_label -> None

        --``sheet`` is an instance of openpyxl.worksheet
        --``label`` is a the string name of the area to label
        --``row_num`` is the row number where the label should be inserted
        --``col_num`` is the column number where the labels should be

        Format area/statement label and dividing border
        """
        side = Side(border_style='double')

        if col_num:
            col = get_column_letter(col_num)
        else:
            col = 'A'

        cell_cos = col + str(row_num)
        cell = sheet[cell_cos]
        cell.font = Font(bold=True)
        cell.set_explicit_value(label.title(),
                                data_type=TypeCodes.FORMULA_CACHE_STRING)

        if AREA_BORDER:
            rows = sheet.iter_rows(row_offset=row_num - 1)
            row = rows.__next__()
            for cell in row:
                border = Border(top=cell.border.top)
                border.top = side
                cell.border = border

    @staticmethod
    def format_calculation(cell):
        """


        CellStyles.format_calculation -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing calculations to conform with Chef standard.
        """
        # font = Font(italic=True, color=CALCULATION_COLOR)
        # cell.font = font

        cell.number_format = NumberFormats.DEFAULT_PARAMETER_FORMAT

    @staticmethod
    def format_date(cell):
        """


        CellStyles.format_date -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing dates to conform with Chef standard.
        """

        cell.number_format = NumberFormats.DEFAULT_DATE_FORMAT

    @staticmethod
    def format_hardcoded(cell):
        """


        CellStyles.format_hardcoded -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing hardcoded values to conform with Chef standard.
        """
        font = Font(color=HARDCODED_COLOR, bold=True)
        cell.font = font

    @staticmethod
    def format_integer(cell):
        """


        CellStyles.format_integer -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing integers to conform with Chef standard.
        """
        cell.number_format = NumberFormats.INTEGER_FORMAT

    @staticmethod
    def format_line(line):
        """


        CellStyles.format_line -> None

        --``line`` is an instance of LineItem

        Format cells written to for the provided line to conform with Chef
        standard.
        """

        use_format = line.xl.number_format or NumberFormats.DEFAULT_LINE_FORMAT

        if line.xl.derived.cell:
            line.xl.derived.cell.number_format = use_format

        if line.xl.consolidated.cell:
            line.xl.consolidated.cell.number_format = use_format

        for cell in line.xl.consolidated.array:
            cell.number_format = use_format

        if line.xl.detailed.cell:
            line.xl.detailed.cell.number_format = use_format

        if line.xl.cell:
            line.xl.cell.number_format = use_format

        if line.xl.format.font_format:
            line.xl.cell.font = Font(**line.xl.format.font_format)

    @staticmethod
    def format_parameter(cell):
        """


        CellStyles.format_parameter -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing parameters to conform with Chef standard.
        """
        cell.number_format = NumberFormats.DEFAULT_PARAMETER_FORMAT
        cell.alignment = Alignment(horizontal='right')

    @staticmethod
    def format_scenario_label(cell):
        """


        CellStyles.format_scenario_label -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells containing scenario column labels on Scenario tab.
        """
        CellStyles.format_header_label(cell, alignment='center')

    @staticmethod
    def format_header_label(cell, alignment=None, color=BLACK,
                            font_color=WHITE, bold=True, font_size=11,
                            name='Calibri'):
        """


        CellStyles.format_header_label -> None

        --``cell`` is an instance of openpyxl cell class
        --``alignment`` is the text alignment within the cell
        --``color`` is the cell color, default is black (hex definition)
        --``font_color`` is the font color, default is white (hex definition)
        --``bold`` is a bool, whether or not to bold the font

        Headers at the top, white on black.
        """
        if alignment:
            cell.alignment = Alignment(horizontal=alignment)

        cell.font = Font(color=font_color, bold=bold, size=font_size, name=name)
        cell.fill = PatternFill(start_color=color,
                                end_color=color,
                                fill_type='solid')

    @staticmethod
    def format_subheader_label(cell, alignment=None, color=SUBHEADER_COLOR):
        """


        CellStyles.format_subheader_label -> None

        --``cell`` is an instance of openpyxl cell class
        --``alignment`` is the text alignment within the cell
        --``color`` is the cell color, default defined by SUBHEADER_COLOR

        Looks like Header, but on a different background color.
        """
        CellStyles.format_header_label(cell, alignment, color)

    @staticmethod
    def format_consolidated_label(cell):
        """


        CellStyles.format_consolidated_labelr -> None

        --``cell`` is an instance of openpyxl cell class

        Format cells which are labels for consolidating logic.
        """
        cell.font = Font(italic=True)

    @staticmethod
    def format_scenario_selector_cells(sheet, label_col, selector_col, row,
                                       active=True):
        """


        CellStyles.format_scenario_selector_cells -> None

        --``sheet`` is an instance of openpyxl worksheet class
        --``label_col`` is the column number where the scenario selector label
            lives
        --``selector_col`` is the column number where the scenario selector
            cell lives
        --``row`` is the row where the scenario selector cells live
        --``active`` is a bool, indicates whether scenario selector is live

        Format cells containing scenario selection label and selector.
        """
        side = Side(border_style='thick')

        # Left label cell
        left_cell = sheet.cell(column=label_col, row=row)
        border = Border(left=left_cell.border.left,
                        top=left_cell.border.top,
                        bottom=left_cell.border.bottom)
        border.left = side
        border.top = side
        border.bottom = side
        left_cell.border = border
        left_cell.font = Font(bold=True, size=14)
        left_cell.value = "Active Scenario:"

        # Blank middle cell
        blank_cell = sheet.cell(column=label_col + 1, row=row)
        border = Border(top=blank_cell.border.top,
                        bottom=blank_cell.border.bottom)
        border.top = side
        border.bottom = side
        blank_cell.border = border

        # Rightmost cell where selector lives
        right_cell = sheet.cell(column=selector_col, row=row)
        border = Border(right=right_cell.border.right,
                        top=right_cell.border.top,
                        bottom=right_cell.border.bottom)
        border.right = side
        border.top = side
        border.bottom = side
        right_cell.border = border
        right_cell.alignment = Alignment(horizontal='center')

        if active:
            color = HARDCODED_COLOR
        else:
            color = BLACK

        font = Font(color=color, bold=True, size=14)
        right_cell.font = font

    @staticmethod
    def format_border_group(sheet, st_col, ed_col, st_row, ed_row,
                            border_style='thin'):
        """


        CellStyles.format_border_group -> None

        --``sheet`` is an instance of openpyxl worksheet class
        --``st_col`` is the starting column in the group
        --``ed_col`` is the ending column in the group
        --``st_row`` is the starting row in the group
        --``ed_row`` is the ending row in the group
        --``border_style`` is any openpyxl border style, default is `thin`

        Draws thin border around group of cells defined by starting and ending
        rows and columns.
        """
        side = Side(border_style=border_style)

        # SET TOP BORDER
        row = st_row
        for c in range(st_col, ed_col + 1):
            cell = sheet.cell(column=c, row=row)
            border = Border(top=cell.border.top)
            border.top = side
            cell.border = border

        # SET LEFT BORDER
        col = st_col
        for r in range(st_row, ed_row + 1):
            cell = sheet.cell(column=col, row=r)
            border = Border(left=cell.border.left)
            border.left = side
            cell.border = border

        # SET RIGHT BORDER
        col = ed_col
        if st_col != ed_col:
            for r in range(st_row, ed_row + 1):
                cell = sheet.cell(column=col, row=r)
                border = Border(right=cell.border.right)
                border.right = side
                cell.border = border
        else:
            for r in range(st_row, ed_row + 1):
                cell = sheet.cell(column=col, row=r)
                border = Border(right=cell.border.right,
                                left=cell.border.left)
                border.left = side
                border.right = side
                cell.border = border

        # SET TOP AND BOTTOM BORDERS
        row = ed_row
        if st_row != ed_row:
            for c in range(st_col, ed_col + 1):
                cell = sheet.cell(column=c, row=row)
                border = Border(bottom=cell.border.bottom,
                                left=cell.border.left,
                                top=cell.border.top,
                                right=cell.border.right)
                border.bottom = side
                cell.border = border
        else:
            for c in range(st_col, ed_col + 1):
                cell = sheet.cell(column=c, row=row)
                border = Border(bottom=cell.border.bottom,
                                left=cell.border.left,
                                top=cell.border.top,
                                right=cell.border.right)
                border.top = side
                border.bottom = side
                cell.border = border

        if st_col != ed_col:
            # SET UPPER-LEFT CORNER BORDER
            cell = sheet.cell(column=st_col, row=st_row)
            border = Border(bottom=cell.border.bottom, left=cell.border.left,
                            top=cell.border.top, right=cell.border.right)
            border.top = side
            border.left = side
            cell.border = border

            # SET UPPER-RIGHT CORNER BORDER
            cell = sheet.cell(column=ed_col, row=st_row)
            border = Border(bottom=cell.border.bottom, left=cell.border.left,
                            top=cell.border.top, right=cell.border.right)
            border.top = side
            border.right = side
            cell.border = border

            # SET LOWER-LEFT CORNER BORDER
            cell = sheet.cell(column=st_col, row=ed_row)
            border = Border(bottom=cell.border.bottom, left=cell.border.left,
                            top=cell.border.top, right=cell.border.right)
            border.bottom = side
            border.left = side
            cell.border = border

            # SET LOWER-RIGHT CORNER BORDER
            cell = sheet.cell(column=ed_col, row=ed_row)
            border = Border(bottom=cell.border.bottom, left=cell.border.left,
                            top=cell.border.top, right=cell.border.right)
            border.bottom = side
            border.right = side
            cell.border = border

        else:
            # SET TOP CELL BORDER
            cell = sheet.cell(column=st_col, row=st_row)
            border = Border(top=cell.border.top,
                            bottom=cell.border.bottom,
                            left=cell.border.left,
                            right=cell.border.right)
            border.top = side
            border.left = side
            border.right = side
            cell.border = border

            # SET BOTTOM CELL BORDER
            cell = sheet.cell(column=st_col, row=ed_row)
            border = Border(top=cell.border.top,
                            bottom=cell.border.bottom,
                            left=cell.border.left,
                            right=cell.border.right)
            border.bottom = side
            border.left = side
            border.right = side
            cell.border = border

    @staticmethod
    def format_line_borders(book):
        """


        CellStyles.format_line_borders() -> None

        --``book`` is the workbook to format

        Function manipulates workbook in place, adding borders to rows where
        line items request them.
        """
        for sheet in book.worksheets:

            if not getattr(sheet, 'bb', None):
                continue

            if not sheet.bb.line_directory:
                continue

            # go through all lines in sheet.bb.line_directory and add border
            # formatting, if any
            param_area = getattr(sheet.bb, FieldNames.PARAMETERS)
            st_col = param_area.columns.by_name[FieldNames.VALUES]
            if 'time_line' in sheet.bb.area_names:
                ed_col = min((sheet.bb.time_line.columns.ending,
                              sheet.max_column))
            else:
                ed_col = sheet.max_column - 1

            for xl in sheet.bb.line_directory.values():
                row = xl.cell.row
                border = xl.format.border

                if not border:
                    continue

                CellStyles.format_border_group(sheet, st_col=st_col,
                                               ed_col=ed_col, st_row=row,
                                               ed_row=row, border_style=border)

    @staticmethod
    def format_bold(cell):
        cell.font = Font(bold=True)

    @staticmethod
    def format_alignment(cell, alignment):
        cell.alignment = Alignment(horizontal=alignment)

    @staticmethod
    def format_border(cell, top=False, bottom=False, left=False, right=False,
                      border_style='thin'):

        side = Side(border_style=border_style)
        # no_side = Side(border_style=None)

        border = Border(top=cell.border.top,
                        bottom=cell.border.bottom,
                        left=cell.border.left,
                        right=cell.border.right)

        if top:
            border.top = side

        if bottom:
            border.bottom = side

        if left:
            border.left = side

        if right:
            border.right = side

        cell.border = border
