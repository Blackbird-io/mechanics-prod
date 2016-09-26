# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.transcript_chef
"""

Class for creating the transcript tab in Blackbird workbooks.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
N/A

FUNCTIONS:
n/a

CLASSES:
TranscriptChef         add transcript tab to Blackbird model Excel workbook
====================  =========================================================
"""




# Imports
import chef_settings
import datetime
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import WHITE, BLACK
import string

from .bb_workbook import BB_Workbook as Workbook
from .cell_styles import CellStyles
from .sheet_style import SheetStyle
from .tab_names import TabNames




# Module Globals
cell_styles = CellStyles()
sheet_style = SheetStyle()
tab_names = TabNames()

# Constants
ADDABLE_CAPTION = ' (addable input)'
NOTE_CAP = 'Note'
TABLE_CAPTION = 'table entry'

USAGE_NOTE = 'To ensure compatibility with website script functionality, ' \
             'adjust only values in "Response" column. Within "Response" ' \
             'column, maintain given formatting for ranges, dates, etc.'

IA = 'input_array'
IST = 'input_sub_type'
IT = 'input_type'
LV = 'line_value'
MC = 'main_caption'
NAME = 'name'
PROM = 'prompt'
RA = 'response_array'
RS = 'response'
TARG = 'target'



# Classes
class TranscriptChef:
    """

    Class packages an Engine model into an Excel workbook with dynamic links.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    make_transcript_tab() adds transcript tab to Blackbird model workbook
    ====================  =====================================================
    """

    PROMPT_HEADER = 'Question'
    CAPTION_HEADER = 'Element Caption'
    TARGET_HEADER = 'Target Unit'
    RESPONSE_HEADER = 'Response'
    QUESTION_NAME_HEADER = 'Question Identifier'

    COLUMN_DICT = dict()
    COLUMN_DICT[PROMPT_HEADER] = 'C'
    COLUMN_DICT[CAPTION_HEADER] = 'D'
    COLUMN_DICT[TARGET_HEADER] = 'E'
    COLUMN_DICT[RESPONSE_HEADER] = 'F'
    COLUMN_DICT[QUESTION_NAME_HEADER] = 'G'

    def make_transcript_excel(self, model, book=None, idx=1):
        """


       TranscriptChef.make_transcript_tab() -> None

        --``model`` Blackbird model object
        --``book`` working Excel book where tab should be added
        --``idx`` is the index where tab should be added

        Function generates and excel transcript of the questiosn/responses from the
        analysis' interview.
        """

        if not book:
            book = Workbook()
            sheet = book.active

            date = datetime.date.today()
            date_str = '%s-%s-%s' % (date.month, date.day, date.year)
            title = model.name + ' ' + date_str

            sheet.title = title
        else:
            sheet = book.create_sheet(name=tab_names.TRANSCRIPT, index=idx)

        # get length of interview
        transcript = list()
        for i in model.transcript:
            q = i[0]['q_in']
            if q:
                q['response_array'] = i[0]['r_in']
                transcript.append(q)

        analysis_name = model.name.title()
        sheet = self._prep_output_excel(sheet, analysis_name)

        current_row = 8
        for q in transcript:
            current_row = self._prep_question_output(sheet, q, current_row)

        for col in self.COLUMN_DICT.values():
            num_col = string.ascii_uppercase.find(col.upper()) + 1
            cell_styles.format_border_group(sheet=sheet,
                                            st_col=num_col,
                                            ed_col=num_col,
                                            st_row=8,
                                            ed_row=current_row - 1,
                                            border_style='dashed')

        self._add_note_to_excel(sheet, current_row + 1)

        sheet_style.style_sheet(sheet, label_areas=False)
        sheet.sheet_properties.tabColor = chef_settings.TRANSCRIPT_TAB_COLOR

        return book

    # ************************************************************************#
    #                         NON-PUBLIC METHODS                              #
    # ************************************************************************#

    @staticmethod
    def _add_note_to_excel(sheet, row):
        # set label cell
        cell = sheet['A' + str(row)]
        cell.value = 'Note:'
        cell.font = Font(bold=True)

        # set note
        cell = sheet['C' + str(row)]
        cell.value = USAGE_NOTE

    def _add_record_to_excel(self, sheet, answer_dict, row):
        for key, column in self.COLUMN_DICT.items():
            cell = sheet[column + str(row)]
            cell.value = answer_dict[key]
            cell.alignment = Alignment(horizontal='left')

    @staticmethod
    def _format_response(response):
        response_out = response[RS][0]
        input_type = response[IT]

        if response_out is None and input_type == 'bool':
            response_out = False

        if isinstance(response_out, datetime.date):
            response_out = response_out.strftime('%Y-%m-%d')

        # range
        if isinstance(response_out, list):
            response_out = 'Min: %s; Max: %s' % tuple(response_out)

        return response_out

    def _prep_output_excel(self, sheet, name):
        title_txt = 'INTERVIEW TRANSCRIPT'

        column_widths = dict()
        column_widths['A'] = 9  # Labels
        column_widths['B'] = 4
        column_widths['C'] = 74  # Question prompt column
        column_widths['D'] = 44  # Target BU name
        column_widths['E'] = 44  # Main Caption/Element Caption column
        column_widths['F'] = 44  # Response column
        column_widths['G'] = 40  # Question name/identifier, HIDE THIS COLUMN

        date = datetime.date.today()
        date_str = '%s-%s-%s' % (date.month, date.day, date.year)

        for col, wid in column_widths.items():
            sheet.column_dimensions[col].width = wid

        # set sheet title cell
        cell = sheet['D2']
        cell.value = title_txt
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

        # set company name cell and label
        cell = sheet['A4']
        cell.value = 'Company:'
        cell.font = Font(bold=True)

        cell = sheet[chef_settings.TRANS_COMP_CELL]
        cell.value = name

        # set date cell and label
        cell = sheet['A5']
        cell.value = 'Date:'
        cell.font = Font(bold=True)

        cell = sheet[chef_settings.TRANS_DATE_CELL]
        cell.value = date_str

        for title, col in self.COLUMN_DICT.items():
            cos = col + '7'
            cell = sheet[cos]
            cell.value = title
            cell.font = Font(color=WHITE, bold=True)
            cell.fill = PatternFill(start_color=BLACK,
                                    end_color=BLACK,
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='left')

        sheet.freeze_panes = 'C8'

        sheet.column_dimensions['G'].hidden = True

        return sheet
    
    def _prep_question_output(self, sheet, q, row):
        try:
            target = q[TARG]
        except KeyError:
            target = None
            prompt = q[PROM]
            name = None
        else:
            name = q[NAME]
            prompt = q[PROM]
    
        response_array = q[RA]
    
        if q[IT] == 'end':
            return row
    
        if not response_array:
            return row
    
        if q[IT] == 'table':
            # outer for loop loops through table rows
            for response in response_array:
                row = self._prep_response_output(response=response,
                                                 prompt=prompt,
                                                 target=target,
                                                 name=name,
                                                 current_row=row,
                                                 sheet=sheet,
                                                 table=True)
        else:
            for response in response_array:
                temp_inp = response[IT]
                temp_res = response[RS]
    
                if len(temp_res) > 1 and temp_inp != 'range':
                    # addable input
                    use_res = response.copy()
                    use_res[MC] += ADDABLE_CAPTION
                    for r in temp_res:
                        # use a copy of the response array and replace the response
                        # value for each value added by user so each gets its own
                        # row
                        use_res[RS] = [r]
                        row = self._prep_response_output(response=use_res,
                                                         prompt=prompt,
                                                         target=target,
                                                         name=name,
                                                         current_row=row,
                                                         sheet=sheet,
                                                         table=False)
                else:
                    # regular input
                    row = self._prep_response_output(response=response,
                                                     prompt=prompt,
                                                     target=target,
                                                     name=name,
                                                     current_row=row,
                                                     sheet=sheet,
                                                     table=False)

        try:
            notes = q['notes']
        except KeyError:
            # clp output, question does not have notes
            pass
        else:
            for note in notes:
                answer = self.COLUMN_DICT.copy()
                answer[self.PROMPT_HEADER] = prompt or ''
                answer[self.CAPTION_HEADER] = NOTE_CAP
                answer[self.TARGET_HEADER] = target or ''
                answer[self.RESPONSE_HEADER] = note or ''
                answer[self.QUESTION_NAME_HEADER] = name or ''
                self._add_record_to_excel(sheet, answer, row)
                row += 1
    
        return row

    def _prep_response_output(self, response, prompt, target, name,
                              current_row, sheet, table=False):
        answer = self.COLUMN_DICT.copy()
        answer[self.PROMPT_HEADER] = prompt or ''
        answer[self.TARGET_HEADER] = target or ''
        answer[self.QUESTION_NAME_HEADER] = name or ''
    
        # inner for loop loops over columns
        if table:
            response_array = response
            response_out = ''
            for response in response_array:
                caption = response[MC]
                temp_response = self._format_response(response)
                if temp_response is None:
                    continue

                response_out += '%s: %s; ' % (caption, temp_response)
                answer[self.CAPTION_HEADER] = TABLE_CAPTION
        else:
            caption = response[MC]
            response_out = self._format_response(response)
    
            if response_out is None:
                return current_row
    
            answer[self.CAPTION_HEADER] = caption or ''
    
        answer[self.RESPONSE_HEADER] = response_out
    
        self._add_record_to_excel(sheet, answer, current_row)
        current_row += 1
    
        return current_row
