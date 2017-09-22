# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.bb_workbook
"""

Module defines workbook with custom native-Python data storage on each sheet.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
BB_Workbook           workbook where each sheet has a SheetData record set
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from chef_settings import COLLAPSE_ROWS, DEFAULT_SCENARIOS, \
    SCENARIO_SELECTORS, SAVE_NAMED_RANGES

from ._chef_tools import check_filename_ext
from .data_management import SheetData
from .field_names import FieldNames
from .tab_names import TabNames

if COLLAPSE_ROWS or SCENARIO_SELECTORS:
    import pythoncom

    from ._chef_tools import add_links_to_selectors, collapse_groups




# Constants
# n/a

# Module Globals
# n/a


# Classes
class BB_Workbook(xlio.Workbook):
    """

    Class modifies standard workbook to include a SheetData record set on each
    sheet. As a result, we can do data lookups and reference builds faster and
    more explicitly.
    ====================  ======================================================
    Attribute             Description
    ====================  ======================================================

    DATA:
    scenario_names        None or list; holds names of scenarios in workbook

    FUNCTIONS:
    create_sheet()        returns sheet with a SheetData instance at sheet.bb
    save()                saves workbook to file and runs test on contents
    set_scenario_names()  sets scenario_names attribute
    ====================  ======================================================
    """
    def __init__(self, *pargs, **kwargs):
        xlio.Workbook.__init__(self, *pargs, **kwargs)
        self.scenario_names = None
        self.original_tab_count = 0
        self.drivers_tab_name = TabNames.SCENARIOS

    @staticmethod
    def convert(book):
        """


        BB_WorkBook.convert() -> BB_Workbook

        --``book`` must be a normal OpenPyXl workbook

        Return BB_Workbook version of book.
        """
        new_book = BB_Workbook()

        # Openpyxl version 2.4.8
        new_book._alignments = book._alignments
        new_book._borders = book._borders
        new_book._cell_styles = book._cell_styles
        new_book._colors = book._colors
        new_book._data_only = book._data_only
        new_book._differential_styles = book._differential_styles
        new_book._fills = book._fills
        new_book._fonts = book._fonts
        new_book._named_styles = book._named_styles
        new_book._number_formats = book._number_formats
        new_book._protections = book._protections
        new_book._read_only = book._read_only
        new_book._table_styles = book._table_styles
        new_book.loaded_theme = book.loaded_theme

        if SAVE_NAMED_RANGES:
            for nr in book.get_named_ranges():
                if nr.attr_text != '#REF!' and not nr.hidden:
                    new_book.add_named_range(nr)

        for sheet in book.worksheets:
            sheet._WorkbookChild__parent = new_book
            new_book._sheets.append(sheet)

        return new_book

    def create_sheet(self, name, index=None):
        """


        BB_WorkBook.create_sheet(name) -> Worksheet

        --``name`` must be string name for new worksheet
        --``index`` is the desired index at which to place the new worksheet
            within the workbook

        Return worksheet with a SheetData record set at instance.bb.

        """

        oname = name
        idx = 1

        lower_names = [n.lower().strip() for n in self.sheetnames]
        while name.lower().strip() in lower_names:
            name = oname + '_%s' % idx
            idx += 1
            print(name)

        sheet = xlio.Workbook.create_sheet(self, name, index=index)
        sheet.bb = SheetData()

        return sheet

    def save(self, filename):
        """


        BB_WorkBook.save(filename) -> None

        --``filename`` must be string path at which to save workbook (".xlsx")

        Method saves workbook to specified file and uses VBScript to prettily
        collapse row and column groups.
        """

        if COLLAPSE_ROWS or SCENARIO_SELECTORS:
            pythoncom.CoInitialize()

        # save workbook
        filename = check_filename_ext(filename, 'xlsx')
        xlio.Workbook.save(self, filename)

        if COLLAPSE_ROWS:
            collapse_groups(filename)

        if SCENARIO_SELECTORS:
            sources_dict = self._get_sources_dict()
            filename = add_links_to_selectors(filename, sources_dict)

        if COLLAPSE_ROWS or SCENARIO_SELECTORS:
            pythoncom.CoUninitialize()

        return filename

    def set_scenario_names(self, model):
        """


        BB_Workbook.set_scenario_names() -> None

        --``model`` must be a Blackbird Engine model

        Sets instance.scenario_names list.
        """
        self.scenario_names = [FieldNames.CUSTOM, FieldNames.BASE]
        self.scenario_names.extend(DEFAULT_SCENARIOS)

        for k in sorted(model.scenarios.keys()):
            if k not in self.scenario_names:
                self.scenario_names.append(k.title())

    # *************************************************************************#
    #                           NON-PUBLIC METHODS                             #
    # *************************************************************************#

    def _get_sources_dict(self):
        """


        BB_Workbook._get_sources_dict(self) -> dict

        Method compiles dictionary of worksheet names and cell addresses where
        scenario selector cells live.
        """
        sources_dict = dict()
        for sheet in self.worksheets:
            if getattr(sheet, 'bb', None):
                if sheet.bb.scenario_selector:
                    idx = int(self.get_index(sheet)) + 1
                    sheet_num = "Sheet%s" % idx
                    sources_dict[sheet_num] = (sheet.title,
                                               sheet.bb.scenario_selector)

        return sources_dict
