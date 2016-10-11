# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.summary_chef
"""

Class for creating summaries of Blackbird Engine models.

Chef modules write formulas to cells explicitly, using the set_explicit_value()
method, to make sure Excel interprets the strings correctly.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
SummaryChef           chop Blackbird Engine model into a dynamic Excel workbook
====================  =========================================================
"""




# Imports
import logging
import openpyxl as xlio

from openpyxl.styles import Font, Alignment

import bb_settings
import chef_settings

from .cell_styles import CellStyles, LOWHEADER_COLOR
from .field_names import FieldNames
from .sheet_style import SheetStyle
from .summary_line_chef import SummaryLineChef
from .tab_names import TabNames
from .transcript_chef import TranscriptChef
from .unit_chef import UnitChef




# Constants
# n/a

# Module Globals
line_chef = SummaryLineChef()
transcript_chef = TranscriptChef()
unit_chef = UnitChef()

get_column_letter = xlio.utils.get_column_letter
bounding_box = xlio.drawing.image.bounding_box
logger = logging.getLogger(bb_settings.LOGNAME_MAIN)


# Classes
class SummaryChef:
    """

    Class packages an Engine model into an Excel workbook with dynamic links.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    add_annual_summary()  adds a summary to book
    ====================  =====================================================
    """

    def add_annual_summary(self, model, book):
        """


        ModelChef._add_annual_summary -> None

        --``book`` is an instance ob BB_Workbook
        --``model`` is an instance of Blackbird Engine model

        Adds an annual summary tab.
        """
        # Create summary tab
        tab_idx = 1
        sheet = book.create_sheet(TabNames.SUMMARIES, tab_idx)

        # 2x1 top left corner for company name
        header_rows = sheet.bb.row_axis.add_group('tab_header', size=2)
        header_cols = sheet.bb.col_axis.add_group('tab_header', size=1)

        # Add company name, top left of header section
        company = model.get_company()
        address = header_rows.get_corner_address(header_cols)
        cell = sheet.cell(address)
        cell.value = company.title
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(size=14, bold=True, underline='single')

        # Output area, size to be calculated
        output_rows = sheet.bb.row_axis.add_group('output_rows')
        output_cols = sheet.bb.col_axis.add_group('output_cols')

        # Output column layout
        # blank column on the left
        colspacer_tip = output_cols.add_group('spacer_tip', size=1)
        # columns for row labels
        label_cols = output_cols.add_group('labels', size=5)
        # columns for years and quarters, size to be calculated
        years_cols = output_cols.add_group('years')
        # blank column at the end
        colspacer_end = output_cols.add_group('spacer_end', size=1)

        # Column headers
        time_line = model.get_timeline()
        self._annual_summary_headers(
            sheet, time_line, output_rows, output_cols
        )

        # Add parameters area to set up label and value columns
        area = sheet.bb.add_area(FieldNames.PARAMETERS)
        area.columns.by_name[FieldNames.LABELS] = label_cols.number()
        area.columns.by_name[FieldNames.VALUES] = years_cols.number()

        # Add row and label for Complete T/F
        complete_label_rows = output_rows.add_group(
            'complete_label',
            size=1,
            offset=1
        )
        address = complete_label_rows.get_corner_address(label_cols)
        cell = sheet.cell(address)
        cell.value = chef_settings.COMPLETE_LABEL
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Add row and label for Available Months
        available_months_rows = output_rows.add_group(
            'available_months',
            size=1,
            offset=0
        )
        address = available_months_rows.get_corner_address(label_cols)
        cell = sheet.cell(address)
        cell.value = chef_settings.AVAILABLE_LABEL
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Fill output: monthly summary timeline
        if chef_settings.SUMMARY_INCLUDES_MONTHS:
            # column selector: date -> years_cols.2017.quarters.1Q17.months.
            col_selector = lambda day: years_cols.get_group(
                day.year, 'quarters', self._quarter_name(day), 'months', day
            )
            # chop quarterly
            self._annual_summary_detail(
                sheet,
                time_line,
                output_rows,
                output_cols,
                col_selector=col_selector
            )

        # Fill output: quarterly summary timeline, may be excluded
        if chef_settings.SUMMARY_INCLUDES_QUARTERS:
            key = time_line.summary_builder.QUARTERLY_KEY
            qtr_timeline = model.get_timeline(key)
            # calculate the column layout for quarters and years
            # column selector: date -> years_cols.2017.quarters.1Q17
            col_selector = lambda day: years_cols.get_group(
                day.year, 'quarters', self._quarter_name(day), 'quarter'
            )
            # chop quarterly
            self._annual_summary_detail(
                sheet,
                qtr_timeline,
                output_rows,
                output_cols,
                col_selector=col_selector
            )

        # Fill output: annual summary timeline, always included
        key = time_line.summary_builder.ANNUAL_KEY
        sum_timeline = model.get_timeline(key)
        # column selector: date -> years_cols.2017.year
        col_selector = lambda day: years_cols.get_group(
            day.year, 'year'
        )
        self._annual_summary_detail(
            sheet,
            sum_timeline,
            output_rows,
            output_cols,
            col_selector=col_selector,
        )

        # blank row before bottom border
        output_rows.add_group('spacer_end', size=1)

        # calculate axis locations and resolve the cell addresses
        output_rows.calc_size()
        output_rows.resolve_cells()

        # Styling and formatting that's left
        SheetStyle.style_sheet(sheet, label_areas=False)

        # Make pretty border
        CellStyles.format_border_group(
            sheet=sheet,
            st_col=output_cols.number(),
            ed_col=output_cols.number() + output_cols.size - 1,
            st_row=output_rows.number(),
            ed_row=output_rows.number() + output_rows.size - 1,
            border_style='thin'
        )

        # Label financial statements
        statement_rowgroup = output_rows.get_group('statements')
        for statement_rows in statement_rowgroup.groups:
            self._annual_summary_labels(
                sheet, statement_rows.groups, label_cols
            )

        # Spacer column width, 2 cols on left and right of output
        for spacer in (colspacer_tip, colspacer_end):
            letter = get_column_letter(spacer.number())
            column = sheet.column_dimensions[letter]
            column.width = 4

        sheet.sheet_properties.tabColor = chef_settings.SUMMARY_TAB_COLOR

        corner_col = area.columns.by_name[FieldNames.VALUES] - 1
        corner_row = complete_label_rows.number() - 1
        corner_cell = sheet.cell(column=corner_col, row=corner_row)
        sheet.freeze_panes = corner_cell

    # *************************************************************************#
    #                           NON-PUBLIC METHODS                             #
    # *************************************************************************#

    def _annual_summary_labels(self, sheet, groups, label_cols, level=0):
        """


        ModelChef._annual_summary_labels() -> None

        Writes row labels on annual summary. To show up on the axis, a group
        1. should have no subgroups
        2. should have a label
        To add a title row for a group with subgroups, create a one-row
        'title' subgroup.
        """
        for group in groups:
            if group.groups:
                self._annual_summary_labels(
                    sheet, group.groups, label_cols, level=level + 1
                )
            elif group.size:
                label = group.extra.get('label')
                if label:
                    row = group.number()
                    col = label_cols.number()
                    if group.name == 'title' and level == 0:
                        formatter = CellStyles.format_area_label
                        formatter(sheet, label, row, col_num=col)
                    else:
                        label_cell = sheet.cell(row=row, column=col)
                        label_cell.value = label

    def _quarter_name(self, date):
        """


        ModelChef._quarter_name() -> str

        Convenience: datetime.date -> '1Q17'.
        """
        return '{}Q{:02d}'.format(1 + (date.month - 1) // 3, date.year % 100)

    def _year_headers(self, sheet, years_cols, year_headrow):
        """


        ModelChef._year_headers() -> None

        --``years_cols`` main column header container
        --``year_headrow`` header row, size 1

        Sets up column header layout at the intersection of ``year_headrow``
        and year_colgroup in ``years_cols``.
        """
        for year_colgroup in years_cols.groups:
            # set width of the column holding annual numbers
            year_col = year_colgroup.get_group('year')
            address = year_headrow.get_corner_address(year_col)
            cell = sheet.cell(address)
            column = sheet.column_dimensions[cell.column]
            column.width = chef_settings.COLUMN_WIDTH

            # set year label at the start of year block
            address = year_headrow.get_corner_address(year_colgroup)
            cell = sheet.cell(address)
            cell.value = year_colgroup.name
            CellStyles.format_header_label(cell, alignment='right')

            # merge header cells
            if year_colgroup.size > 1:
                stretch = year_headrow.get_range_address(year_colgroup)
                sheet.merge_cells(stretch)

    def _quarter_headers(self, sheet, years_cols, qtr_headrow):
        """


        ModelChef._quarter_headers() -> None

        --``years_cols`` main column header container
        --``qtr_headrow`` header row, size 1

        Sets up column header layout at the intersection of ``qtr_headrow``
        and 'quarters' in ``years_cols``.
        """
        # set width of the column holding quarterly numbers
        for qtr_colgroup in years_cols.find_all(
            None, 'quarters', None
        ):
            qtr_col = qtr_colgroup.get_group('quarter')
            address = qtr_headrow.get_corner_address(qtr_col)
            cell = sheet.cell(address)
            column = sheet.column_dimensions[cell.column]
            column.width = chef_settings.COLUMN_WIDTH
            column.outlineLevel = 1
            column.hidden = True

            # set quarter label at the start of quarterly block
            address = qtr_headrow.get_corner_address(qtr_colgroup)
            cell = sheet.cell(address)
            cell.value = qtr_colgroup.name
            CellStyles.format_subheader_label(cell, alignment='right')
            column = sheet.column_dimensions[cell.column]
            column.width = chef_settings.COLUMN_WIDTH

            # merge header cells
            if qtr_colgroup.size > 1:
                stretch = qtr_headrow.get_range_address(qtr_colgroup)
                sheet.merge_cells(stretch)

    def _month_headers(self, sheet, years_cols, mon_headrow):
        """


        ModelChef._month_headers() -> None

        --``years_cols`` main column header container
        --``mon_headrow`` header row, size 1

        Sets up column header layout at the intersection of ``mon_headrow``
        and month's location in ``years_cols``.
        """
        # iterate over all locators of the form:
        # 2017.quarters.1Q17.months.2017-02-28
        for mon_col in years_cols.find_all(
            None, 'quarters', None, 'months', None
        ):
            address = mon_headrow.get_corner_address(mon_col)
            cell = sheet.cell(address)
            cell.value = mon_col.name
            CellStyles.format_subheader_label(
                cell,
                alignment='right',
                color=LOWHEADER_COLOR
            )
            column = sheet.column_dimensions[cell.column]
            column.width = chef_settings.COLUMN_WIDTH
            column.outlineLevel = 1 + chef_settings.SUMMARY_INCLUDES_QUARTERS
            column.hidden = True

    def _annual_summary_headers(
        self, sheet, timeline, output_rows, output_cols
    ):
        """


        ModelChef._annual_summary_headers() -> None

        --``timeline`` quarterly summary timeline

        Create the layout for column headers on annual summary sheet.
        Fill in year, quarter and month labels in headers.
        """
        # column group for years
        years_cols = output_cols.get_group('years')
        # row for years headers
        year_headrow = output_rows.add_group('years', offset=1, size=1)
        if chef_settings.SUMMARY_INCLUDES_QUARTERS:
            # row for quarter headers, if needed
            qtr_headrow = output_rows.add_group('quarters', size=1)
        if chef_settings.SUMMARY_INCLUDES_MONTHS:
            # row for month headers, if needed
            mon_headrow = output_rows.add_group('months', size=1)

        # actual header labels, years and (possibly) quarters and months
        # nested in the form: years.2017.quarters.1Q17.months.2017-01-01
        company = timeline.current_period.content
        for date, period in sorted(timeline.items()):
            if date < company.period.end:
                continue
            if not period.content:
                continue
            # container for quarters (if requested) and year
            year_colgroup = years_cols.add_group(date.year)
            # label for quarter column
            qtr_name = self._quarter_name(date)
            # sub-container for quarters:
            # years.2017.quarters.1Q17
            qtr_colgroup = year_colgroup.add_group('quarters', qtr_name)
            if chef_settings.SUMMARY_INCLUDES_MONTHS:
                # terminal leaf for monthly values
                # years.2017.quarters.1Q17.months.2017-01-31
                qtr_colgroup.add_group('months', date, size=1)
            if chef_settings.SUMMARY_INCLUDES_QUARTERS:
                # terminal leaf for quarterly values, after months
                # years.2017.quarters.1Q17.quarter
                qtr_colgroup.add_group('quarter', size=1)
            # terminal leaf for year itself, after quarters
            # years.2017.year
            year_colgroup.add_group('year', size=1)
        # column layout is known at this point, calculate all col locations
        output_cols.calc_size()

        # fill out the headers, now that the column positions are known
        if chef_settings.SUMMARY_INCLUDES_MONTHS:
            self._month_headers(sheet, years_cols, mon_headrow)

        if chef_settings.SUMMARY_INCLUDES_QUARTERS:
            self._quarter_headers(sheet, years_cols, qtr_headrow)

        self._year_headers(sheet, years_cols, year_headrow)

    def _annual_summary_detail(
        self, sheet, timeline, output_rows, output_cols, col_selector
    ):
        """


        ModelChef._annual_summary_detail() -> None

        Fills in the periodic data on the annual summary sheet. Period is
        specified by ``col_selector``, which finds the matching output column
        in column headers, and needs to match the ``timeline``.
        """
        complete_label_rows = output_rows.get_group('complete_label')
        available_months_rows = output_rows.get_group('available_months')

        for period in timeline.iter_ordered():
            column = col_selector(period.end)
            if not column:
                continue

            unit = period.content
            if not (unit):
                continue

            unit.xl.set_sheet(sheet)
            sheet.bb.outline_level = 0

            # Complete T/F
            address = complete_label_rows.get_corner_address(column)
            cell = sheet.cell(address)
            cell.value = unit.complete
            cell.alignment = Alignment(horizontal='right', vertical='center')

            # Available months
            address = available_months_rows.get_corner_address(column)
            cell = sheet.cell(address)
            cell.value = unit.periods_used
            cell.alignment = Alignment(horizontal='right', vertical='center')

            # Statements
            statement_rowgroup = output_rows.add_group('statements', offset=1)
            for name, statement in unit.financials.chef_ordered():
                if statement is not None:
                    # to handle the hard link from starting to ending financials
                    if name == 'starting':
                        title = 'Starting Balance Sheet'
                    else:
                        title = statement.title
                    line_chef.chop_summary_statement(
                        sheet=sheet,
                        statement=statement,
                        row_container=statement_rowgroup,
                        col_container=column,
                        title=title,
                    )
