# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.delayed_cell
"""

Module defines an openpyxl cell whose location is yet to be determined.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
DelayedCell           defines an openpyxl cell whose location is not known yet
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from .cell_styles import CellStyles
from .data_types import TypeCodes




# Constants
# n/a

# Module Globals
get_column_letter = xlio.utils.get_column_letter


# Classes
class DelayedCell:
    """

    Class defines an openpyxl cell whose location is yet to be determined.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    from_cell()           wraps a real cell into a DelayedCell
    set_resolved()        replaces a DelayedCell with a real one
    resolve()             computes a real cell from a DelayedCell
    ====================  =====================================================
    """

    def __init__(self, line, **kargs):
        # must have
        self.line = line
        self.sheet = kargs['sheet']
        # unresolved locators
        self.row_container = kargs.get('row_container')
        self.col_container = kargs.get('col_container')
        # resolved cell and locators
        self.cell = kargs.get('cell')
        self.rownum = kargs.get('rownum')
        self.colnum = kargs.get('colnum')
        # cell content
        self.template = kargs.get('template')
        self.inputs = kargs.get('inputs', {})
        # a callable to do postprocessing on inputs dict after filling it
        self.inputs_mold = kargs.get('inputs_mold')
        self.value = kargs.get('value')
        self.cell_type = kargs.get('cell_type')
        self.update_cell = kargs.get('update_cell', True)
        self.formatter = kargs.get('formatter')
        # hook selves unto the line
        self.line.xl_data.cell_delayed = self
        # hook selves unto layout containers
        if self.row_container and self.col_container:
            self.row_container.cells[self.col_container.path] = self

    @classmethod
    def from_cell(cls, line, **kargs):
        """

        --``line`` LineItem with .xl_data property.

        Returns DelayedCell, two scenarios possible:
        1. ``line`` already has a DelayedCell attached
        2. wrap the actual cell attached to the line
        """
        if line.xl_data :
            cell = getattr(line.xl_data , 'cell_delayed', None)
            if cell:
                return cell
            cell = getattr(line.xl_data , 'cell', None)
            if cell:
                kargs['sheet'] = cell.parent
                kargs['cell'] = cell
                kargs['rownum'] = cell.row
                kargs['colnum'] = cell.col_idx
                return cls(line, **kargs)

    def set_resolved(self, cell):
        """

        --``cell`` openpyxl cell object, "hard" cell

        Hardens the cell object attached to line, completes cell resolution.
        """
        # we have been resolved
        self.cell = cell
        line = self.line
        # remove reference to self from target line
        line.xl_data.cell_delayed = None
        if self.update_cell:
            # attach new cell to target line, possibly replacing existing
            line.xl_data.cell = cell
            if self.cell_type:
                base = getattr(line.xl_data , self.cell_type)
                setattr(base, 'cell', cell)
        # apply formats
        CellStyles.format_line(self.line)
        if self.formatter:
            self.formatter(cell)

    def resolve(self):
        """

        resolve() -> None

        Must be run after AxisGroup.set_size(), i.e. when all locations are
        known. Determines our position on the sheet, computes our formula
        and create the actual Excel cell with contents.
        """
        self.rownum = self.row_container.number()
        self.colnum = self.col_container.number()
        target_cell = self.sheet.cell(
            row=self.rownum, column=self.colnum
        )

        if self.template is not None:
            # formula inputs: some_name = 'sheet1'!A1
            parsed = {}
            for name, source_cell in self.inputs.items():
                if not source_cell.cell:
                    source_cell.resolve()
                hard_cell = source_cell.cell
                sheet = hard_cell.parent
                row = hard_cell.row
                col = hard_cell.column
                if sheet.title != self.sheet.title:
                    sname = "'{}'!".format(sheet.title)
                else:
                    sname = ''
                parsed[name] = '{}{}{}'.format(sname, col, row)
            # postprocessing of inputs, e.g. to nest within another dict
            if self.inputs_mold:
                parsed = self.inputs_mold(parsed)
            formula = self.template.format(**parsed)
            target_cell.set_explicit_value(
                formula, data_type=TypeCodes.FORMULA
            )
        elif self.value is not None:
            target_cell.value = self.value

        self.set_resolved(target_cell)
