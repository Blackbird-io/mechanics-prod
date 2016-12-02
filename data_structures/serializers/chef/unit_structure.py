# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.unit_structure
"""

Module defines a class for displaying the company structure in Excel.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
StructureChef         class containing methods to display BusinessUnits
                      components on a sheet
====================  =========================================================
"""




# Imports
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_interval

import chef_settings

from .tab_names import TabNames
from .sheet_style import SheetStyle




# Constants
# n/a

# Module Globals
# n/a


# Classes
class StructureChef:
    """
    Creates a tab that lays out component structure.

    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    BOX_HEIGHT            box sizing
    BOX_ACROSS            box sizing
    SEP_ACROSS            width of columns between parent and children

    FUNCTIONS:
    chop()                creates a tab for unit structure
    chop_report()         creates spacer tab for reports including TOC
    unit_box()            creates a top-level box and recurs to components
    layout()              once all boxes are set, computes the layout
    ====================  =====================================================
    """
    BOX_HEIGHT = 2
    BOX_ACROSS = 2
    SEP_ACROSS = 4

    def __init__(self, model):
        self.model = model

    def chop(self, book):
        """


        StructureChef.chop() -> None

        --``book`` must be a Workbook

        Method recursively walks through top-level unit and components
        and displays a diagram of BU structure on an Excel sheet.
        """
        # sheet location
        ahead = book.get_sheet_by_name('Transcript')
        index = book.get_index(ahead) + 1

        # sheet layout
        sheet = book.create_sheet(TabNames.STRUCTURE, index)
        sheet.sheet_properties.tabColor = chef_settings.COVER_TAB_COLOR
        SheetStyle.style_sheet(sheet)

        header_rows = sheet.bb.row_axis.add_group('header', size=1)
        header_cols = sheet.bb.col_axis.add_group('header', size=1)
        body_rows = sheet.bb.row_axis.add_group('body')
        body_cols = sheet.bb.col_axis.add_group('body')

        unit = self.model.get_company()
        # start with top box
        self.unit_box(sheet, unit, body_rows, body_cols, level=0)
        body_rows.calc_size()
        body_cols.calc_size()
        self.layout(sheet, body_rows.groups)

    def chop_report(self, book, tab_color):
        """


        StructureChef.chop() -> None

        --``book`` must be a Workbook
        --``tab_color`` must be a Hex color

        Method adds a spacer tab before reports.  Method will eventually
        add a table of contents.
        """
        sheet = book.create_sheet(TabNames.REPORT_SPACER, 1)
        sheet.sheet_properties.tabColor = tab_color
        SheetStyle.style_sheet(sheet, label_areas=False)

    def unit_box(self, sheet, unit, row_container, col_container, level=0):
        """


        StructureChef.unit_box() -> None

        --``sheet`` must be a Worksheet
        --``unit`` must be an instance of BusinessUnit
        --``row_container`` AxisGroup on the rows
        --``col_container`` AxisGroup on the cols

        Method recursively walks through ``unit`` creating a nested column
        layout for unit boxes.
        """
        # col is determined by depth
        unit_cols = col_container.add_group(
            level,
            offset=2 if level else 0,
            size=self.BOX_ACROSS,
        )
        # row is determined by parent unit
        unit_rows = row_container.add_group(
            unit.title,
            offset=1 if row_container.groups else 0,
            label=unit.title,
            # intersection with the col axis
            across=unit_cols,
        )

        children = unit.components.get_ordered()
        if children:
            for child in children:
                self.unit_box(
                    sheet, child, unit_rows, col_container, level=level + 1
                )
        else:
            unit_rows.size = self.BOX_HEIGHT

    def layout(self, sheet, groups, level=0):
        """


        StructureChef.layout() -> None

        --``sheet`` must be a Worksheet
        --``groups`` list of AxisGroup on the row axis

        Method adds boxes representing company structure.
        """
        for group in groups:
            across = group.extra['across']
            if group.groups:
                self.layout(
                    sheet, group.groups, level=level + 1
                )
            self._make_cell(sheet, group, across)

    # *************************************************************************#
    #                          NON-PUBLIC METHODS                              #
    # *************************************************************************#

    def _make_cell(self, sheet, group, across):
        """

        StructureChef._make_cell() -> None

        --``sheet`` must be a Worksheet
        --``group`` AxisGroup on the row axis for this company layout

        Method draws a box representing a business unit.
        """
        # for an even-sized cell, first edge of "center" is the middle of cell
        center = int(group.size / 2) + group.number()
        # for 2-cell box, 1 before of center
        tip = center - int(self.BOX_HEIGHT / 2)
        col = across.number()
        label = group.extra.get('label')

        # all the cells that will be merged into a combined cell, row-col
        box = []
        side = Side(border_style='double')
        border = Border(left=side, right=side, top=side, bottom=side)
        for yspan in range(self.BOX_HEIGHT):
            box.append([])
            for xspan in range(self.BOX_ACROSS):
                c = sheet.cell(row=tip + yspan, column=col + xspan)
                c.border = border
                box[-1].append(c)
        sheet.merge_cells(
            start_row=tip, start_column=col,
            end_row=tip + 1, end_column=col + 1
        )

        # value and styling of the merged cell
        cell = box[0][0]
        cell.value = label
        cell.alignment = Alignment(
            wrapText=True, vertical='center', horizontal='center'
        )

        # hyperlink
        link = "#'{}'!B2".format(label)
        cell.hyperlink = link

        # draw connecting lines to children, if present
        group.extra['center'] = center
        if group.groups:
            self._child_line(sheet, group, across)

    def _child_line(self, sheet, group, across):
        """

        StructureChef._child_line() -> None

        Draws lines from parent to children.
        """
        # where the parent cell ends and the connection to children begins
        center = group.extra['center']
        tip = across.number() + self.BOX_ACROSS
        # stem from parent to children
        side = Side(border_style='thick')
        upper = sheet.cell(row=center, column=tip)
        upper.border = Border(top=side)
        under = sorted(g.extra['center'] for g in group.groups)

        # stems from children up to parent
        for row in under:
            c = sheet.cell(row=row, column=tip + 1)
            border = c.border.copy()
            border.top = side
            c.border = border

        # branch on which children hang
        if len(under) > 1:
            top = min(under)
            end = max(under)
            for row in range(top, end):
                c = sheet.cell(row=row, column=tip + 1)
                border = c.border.copy()
                border.left = side
                c.border = border

        # col sizing
        for c in get_column_interval(tip, tip + 1):
            sheet.column_dimensions[c].width = self.SEP_ACROSS
