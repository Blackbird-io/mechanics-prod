# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.model_chef
"""

Class for creating dynamic Excel representations of Blackbird Engine models.

Chef modules write formulas to cells explicitly, using the set_explicit_value()
method, to make sure Excel interprets the strings correctly.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
ModelChef             chop Blackbird Engine model into a dynamic Excel workbook
====================  =========================================================
"""




# Imports
import os
import openpyxl as xlio

import bb_settings
import chef_settings

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

from .bb_workbook import BB_Workbook as Workbook
from .cell_styles import CellStyles
from .garnish_chef import GarnishChef
from .line_chef import LineChef
from .report_chef import ReportChef
from .sheet_style import SheetStyle
from .summary_chef import SummaryChef
from .unit_chef import UnitChef
from .unit_structure import StructureChef




# Constants
IMAGE_PATH = os.path.join(
    os.path.dirname(__file__), 'static', 'blackbird_engine_2X_410x120.png'
)

# Module Globals
line_chef = LineChef()

get_column_letter = xlio.utils.get_column_letter
bounding_box = xlio.drawing.image.bounding_box

# Classes
class ModelChef:
    """

    Class packages an Engine model into an Excel workbook with dynamic links.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    chop_model()          returns BB_Workbook containing an Excel workbook with
                          dynamic links
    build_report()        returns BB_Workbook containing formatted report(s)
    ====================  =====================================================
    """

    def chop_model(self, model, base_file=None, include_ids=False):
        """


        ModelChef.chop_model -> BB_Workbook

        --``model`` is an instance of Blackbird Engine model

        Method delegates to UnitChef to chop BusinessUnits and returns an
        instance of BB_Workbook.  BB_Workbook contains an Excel workbook with
        dynamic links.
        """
        model.populate_xl_data()

        if base_file is not None:
            obook = xlio.load_workbook(base_file)
            preserved_styles = self.save_styles_from_file(obook)
            book = Workbook.convert(obook)
        else:
            book = Workbook()

        book = GarnishChef.add_garnishes(model, book=book)

        # add a tab with unit structure map
        structure_chef = StructureChef(model)
        structure_chef.chop(book)

        proj = model.get_timeline(resolution='monthly', name='default')
        unit_chef = UnitChef(model, timeline=proj, include_ids=include_ids)
        unit_chef.chop_multi(book)

        actl = model.get_timeline(resolution='monthly', name='actual')
        if actl:
            unit_chef = UnitChef(model, timeline=actl)
            unit_chef.chop_multi(book, tab_name='Actual')

        if bb_settings.MAKE_ANNUAL_SUMMARIES:
            summary_chef = SummaryChef(model)
            summary_chef.add_annual_summary(book)

        unit_chef.chop_multi_valuation(model, book, index=2, recur=False)

        CellStyles.format_line_borders(book)

        if base_file:
            self.apply_saved_styles_to_file(book, preserved_styles)

        return book

    def build_report(self, model, report='latest', specific_dates=None,
                     base_file=None):
        """


        ModelChef.build_report() -> BB_Workbook

        --``model`` is an instance of Blackbird Engine model
        --``report`` must be a string in {'all', 'latest', 'specific_dates'}
        --``specific_dates`` is a tuple of the date range for which to produce
                             reports if report == 'specific_dates'

        Method prepares and formats reports for the specified date range, or
        all available if no date range is provided.

        ``report`` keyword specifies which reports to produce.  The default is
        to produce only the latest report.  If report == ``specific_dates``,
        specific_dates keyword must be set, otherwise will produce latest
        report.
        """
        model.populate_xl_data()

        forecast_color = '4f6228'
        actual_color = '000000'
        spacer_color = '4f81bd'

        # Get timelines to report from
        proj = model.get_timeline(resolution='monthly', name='default')
        actl = model.get_timeline(resolution='monthly', name='actual')
        budg = model.get_timeline(resolution='monthly', name='budget')

        last_date = max(actl.keys())

        excel_data = model.portal_data.get('excel_data', dict())
        actl_pd = actl.find_period(last_date)
        co = model.get_company()
        actl_fins = co.get_financials(actl_pd)

        if base_file is not None:
            obook = xlio.load_workbook(base_file)
            preserved_styles = self.save_styles_from_file(obook)
            book = Workbook.convert(obook)
        else:
            book = Workbook()

        self.update_fins_refs(actl_fins, excel_data, book)

        # Make workbook and add Cover tab
        book = GarnishChef.add_garnishes(model, book=book, report=True,
                                         last_date=last_date)

        # Add "Forecast" tab filled with projections and "Actual" tab filled
        # with reported values.
        unit_chef = UnitChef(model, timeline=proj)
        unit_chef.chop_multi(book, values_only=True, tab_name='Forecast',
                             tab_color=forecast_color, index=0)

        if budg is not None:
            unit_chef = UnitChef(model, timeline=budg)
            unit_chef.chop_multi(book, values_only=True,
                                 tab_name='Budget Projections',
                                 tab_color=actual_color, index=0)

        unit_chef = UnitChef(model, timeline=actl)
        unit_chef.chop_multi(book, values_only=True, tab_name='Actual',
                             tab_color=actual_color, index=0)

        # Build reports
        report_chef = ReportChef(model, proj, actl, report,
                                 dates=specific_dates)
        report_chef.build_reports(book)

        if base_file:
            self.apply_saved_styles_to_file(book, preserved_styles)

        return book

    @staticmethod
    def update_fins_refs(financials, excel_data, wb):
        """


        update_fins_refs() -> None

        --``financials`` is the instance of financials on which to set refs
        --``excel_data`` is a dictionary of BBID: transcript cell address
        --``wb`` is the original workbook on which Report will be built

        Method works in-place to temporarily set direct references on
        last final period of actuals timeline based on transcript data from
        Portal.
        """

        if excel_data:
            for stmt in financials.full_ordered:
                if stmt:
                    for line in stmt.get_full_ordered():
                        bbid = line.id.bbid.hex
                        if bbid in excel_data:
                            source = excel_data[bbid]

                            sheet_name, cell_coord = source.split('!')
                            sheet_name = sheet_name.replace("'", "")

                            sheet = wb.get_sheet_by_name(sheet_name)
                            cell = sheet[cell_coord]

                            if cell.data_type == "f":
                                source = cell.value.replace("=", "")

                            line.xl_data.set_ref_direct_source(source,
                                                               update=False)

    @staticmethod
    def save_styles_from_file(wb):
        data = dict()
        for worksheet in wb.worksheets:
            sheet_dict = data.setdefault(worksheet.title, dict())
            for cell in worksheet.get_cell_collection():
                cell_dict = sheet_dict.setdefault(cell.coordinate, dict())

                cell_dict["number_format"] = cell.number_format

                #  FONT
                font_keys = ['family', 'vertAlign', 'strike', 'charset',
                             'color', 'b', 'scheme', 'condense', 'i',
                             'outline', 'u', 'shadow', 'extend', 'name', 'sz']
                font_dict = dict()
                for k in font_keys:
                    font_dict[k] = cell.font.__dict__[k]

                cell_dict["font"] = font_dict

                #  BORDER
                border_keys = ['diagonal_direction', 'vertical', 'horizontal',
                               'diagonalUp', 'diagonalDown', 'outline',
                               'start', 'end']
                border_sides = ['left', 'right', 'top', 'bottom', 'diagonal']
                side_keys = ['style', 'color']

                border_dict = dict()
                for k in border_keys:
                    border_dict[k] = cell.border.__dict__[k]
                cell_dict["border"] = border_dict

                side_dict = border_dict.setdefault('sides', dict())
                for s in border_sides:
                    side = cell.border.__dict__[s]
                    this_dict = side_dict.setdefault(s, dict())
                    for k in side_keys:
                        this_dict[k] = side.__dict__[k]

                #  FILL
                fill_keys = ['patternType', 'fgColor', 'bgColor']
                fill_dict = dict()
                for k in fill_keys:
                    fill_dict[k] = cell.fill.__dict__[k]

                cell_dict["fill"] = fill_dict

        return data

    @staticmethod
    def apply_saved_styles_to_file(wb, styles, skip_sheet=None):
        for worksheet in wb.worksheets:
            if worksheet.title == skip_sheet:
                continue

            if worksheet.title in styles:
                sheet_styles = styles[worksheet.title]
                for cell in worksheet.get_cell_collection():
                    cell_dict = sheet_styles.get(cell.coordinate, None)
                    if cell_dict:
                        cell.number_format = cell_dict["number_format"]

                        cell.font = Font(**cell_dict["font"])

                        border_dict = cell_dict["border"]
                        sides_dict = border_dict.pop('sides')
                        for side, info in sides_dict.items():
                            side_obj = Side(**info)
                            border_dict[side] = side_obj
                        cell.border = Border(**border_dict)

                        cell.fill = PatternFill(**cell_dict["fill"])