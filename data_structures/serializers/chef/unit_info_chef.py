# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.unit_info_chef
"""

Module defines a class that represents arbitrarily rich BusinessUnit instances
as a collection of linked Excel worksheets.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
UnitInfoChef          class containing methods to ad BusinessUnit attributes
                      (context info, e.g. life and parameters) to sheet
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from ._chef_tools import add_scenario_selector, group_lines
from .cell_styles import CellStyles
from .data_types import TypeCodes
from .field_names import FieldNames
from .formulas import FormulaTemplates
from .sheet_style import SheetStyle
from .tab_names import TabNames

from chef_settings import SCENARIO_SELECTORS, HIDE_LIFE_EVENTS
from data_structures.modelling import common_events




# Constants
# n/a

# Module Globals
_INVALID_CHARS = r"\/*[]:?"
# Excel forbids the use of these in sheet names.
REPLACEMENT_CHAR = None
# When the replacement character is None, UnitChef will remove all bad chars
# from sheet titles.
bad_char_table = {ord(c): REPLACEMENT_CHAR for c in _INVALID_CHARS}
get_column_letter = xlio.utils.get_column_letter

# Classes
class UnitInfoChef:
    """

    One tab per unit
    Children first
    Arbitrarily recursive (though should max out at sheet limit; alternatively,
    book should prohibit new sheets after that? or can have 2 limits: a soft
    limit where ModelChopper shifts into different representation mode, and a
    hard limit, where you just cant create any more sheets.)

    Most non-public methods force keyword-based arg entry to avoid potentially
    confusing erros (switching rows for columns, etc.)

    Methods generally leave current row pointing to their last completed
    (filled) row.

    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    MAX_CONSOLIDATION_ROWS = 15
    MAX_LINKS_PER_CELL = 1
    MAX_TITLE_CHARACTERS = 30

    FUNCTIONS:
    add_items_to_area()   adds dictionary items to specified area
    add_unit_life()       adds unit life to sheet for a particular period
    add_scenario_selector_logic() adds logic to use scenario selector to sheet
    add_unit_size()       adds unit's size to sheet
    create_unit_sheet()   creates tab for unit's information
    unit_life()           adds unit life information over time
    ====================  =====================================================
    """
    MAX_CONSOLIDATION_ROWS = 15
    MAX_LINKS_PER_CELL = 1

    MAX_TITLE_CHARACTERS = 30

    LABEL_COLUMN = 2
    MASTER_COLUMN = 4
    VALUE_COLUMN = 6

    TITLE_ROW = 3
    VALUES_START_ROW = 6

    SCENARIO_ROW = 2

    def __init__(self, model, timeline):
        self.model = model
        self.timeline = timeline

    def add_items_to_area(self, *pargs, sheet, area, items, active_column,
                          set_labels=True, format_func=None,
                          preference_order=[], group=True):
        """


        UnitChef.add_items_to_area() -> Worksheet

        --``sheet`` must be an instance of Worksheet
        --``area`` must be an instance of Area
        --``items`` must be a dictionary of items to add
        --``active_column`` must be current column index
        --``set_labels`` must be a boolean; whether or not to label row
        --``format_func`` is a function to use on cells for formatting
        --``preference_order`` is optionally a specific ordering
        --``group`` bool; whether to group the rows these items are added to

        Method adds names to Area in sorted order.  Method starts work after
        the current row and expects sheet to come with parameters unless you
        specify the label and master column.
        """

        parameters = getattr(sheet.bb, FieldNames.PARAMETERS)
        label_column = parameters.columns.get_position(FieldNames.LABELS)
        master_column = parameters.columns.get_position(FieldNames.MASTER)

        SheetStyle.set_column_width(sheet, master_column)

        starting = sheet.bb.current_row or 0
        new_row = starting + 1

        for name in self._sort_bypreference(items, preference_order or []):
            value = items[name]

            # Register the new row
            area.rows.by_name[name] = new_row

            # Add the label
            if set_labels:
                label_cell = sheet.cell(column=label_column, row=new_row)
                label_cell.value = name
                # Or run through labels routine

            # Add the master value (from this period)
            master_cell = sheet.cell(column=master_column, row=new_row)
            master_cell.value = value
            CellStyles.format_parameter(master_cell)
            CellStyles.format_hardcoded(master_cell)

            # Link the period to the master
            current_cell = sheet.cell(column=active_column, row=new_row)
            link = FormulaTemplates.ADD_COORDINATES
            link = link.format(coordinates=master_cell.coordinate)
            current_cell.set_explicit_value(link, data_type=TypeCodes.FORMULA)
            CellStyles.format_parameter(current_cell)

            if format_func:
                format_func(master_cell)
                format_func(current_cell)

            sheet.bb.current_row = new_row

            if group:
                group_lines(sheet)

            new_row += 1

        return sheet

    def add_scenario_selector_logic(self, book, sheet):
        """

        add_scenario_selector_logic() -> None

        --``book`` must be a Workbook
        --``sheet`` must be an instance of openpyxl Worksheet

        Method adds logic to use scenario selector from unit tab.  Scenario
        selector changes parameter values.
        """
        scen_tab = book.get_sheet_by_name(TabNames.SCENARIOS)
        src_sheet = scen_tab.title
        src_row = scen_tab.bb.general.rows.by_name[FieldNames.SELECTOR]

        scen_area = getattr(scen_tab.bb, FieldNames.PARAMETERS)
        num_col = scen_area.columns.by_name[FieldNames.CUSTOM_CASE]
        src_col = get_column_letter(num_col)

        info = dict(sheet=src_sheet, alpha_column=src_col, row=src_row)

        active_label_cell = sheet.cell(column=self.LABEL_COLUMN,
                                       row=self.SCENARIO_ROW)
        active_label_cell.value = FieldNames.IN_EFFECT

        template = FormulaTemplates.LINK_TO_CELL_ON_SHEET
        link = template.format(**info)
        scen_cell = sheet.cell(column=self.MASTER_COLUMN,
                               row=self.SCENARIO_ROW)
        scen_cell.value = link

        if SCENARIO_SELECTORS:
            label_column = self.LABEL_COLUMN
            add_scenario_selector(sheet, label_column, self.SCENARIO_ROW,
                                  book.scenario_names)
        else:
            CellStyles.format_scenario_selector_cells(sheet,
                                                      self.LABEL_COLUMN,
                                                      self.MASTER_COLUMN,
                                                      self.SCENARIO_ROW,
                                                      active=False)

    def add_unit_size(self, sheet, unit):
        """


        UnitChef._add_unit_size() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``unit`` must be an instance of BusinessUnit
        --``column`` must be a column number reference

        Method adds "size" Area to unit sheet and populates it with values.
        Will start writing on current row.
        """
        body_rows = sheet.bb.row_axis.get_group('body')
        size_group = body_rows.add_group(
            'size', offset=1
        )
        size_title = size_group.add_group(
            'title', size=1, label='Size', rank=1
        )
        size_lines = size_group.add_group('lines')
        body_rows.calc_size()

        timeline_range = getattr(sheet.bb, FieldNames.TIMELINE)
        time_line = self.model.get_timeline()
        now = time_line.current_period
        template = FormulaTemplates.ADD_COORDINATES
        master_column = self.MASTER_COLUMN

        size = sheet.bb.add_area(FieldNames.SIZE)
        size.rows.by_name[FieldNames.SIZE_LABEL] = size_lines.number()

        for period in time_line.iter_ordered(open=now.end):
            period_column = timeline_range.columns.get_position(period.end)
            rowbox = size_lines.add_group(
                FieldNames.SIZE_LABEL,
                size=1,
                label=FieldNames.SIZE_LABEL,
                outline=1
            )
            master_cell = sheet.cell(
                column=master_column, row=rowbox.number()
            )
            active_cell = sheet.cell(
                column=period_column, row=rowbox.number()
            )
            # TODO: unit.get_size(period)
            value = unit.size
            if value and not master_cell.value:
                master_cell.value = value
                CellStyles.format_parameter(master_cell)
                CellStyles.format_hardcoded(master_cell)
                CellStyles.format_integer(master_cell)
            if value == master_cell.value:
                link = template.format(coordinates=master_cell.coordinate)
                active_cell.set_explicit_value(
                    link, data_type=TypeCodes.FORMULA
                )
            else:
                active_cell.value = value
                CellStyles.format_hardcoded(master_cell)
            CellStyles.format_integer(active_cell)

        return sheet

    def create_unit_sheet(
        self, book, unit, index, name=None, current_only=False,
        values_only=False, tab_color='',
    ):
        """


        UnitChef._create_unit_sheet() -> Worksheet

        --``book`` must be a Workbook
        --``unit`` must be an instance of BusinessUnit
        --``index`` is optionally the index at which to insert the tab
        --``name`` is the name to give the tab
        --``current_only`` bool; whether to add all periods or only current
        --``values_only`` bool; whether to write flowing Excel formulas and
                          include background info, like life and parameters, on
                          the unit sheet
        --``tab_color`` str; Hex color definition for tab color

        Returns sheet with current row pointing to last parameter row.
        ``values_only`` = True will print only values to Excel, no driver
        calculations, no consolidation, so Life, Events, and Drivers sections
        will not be written to the sheet.
        """
        if not name:
            name = unit.tags.title

        if name in book:
            rev_name = name + " ..." + str(unit.id.bbid)[-8:]
            name = rev_name

            if name in book:
                name = str(unit.id.bbid)

        name = name.translate(bad_char_table)
        name = name[:self.MAX_TITLE_CHARACTERS]
        # Replace forbidden characters, make sure name is within length limits

        sheet = book.create_sheet(name, index)

        req_rows = len(unit.components.by_name) // self.MAX_LINKS_PER_CELL
        req_rows = min(req_rows, self.MAX_CONSOLIDATION_ROWS)
        req_rows = max(1, req_rows)

        sheet.bb.consolidation_size = req_rows
        # Compute the amount of rows we will use for consolidation on this
        # sheet. In future periods, the number of kids may grow or shrink.
        # LineChef will modify the number of links in each cell to make sure
        # the full consolidation work always takes up the same amount of row
        # space, to make sure our spreadsheet aligns.

        unit.xl.set_sheet(sheet)

        # Hide sheets for units below a certain depth. The depth should be a
        # Chef-level constant. Use ``sheet_state := "hidden"`` to implement.

        head_rows = sheet.bb.row_axis.add_group('head', size=self.TITLE_ROW)
        head_cols = sheet.bb.col_axis.add_group('head', size=self.MASTER_COLUMN)
        sheet.bb.row_axis.add_group('top_spacer', size=1)
        sheet.bb.col_axis.add_group('top_spacer', size=1)
        body_rows = sheet.bb.row_axis.add_group('body')
        body_cols = sheet.bb.col_axis.add_group('body')

        if not values_only:
            # link to TimeLine parameters
            self._link_to_time_line(book=book, sheet=sheet, unit=unit,
                                    current_only=current_only)

            # unit-specific parameters
            # Add unit parameters and update TimeLine/Period params as necessary
            self._add_unit_params(sheet, unit, current_only=current_only)

            val_col = sheet.bb.time_line.columns.get_position(unit.period.end)
            param_area = getattr(sheet.bb, FieldNames.PARAMETERS)
            param_area.columns.by_name[FieldNames.VALUES] = val_col
            # At this point, sheet.bb.current_row will point to the last parameter.

            corner_col = param_area.columns.get_position(FieldNames.MASTER) + 1
        else:
            #  Will need to add timeline headers manually
            self._add_independent_timeline(sheet=sheet)

            corner_col = self.VALUE_COLUMN - 1

        # Freeze panes:
        corner_row = sheet.bb.time_line.rows.ending
        corner_row += 1

        corner_cell = sheet.cell(column=corner_col, row=corner_row)
        sheet.freeze_panes = corner_cell

        if tab_color:
            sheet.sheet_properties.tabColor = tab_color

        # Return sheet
        return sheet

    def unit_life(self, sheet, unit, current_only=False):
        """


        UnitChef.unit_life() -> None

        Method adds Life and Events sections to unit sheet and delegates to
        _add_unit_life to write each time period.
        """
        body_rows = sheet.bb.row_axis.get_group('body')

        # layout for Life
        life_group = body_rows.add_group(
            'life',
            offset=1 if body_rows.groups else 0,
            hidden=HIDE_LIFE_EVENTS,
        )
        life_title = life_group.add_group(
            'title', size=1, label='Life', rank=1,
        )
        life_lines = life_group.add_group('lines')
        for label in (
            FieldNames.REF_DATE,
            FieldNames.START_DATE,
            # blank line before age
            None,
            FieldNames.AGE,
            FieldNames.ALIVE,
            FieldNames.SPAN,
            FieldNames.PERCENT,
        ):
            life_lines.add_group(
                label or '_spacer', size=1, label=label,
                outline=int(not HIDE_LIFE_EVENTS),
            )

        # layout for Events
        event_group = body_rows.add_group(
            'events', offset=1,
            hidden=HIDE_LIFE_EVENTS,
        )
        event_title = event_group.add_group(
            'title', size=1, label='Events', rank=1,
        )
        event_lines = event_group.add_group('lines')
        for label in unit.life.ORDER:
            box = event_lines.add_group(
                label, size=1, label=label,
                outline=int(not HIDE_LIFE_EVENTS),
            )
        body_rows.calc_size()

        if not getattr(sheet.bb, "life", None):
            sheet.bb.add_area("life")
            for box in life_lines.groups:
                if box.extra.get('label'):
                    sheet.bb.life.rows.by_name[box.name] = box.number()

        if not getattr(sheet.bb, "events", None):
            sheet.bb.add_area("events")
            for box in event_lines.groups:
                if box.extra.get('label'):
                    sheet.bb.events.rows.by_name[box.name] = box.number()

        # fill in cell by period
        timeline_range = getattr(sheet.bb, FieldNames.TIMELINE)
        time_line = self.model.get_timeline()
        now = time_line.current_period
        for period in time_line.iter_ordered(open=now.end):
            period_column = timeline_range.columns.get_position(period.end)
            self._add_events(sheet, unit, period, period_column)
            self._add_life(sheet, unit, period, period_column)
            if current_only:
                break

    # *************************************************************************#
    #                          NON-PUBLIC METHODS                              #
    # *************************************************************************#
    def _add_independent_timeline(self, sheet):
        """
        Method adds stand-alone timeline to unit sheet.  Method writes values
        to header row (does not link to another tab).  Method will not write
        periods preceding the current_period, if defined.
        """
        timeline_area = sheet.bb.add_area(FieldNames.TIMELINE)
        timeline_area.rows.by_name[FieldNames.TITLE] = self.TITLE_ROW
        active_column = self.VALUE_COLUMN
        active_row = self.TITLE_ROW

        now = getattr(self.timeline, 'current_period', None)
        if now is None:
            now = self.timeline[min(self.timeline.keys())]

        for date in sorted(self.timeline.keys()):
            if date < now.end:
                continue

            timeline_area.columns.by_name[date] = active_column
            cell = sheet.cell(column=active_column, row=active_row)
            cell.value = date
            CellStyles.format_date(cell)
            active_column += 1

    def _add_labels(self, sheet, groups, label_col, level=0):
        """


        UnitChef._add_labels() -> None

        Writes row labels on sheet. To show up on the axis, a group
        1. should have no subgroups
        2. should have a label
        To add a title row for a group with subgroups, create a one-row
        'title' subgroup.
        """
        for group in groups:
            if group.groups:
                self._add_labels(
                    sheet, group.groups, label_col, level=level + 1
                )
            elif group.size:
                label = group.extra.get('label')
                if label:
                    row = group.number()
                    col = label_col.number()
                    rank = group.extra.get('rank')
                    if group.name == 'title' and rank == 1:
                        formatter = CellStyles.format_area_label
                        formatter(sheet, label, row, col_num=col)
                    else:
                        label_cell = sheet.cell(row=row, column=col + 1)
                        label_cell.value = label
                        formatter = group.extra.get('formatter')
                        if formatter:
                            formatter(label_cell)
                        if group.outline:
                            r = sheet.row_dimensions[row]
                            r.outline_level = group.outline

    def _add_life(self, sheet, unit, period, active_column):
        """


        UnitChef._add_life_analysis() -> Worksheet

        Method adds unit life for a single period to show unit state
        (alive/dead/etc.), age, etc.  Method assumes events are filled out.
        """
        event_lines = sheet.bb.row_axis.get_group('body', 'events', 'lines')
        life_lines = sheet.bb.row_axis.get_group('body', 'life', 'lines')

        timeline_range = getattr(sheet.bb, FieldNames.TIMELINE)
        timeline_row = timeline_range.rows.get_position(FieldNames.TITLE)

        birth = sheet.cell(
            column=active_column,
            row=event_lines.get_group(common_events.KEY_BIRTH).number()
        )
        death = sheet.cell(
            column=active_column,
            row=event_lines.get_group(common_events.KEY_DEATH).number()
        )
        conception = sheet.cell(
            column=active_column,
            row=event_lines.get_group(common_events.KEY_CONCEPTION).number()
        )

        cells = dict()
        cells["birth"] = birth
        cells["death"] = death
        cells["conception"] = conception

        # 1. Add ref_date
        box = life_lines.get_group(FieldNames.REF_DATE)
        ref_date = sheet.cell(column=active_column, row=box.number())
        time_line = sheet.cell(column=active_column, row=timeline_row)
        formula = FormulaTemplates.LINK_TO_COORDINATES.format(
            coordinates=time_line.coordinate
        )
        ref_date.value = formula
        CellStyles.format_date(ref_date)
        cells["ref_date"] = ref_date

        # 2. Add period start date
        box = life_lines.get_group(FieldNames.START_DATE)
        start_date = sheet.cell(column=active_column, row=box.number())
        start_date.value = period.start
        CellStyles.format_date(start_date)

        # 3. Add age
        box = life_lines.get_group(FieldNames.AGE)
        age = sheet.cell(column=active_column, row=box.number())
        cos = {k: v.coordinate for k, v in cells.items()}
        formula = FormulaTemplates.COMPUTE_AGE_IN_DAYS.format(**cos)
        age.set_explicit_value(formula, data_type=TypeCodes.FORMULA)
        CellStyles.format_parameter(age)
        cells["age"] = age
        cos[FieldNames.AGE] = age.coordinate

        # 4. Add alive
        box = life_lines.get_group(FieldNames.ALIVE)
        alive = sheet.cell(column=active_column, row=box.number())
        formula = FormulaTemplates.IS_ALIVE.format(**cos)
        alive.set_explicit_value(formula, data_type=TypeCodes.FORMULA)
        CellStyles.format_parameter(alive)
        cells["alive"] = alive
        cos["alive"] = alive.coordinate

        # 5. Add span (so we can use it as the denominator in our percent
        # computation below).
        box = life_lines.get_group(FieldNames.SPAN)
        span = sheet.cell(column=active_column, row=box.number())
        formula = FormulaTemplates.COMPUTE_SPAN_IN_DAYS.format(**cos)
        span.set_explicit_value(formula, data_type=TypeCodes.FORMULA)
        CellStyles.format_parameter(span)
        cells[FieldNames.SPAN] = span
        cos[FieldNames.SPAN] = span.coordinate

        # 6. Add percent
        box = life_lines.get_group(FieldNames.PERCENT)
        percent = sheet.cell(column=active_column, row=box.number())
        formula = FormulaTemplates.COMPUTE_AGE_IN_PERCENT.format(**cos)
        percent.set_explicit_value(formula, data_type=TypeCodes.FORMULA)

        # Return sheet
        return sheet

    def _add_events(self, sheet, unit, period, active_column):
        """


        UnitChef._add_life_events() -> Worksheet


        Method adds life events to unit Worksheet and returns updated sheet.
        Expects sheet to include areas for events and parameters.
        Runs through add_items() [which is why we get name-based sorting]
        """
        event_lines = sheet.bb.row_axis.get_group('body', 'events', 'lines')
        master_column = self.MASTER_COLUMN
        template = FormulaTemplates.ADD_COORDINATES

        for name, event_date in unit.life.events.items():
            row = event_lines.get_group(name).number()
            master_cell = sheet.cell(column=master_column, row=row)
            active_cell = sheet.cell(column=active_column, row=row)
            if not master_cell.value:
                master_cell.value = event_date
                CellStyles.format_date(master_cell)
                CellStyles.format_hardcoded(master_cell)
            if event_date == master_cell.value.date():
                link = template.format(coordinates=master_cell.coordinate)
                active_cell.set_explicit_value(
                    link, data_type=TypeCodes.FORMULA
                )
                CellStyles.format_date(active_cell)
            else:
                print(master_cell.value, event_date)
                active_cell.value = event_date
                CellStyles.format_date(active_cell)
                CellStyles.format_hardcoded(active_cell)

        return sheet

    def _add_unit_params(
        self, sheet, unit, current_only=False, set_labels=True
    ):
        """


        UnitChef._add_unit_params() -> Worksheet

        - Add any new unit parameters, place master value in MASTER column
        - Check on timeline_params and update with hardcoded value as applicable
        """
        parameters = getattr(sheet.bb, FieldNames.PARAMETERS)
        timeline_range = getattr(sheet.bb, FieldNames.TIMELINE)
        timeline_params = timeline_range.rows.by_name.keys()

        param_lines = sheet.bb.row_axis.get_group('body', 'drivers', 'lines')
        time_line = self.model.get_timeline()
        now = time_line.current_period
        template = FormulaTemplates.LINK_TO_COORDINATES

        for period in time_line.iter_ordered(open=now.end):
            period_column = timeline_range.columns.get_position(period.end)

            # period, unit, and period-unit parameters
            # combined in order of precedence
            allpar = unit.get_parameters(period)

            for param, value in sorted(allpar.items()):
                rowbox = param_lines.add_group(
                    param, size=1, label=param, outline=1
                )
                this_row = rowbox.number()
                parameters.rows.by_name[param] = this_row
                cell = sheet.cell(
                    row=this_row, column=period_column
                )
                if param in timeline_params:
                    if value != cell.value:
                        cell.value = value
                        CellStyles.format_hardcoded(cell)
                else:
                    master_cell = sheet.cell(
                        row=this_row, column=self.MASTER_COLUMN
                    )
                    if value and not master_cell.value:
                        master_cell.value = value
                        CellStyles.format_parameter(master_cell)
                        CellStyles.format_hardcoded(master_cell)
                    if value == master_cell.value:
                        info = dict(coordinates=master_cell.coordinate)
                        link = template.format(**info)
                        cell.set_explicit_value(
                            link, data_type=TypeCodes.FORMULA
                        )
                    else:
                        cell.value = value
                        CellStyles.format_hardcoded(cell)
                    CellStyles.format_parameter(cell)

            if current_only:
                break

        sheet.bb.current_row = parameters.rows.ending or self.VALUES_START_ROW

        return sheet

    def _link_to_time_line(self, *pargs, book, sheet, unit,
                           current_only=False):
        """


        UnitChef._link_to_time_line() -> Worksheet


        Link sheet to book's time_line.
        Force keyword-entry for book and sheet to make sure we feed in the
        right arguments.
        """
        source = book.get_sheet_by_name(TabNames.SCENARIOS)
        source_area = getattr(source.bb, FieldNames.TIMELINE)

        param_area = sheet.bb.add_area(FieldNames.PARAMETERS)
        timeline_area = sheet.bb.add_area(FieldNames.TIMELINE)

        # First add labels for parameters
        active_row = self.VALUES_START_ROW
        active_column = self.LABEL_COLUMN
        param_area.columns.by_name[FieldNames.LABELS] = self.LABEL_COLUMN
        param_area.columns.by_name[FieldNames.MASTER] = self.MASTER_COLUMN

        template = FormulaTemplates.ADD_CELL_FROM_SHEET
        source_label_column = source_area.columns.by_name[FieldNames.LABELS]
        src_col = get_column_letter(source_label_column)

        # link to parameters from Drivers tab
        body_rows = sheet.bb.row_axis.get_group('body')
        param_group = body_rows.add_group('drivers')
        param_title = param_group.add_group(
            'title', size=1, label='Drivers', rank=1
        )
        param_lines = param_group.add_group('lines', outline=1)

        src_params = set(source_area.rows.by_name.keys()) - {FieldNames.TITLE}
        for param in sorted(src_params):
            rowbox = param_lines.add_group(
                param, size=1, label=param, outline=1
            )
            param_area.rows.by_name[param] = rowbox.number()
            src_row = source_area.rows.by_name[param]
            cell = sheet.cell(column=active_column, row=rowbox.number())

            info = dict(sheet=source.title, alpha_column=src_col, row=src_row)
            link = template.format(**info)
            cell.set_explicit_value(link, data_type=TypeCodes.FORMULA)

        # Next add timeline header row and parameters from Scenarios tab
        timeline_area.rows.by_name[FieldNames.TITLE] = self.TITLE_ROW
        template = FormulaTemplates.ADD_CELL_FROM_SHEET
        src_vals = \
            set(source_area.columns.by_name.keys()) - {FieldNames.LABELS}

        if current_only:
            src_vals = [unit.period.end]

        active_column = self.VALUE_COLUMN
        for date in sorted(src_vals):
            active_row = self.TITLE_ROW
            src_row = source_area.rows.by_name[FieldNames.TITLE]

            # make header cell
            col_num = source_area.columns.by_name[date]
            src_col = get_column_letter(col_num)

            timeline_area.columns.by_name[date] = active_column
            cell = sheet.cell(column=active_column, row=active_row)

            info = dict(sheet=source.title, alpha_column=src_col, row=src_row)
            link = template.format(**info)
            cell.set_explicit_value(link, data_type=TypeCodes.FORMULA)
            CellStyles.format_date(cell)

            # now link parameters to value cells
            for param in src_params:
                src_row = source_area.rows.by_name[param]
                active_row = param_area.rows.by_name[param]

                cell = sheet.cell(column=active_column, row=active_row)

                info = dict(sheet=source.title, alpha_column=src_col,
                            row=src_row)
                link = template.format(**info)
                cell.set_explicit_value(link, data_type=TypeCodes.FORMULA)
                CellStyles.format_parameter(cell)

            SheetStyle.set_column_width(sheet, active_column)

            active_column += 1

        body_rows.calc_size()

        return sheet

    def _sort_bypreference(self, items, preference_order=[]):
        """


        UnitChef._sort_bypreference() -> list (of items' keys)

        --``items`` is any dict
        --``preference_order`` any iterable giving the sorted order of items;
            items' keys which are not in preference_order will be tacked on at
            the end in sorted order
        """
        result = []

        for k in preference_order:
            if k in items:
                result.append(k)

        leftover = set(items.keys()) - set(result)
        result.extend(sorted(leftover))

        return result
