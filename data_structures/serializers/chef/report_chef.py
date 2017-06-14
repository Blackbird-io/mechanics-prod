# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.report_chef
"""

Class for creating dynamic Excel reports showing forecast and actual values.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
ReportChef            class to generate reports from forecast and actual values
====================  =========================================================
"""




# Imports
from data_structures.modelling.line_item import LineItem
from openpyxl.styles.colors import WHITE, BLACK
from openpyxl.utils import get_column_letter

from .cell_styles import CellStyles
from .data_types import TypeCodes, NumberFormats
from .formulas import FormulaTemplates
from .sheet_style import SheetStyle




# Constants
# N/A

# Module Globals
# N/A

# Classes
class ReportChef:
    """

    Class packages an Engine model containing forecast and actual values into
    and Excel workbook containing reports for the specified dates (or all) 
    with dynamic links and (someday) graphics and analysis.
    ==============  ===========================================================
    Attribute             Description
    ==============  ===========================================================

    DATA:
    actual          obj; instance of Timeline containing actual financials
    end_date        datetime.date; last date for which to make reports
    forecast        obj; instance of Timeline containing forecast financials
    model           obj; instance of blackbird model
    start_date      datetime.date; first date for which to make reports

    FUNCTIONS:
    add_report()     method adds a specific report
    build_reports()  method builds reports within date range
    ===============  ==========================================================
    """

    def __init__(self, model, proj, actl, report_type, dates=None):
        self.model = model
        self.forecast = proj
        self.actual = actl

        # Validate arguments
        options = ['all', 'latest', 'specific_dates']

        if report_type not in options:
            report_type = 'latest'

        if report_type == 'specific_dates' and dates is None:
            report_type = 'latest'

        if report_type == 'all':
            self.start_date = min(self.actual.keys())
            self.end_date = max(self.actual.keys())
        elif report_type == 'latest':
            last_period = self.actual[max(self.actual.keys())]
            self.start_date = last_period.start
            self.end_date = last_period.end
        elif report_type == 'specific_dates':
            self.start_date = min(dates)
            self.end_date = max(dates)
        else:
            c = 'Report type %s is unknown, select from {all, latest,' \
                ' specific_dates}' % report_type
            raise KeyError(c)

        self._placeholder = '--'  # value to use in reports for Divide by Zero
                                  # error or unavailable data

        # FORMATTING
        self._report_tab_color = WHITE
        self._banner_color = '002060'  # navy blue
        self._banner_font_color = WHITE
        self._tbl_hdr_color = '00FFFFFF'  # transparent
        self._tbl_hdr_font_color = BLACK

    def add_report(self, wb, act_period, for_period):
        """


        ReportChef.add_report() -> None

        --``wb`` must be a BB_Workbook
        --``act_period`` must be a TimePeriod object containing actual reported
                         financials
        --``for_period`` must be a TimePeriod object containing forecast
                         financials corresponding to those in ``act_period``

        Method creates a new report tab and fills in the forecast and actual
        data.  Method adds columns for delta and percent difference.
        """

        # Do preliminary formatting
        use_date = act_period.end
        tab_name = use_date.strftime('%Y-%m-%d')
        title = '%s: %s Performance' % (self.model.title, use_date.strftime('%m-%Y'))
        sheet = wb.create_sheet(tab_name, index=0)
        self._make_report_header(sheet, title, use_date)

        co = self.model.get_company()

        act_fins = self.model.get_financials(co.id.bbid, act_period)
        for_fins = self.model.get_financials(co.id.bbid, for_period)

        for name, act_statement in act_fins.chef_ordered():
            if act_statement is not None:
                if name == 'starting':
                    label = 'Starting Balance Sheet'
                else:
                    label = None

                for_statement = for_fins.get_statement(name)
                self._report_statement(sheet, act_statement, for_statement, name=label)

        # Add Finishing Touches
        sheet.bb.calc_sizes()

        all_col = sheet.bb.col_axis.get_group('all')
        st_col = all_col.get_group('space_left').number()
        ed_col = all_col.get_group('space_right').number()

        all_row = sheet.bb.row_axis.get_group('all')
        st_row = all_row.number()
        ed_row = st_row + all_row.size

        freeze_cell = 'D7'
        corner_cell = sheet[freeze_cell]
        sheet.freeze_panes = corner_cell

        CellStyles.format_border_group(sheet, st_col, ed_col, st_row, ed_row)

    def build_reports(self, wb):
        """


        ReportChef.build_reports() -> None

        --``wb`` must be a BB_Workbook

        Method delegates to add_report() to build reports within the time
        range specified on instantiation (between self.start_date and
        self.end_date).
        """
        actual_periods = sorted(self.actual.values(), key=lambda x: x.start)

        for act_period in actual_periods:
            if act_period.start >= self.start_date and act_period.end <= self.end_date:
                for_period = self.forecast.find_period(act_period.end)

                self.add_report(wb, act_period, for_period)

            elif act_period.end > self.end_date:
                # since we sorted keys first, we know we won't want
                # any future periods
                break

    # ************************************************************************#
    #                          NON-PUBLIC METHODS                             #
    # ************************************************************************#

    def _add_details(self, sheet, act_line, for_line, row_container, indent=0):
        """
        Method adds details within report format.  Method formulates a
        summation equation which will be set as the value for the parent
        line to show this simple relationship.
        """
        all_cols = sheet.bb.col_axis.get_group('all')
        forecast_col = all_cols.get_group('Forecast')
        actual_col = all_cols.get_group('Actual')

        actual_details = act_line.get_ordered()

        if actual_details:
            sub_indent = indent + LineItem.TAB_WIDTH
            act_detail_summation = ""
            for_detail_summation = ""

            sheet.bb.outline_level += 1
            for act_det in actual_details:
                for_det = for_line.find_first(act_det.name)

                if for_det is not None:
                    self._report_line(
                        sheet=sheet,
                        act_line=act_det,
                        for_line=for_det,
                        row_container=row_container,
                        indent=sub_indent,
                    )

                    link_template = FormulaTemplates.ADD_COORDINATES

                    cos = act_det.xl_data.get_coordinates(include_sheet=False)
                    link = link_template.format(coordinates=cos)
                    act_detail_summation += link

                    cos = for_det.xl_data.get_coordinates(include_sheet=False)
                    link = link_template.format(coordinates=cos)
                    for_detail_summation += link

            sheet.bb.outline_level -= 1
            row_container.calc_size()

            if act_line.sum_details:
                if act_line.xl_format.blank_row_before:
                    row_container.add_group('spacer_details', size=1)

                # subtotal row for details
                line_label = indent * " " + act_line.title
                finish = row_container.add_group(act_line.title, size=1) # , label=line_label

                sheet.bb.calc_sizes()

                # ACTUAL
                act_cell = sheet.cell(column=actual_col.number(), row=finish.number())
                act_cell.set_explicit_value(act_detail_summation, data_type=TypeCodes.FORMULA)
                act_line.xl_data.cell = act_cell
                CellStyles.format_line(act_line)

                # FORECAST
                for_cell = sheet.cell(column=forecast_col.number(), row=finish.number())
                for_cell.set_explicit_value(for_detail_summation, data_type=TypeCodes.FORMULA)
                for_line.xl_data.cell = for_cell
                CellStyles.format_line(for_line)

    def _make_report_header(self, sheet, title, date):
        """
        Method formats header portion of report tabs.
        """
        sheet.sheet_properties.tabColor = self._report_tab_color
        SheetStyle.style_sheet(sheet)

        all_rows = sheet.bb.row_axis.add_group('all', offset=1)
        all_cols = sheet.bb.col_axis.add_group('all', size=8, offset=1)

        #  set up rows
        ban_row = all_rows.add_group('banner', size=1, offset=1)
        date_row = all_rows.add_group('report_date', size=1, offset=1)
        header_row = all_rows.add_group('column_headers', size=1)

        # set up columns
        space_left = all_cols.add_group('space_left', size=1)
        label_col = all_cols.add_group('labels', size=1)
        line_col = all_cols.add_group('lines', size=1)
        fore_col = all_cols.add_group('Forecast', size=1)
        actl_col = all_cols.add_group('Actual', size=1)
        delta_col = all_cols.add_group('Delta', size=1)
        perc_diff = all_cols.add_group('Percent Difference', size=1)
        space_right = all_cols.add_group('space_right', size=1)

        sheet.bb.calc_sizes()

        # Make Banner Cell
        # Get furthest left banner cell and format as desired
        banner_cell = sheet.cell(row=ban_row.number(), column=label_col.number())
        banner_cell.value = title
        CellStyles.format_header_label(banner_cell, alignment='center',
                                       color=self._banner_color,
                                       font_color=self._banner_font_color,
                                       bold=True, font_size=16, name='Arial')

        # merge banner cells
        st_col = get_column_letter(label_col.number())
        ed_col = get_column_letter(perc_diff.number())
        row_num = ban_row.number()
        merge_range = '%s%s:%s%s' % (st_col, row_num, ed_col, row_num)
        sheet.merge_cells(merge_range)

        # Set date cell (should just be its own row/col group with offset = 1)
        label_cell = sheet.cell(row=date_row.number(), column=label_col.number())
        label_cell.value = 'Date of collection:'
        CellStyles.format_bold(label_cell)

        value_cell = sheet.cell(row=date_row.number(), column=line_col.number())
        value_cell.value = date
        CellStyles.format_date(value_cell)
        CellStyles.format_bold(value_cell)
        CellStyles.format_alignment(value_cell, 'left')

        # Set header columns (should be own row group w/offset=1, col group with offset=3)
        groups = [fore_col, actl_col, delta_col, perc_diff]
        for col in groups:
            cell = sheet.cell(row=header_row.number(), column=col.number())
            cell.value = col.name
            CellStyles.format_header_label(cell, alignment='center',
                                           color=self._tbl_hdr_color,
                                           font_color=self._tbl_hdr_font_color)
            CellStyles.format_border(cell, bottom=True)

        # Set Column Widths
        SheetStyle.set_column_width(sheet, space_left.number(), 2.4)
        SheetStyle.set_column_width(sheet, space_right.number(), 2.4)
        for c in range(label_col.number(), perc_diff.number()+1):
            SheetStyle.set_column_width(sheet, c, 16.63)

    def _report_line(self, sheet, act_line, for_line, row_container, indent=0):
        """
        Method adds a specific LineItem to the report.  Method preserves
        original cell designation on the ``Forecast`` or ``Actual`` tab.
        Method adds links to source data for Forecast and Actual columns and
        inserts formulas to calculate Delta and Percent Difference.
        """

        all_cols = sheet.bb.col_axis.get_group('all')
        line_col = all_cols.get_group('lines')
        forecast_col = all_cols.get_group('Forecast')
        actual_col = all_cols.get_group('Actual')
        delta_col = all_cols.get_group('Delta')
        diff_col = all_cols.get_group('Percent Difference')

        details = act_line.get_ordered()

        if act_line.xl_format.blank_row_before and not details:
            # if row_container.groups or not row_container.offset:
            sheet.bb.need_spacer = True

        line_label = indent * " " + act_line.title
        line_rows = row_container.add_group(act_line.title, offset=int(sheet.bb.need_spacer))

        if sheet.bb.need_spacer:
            row_size = row_container.calc_size()
            r = sheet.row_dimensions[row_container.tip + row_size]
            r.outline_level = sheet.bb.outline_level

        sheet.bb.need_spacer = False

        # a line with own content should have no children with own content,
        # and should not consolidate
        if details and act_line.sum_details:
            self._add_details(
                sheet=sheet,
                act_line=act_line,
                for_line=for_line,
                row_container=line_rows,
                indent=indent,
            )
        else:
            if details:
                self._add_details(
                    sheet=sheet,
                    act_line=act_line,
                    for_line=for_line,
                    row_container=line_rows,
                    indent=indent,
                )

            # this is the logic for lines without details
            line_row = line_rows.add_group(act_line.title, size=1)
            sheet.bb.calc_sizes()

            # need to write actual and forecast values here (link to source)
            formula_string = '=%s' % act_line.xl_data.get_coordinates(
                include_sheet=True)
            act_cell = sheet.cell(row=line_row.number(),
                                  column=actual_col.number())
            act_cell.set_explicit_value(formula_string,
                                        data_type=TypeCodes.FORMULA)
            act_line.xl_data.cell = act_cell
            CellStyles.format_line(act_line)

            formula_string = '=%s' % for_line.xl_data.get_coordinates(
                include_sheet=True)

            for_cell = sheet.cell(row=line_row.number(),
                                  column=forecast_col.number())
            for_cell.set_explicit_value(formula_string,
                                        data_type=TypeCodes.FORMULA)
            for_line.xl_data.cell = for_cell
            CellStyles.format_line(for_line)

        sheet.bb.calc_sizes()
        # *************************************************************
        line_row = line_rows.get_group(act_line.title)
        # write line label
        label_cell = sheet.cell(row=line_row.number(), column=line_col.number())

        label_cell.value = line_label
        # *************************************************************

        # *************************************************************
        # Do the actual work here
        materials = dict()
        materials['actual'] = act_line.xl_data.get_coordinates(include_sheet=False)
        materials['forecast'] = for_line.xl_data.get_coordinates(include_sheet=False)
        materials['placeholder'] = self._placeholder

        delta_cell = sheet.cell(row=line_row.number(), column=delta_col.number())
        temp = FormulaTemplates.REPORT_DELTA
        formula_string = temp.format(**materials)
        delta_cell.set_explicit_value(formula_string,
                                      data_type=TypeCodes.FORMULA)

        working_row = delta_cell.row

        save_cell = act_line.xl_data.cell
        act_line.xl_data.cell = delta_cell
        CellStyles.format_line(act_line)
        act_line.xl_data.cell = save_cell

        diff_cell = sheet.cell(row=line_row.number(), column=diff_col.number())
        temp = FormulaTemplates.REPORT_DIFF
        formula_string = temp.format(**materials)
        diff_cell.set_explicit_value(formula_string,
                                     data_type=TypeCodes.FORMULA)
        CellStyles.format_parameter(diff_cell)
        diff_cell.number_format = NumberFormats.PERCENT_FORMAT
        # *************************************************************

        r = sheet.row_dimensions[working_row]
        r.outline_level = sheet.bb.outline_level

        if act_line.xl_format.blank_row_after:
            sheet.bb.need_spacer = True
        else:
            sheet.bb.need_spacer = False

        row_container.calc_size()

    def _report_statement(self, sheet, act_statement, for_statement, name=None):
        """
        Method adds a financial statement to the report, delegating to
        the _report_line() method for most of the work.
        """
        all_cols = sheet.bb.col_axis.get_group('all')
        lab_col = all_cols.get_group('labels')

        all_rows = sheet.bb.row_axis.get_group('all')

        if name is None:
            name = act_statement.name

        stat_rows = all_rows.add_group(name, offset=1)

        lab_row = stat_rows.add_group('label', size=1)
        sheet.bb.calc_sizes()

        cell = sheet.cell(row=lab_row.number(), column=lab_col.number())
        cell.value = name.title()
        CellStyles.format_bold(cell)

        for act_line in act_statement.get_ordered():
            sheet.bb.outline_level = 1
            for_line = for_statement.find_first(act_line.name)
            if for_line is not None:
                self._report_line(sheet, act_line, for_line, stat_rows)

                sheet.bb.calc_sizes()
