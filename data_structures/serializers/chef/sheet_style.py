# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.sheet_style
"""

Module defines a class that stores standard row and column names.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
SheetStyle            standard styles for worksheets
====================  =========================================================
"""




# Imports
import openpyxl as xlio

from chef_settings import COLUMN_WIDTH

from .cell_styles import CellStyles
from .field_names import FieldNames




# Constants
SHOW_GRID_LINES = False
ZOOM_SCALE = 80

# Module Globals
get_column_letter = xlio.utils.get_column_letter

# Classes
class SheetStyle:
    """

    Class stores standard keys for row and column lookups within SheetData.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    label_areas()         method labels all Areas on the sheet
    set_column_width()    method sets the column width for provided column
    style_sheet()         method formats sheet view, sets zoom, turns off lines
    ====================  =====================================================
    """
    @staticmethod
    def label_areas(sheet):
        """


        SheetStyle.label_areas() -> None

        --``sheet`` is an openpyxl worksheet

        Method iterates over all Areas attached to sheet.bb and labels them.
        Method excludes the General and Timeline areas which are standard on
        many sheets but do not get labeled by default.
        """
        areas_exclude = set((FieldNames.GENERAL, FieldNames.TIMELINE))
        areas = set(sheet.bb.area_names)
        areas = areas - areas_exclude

        for name in areas:
            area = getattr(sheet.bb, name)
            if not area.rows.by_name:
                continue

            row_num = min(area.rows.by_name.values()) - 1
            CellStyles.format_area_label(sheet, name, row_num)

    @staticmethod
    def set_column_width(sheet, column, width=COLUMN_WIDTH):
        """


        SheetStyle.set_column_width() -> None

        --``sheet`` is an openpyxl worksheet
        --``column`` is the number of the column you which to format
        --``width`` is the width to set on the column

        Method sets the Excel column width the for column.
        """
        column = sheet.column_dimensions[get_column_letter(column)]
        column.width = width

    @staticmethod
    def style_sheet(sheet, label_areas=True):
        """


        SheetStyle.style_sheet() -> None

        --``sheet`` is an openpyxl worksheet
        --``label_areas`` whether or not to label areas, default is true

        Method labels areas on sheet (label_areas == True) and implements
        Blackbird sheet styling.
        """
        if label_areas:
            SheetStyle.label_areas(sheet)

        sheet.sheet_view.showGridLines = SHOW_GRID_LINES
        sheet.sheet_view.zoomScale = ZOOM_SCALE
