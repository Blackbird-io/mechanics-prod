# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef._chef_tools
"""

Module defines functions that are tools for Chef module.
====================  ==========================================================
Attribute             Description
====================  ==========================================================

DATA:
n/a

FUNCTIONS:
add_links_to_selectors function adds VB macros to link cells in sheets
add_scenario_selector function adds a scenario selector cell to the sheet
calculate_formulas    function opens-saves-closes Excel book do calculations
check_alignment       function checks that line is correctly aligned
check_filename_ext    function checks that file has correct extension
close_excel_by_force  function closes Excel application by force
collapse_groups       function opens an Excel file and collapses all groups
group_lines           function adds row to outline group for pretty collapsing
is_close              function for fuzzy equals between numeric values
rows_to_coordinates   creates directory of row name:row coordinate in a column
set_label             function sets the label for a given row
set_param_rows        function sets the SSOT parameter rows for a line item

CLASSES:
n/a
====================  ==========================================================
"""




# Imports
import os
import openpyxl as xlio
import re

from bb_exceptions import ExcelPrepError
from chef_settings import SCENARIO_SELECTORS, FILTER_PARAMETERS

if SCENARIO_SELECTORS:
    import win32api
    import win32com.client as win32
    import win32con
    import win32gui
    import win32process

from openpyxl.worksheet.datavalidation import DataValidation

from .cell_styles import CellStyles
from .field_names import FieldNames




# Constants
_COLLAPSE_GROUPS_VBS_FILE = "excel_collapse_pretty_rows.vbs"
_OPEN_SAVE_CLOSE_VBS_FILE = "open_save_close_excel.vbs"
_VBS_FILENAME_BOOKMARK = "FILENAME_PLACEHOLDER"
_VBS_PATH = os.path.dirname(os.path.realpath(__file__))

# Module Globals
get_column_letter = xlio.utils.get_column_letter

# Classes
# n/a

# Functions
def add_links_to_selectors(filename, sources_dict):
    """


    add_links_to_selectors() -> None

    --``filename`` is the xlsx file (Chef chopped) to work on
    --``sources_dict`` is a dictionary containing the Excel sheet names and
       selector cell coordinates to link

    Method adds VB code to a completed Chef workbook to link scenario
    selection cells together to allow universal scenario change from any
    sheet in the workbook.
    """
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    xl.Visible = False

    ss = xl.Workbooks.Open(filename)

    all_sheets = set(sources_dict.keys())
    for sheet in all_sheets:
        macro_code = []
        cos = sources_dict[sheet][1]
        col = cos[0]
        row = cos[1]
        line = '    If Target.Address = "$%s$%s" Then' % (col, row)
        macro_code.append(line)

        line = '        Application.EnableEvents = False'
        macro_code.append(line)

        other_sheets = all_sheets - set(sheet)
        for o in other_sheets:
            name = sources_dict[o][0]
            cos = sources_dict[o][1]
            col = cos[0]
            row = cos[1]
            line = '        ThisWorkbook.Sheets("%s").Range("$%s$%s").Value = Target.Value' % (name, col, row)
            macro_code.append(line)

        line = '        Application.EnableEvents = True'
        macro_code.append(line)

        line = '    End If'
        macro_code.append(line)

        comp = ss.VBProject.VBComponents(sheet)
        module = comp.CodeModule
        line_num = module.CreateEventProc("Change", "Worksheet")

        for i, line in enumerate(macro_code):
            module.InsertLines(line_num+i+1, line)

    xl.DisplayAlerts = False

    newfile = check_filename_ext(filename, "xlsm")
    ss.SaveAs(newfile, FileFormat=52)

    ss.Close()

    xl.Quit()
    close_excel_by_force(xl)

    # delete original file
    os.remove(filename)

    return newfile


def add_scenario_selector(sheet, column, row, selections):
    """


    add_scenario_selector() -> None

    --``filename`` is the xlsx file (Chef chopped) to work on
    --``sources_dict`` is a dictionary containing the Excel sheet names and
       selector cell coordinates to link

    Method adds VB code to a completed Chef workbook to link scenario
    selection cells together to allow universal scenario change from any
    sheet in the workbook.
    """

    select_column = column + 2

    options = ','.join(selections)
    dv = DataValidation(type="list",
                        formula1='"%s"' % options,
                        allow_blank=False)

    # Optionally set a custom error message
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    # Optionally set a custom prompt message
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    sheet.add_data_validation(dv)

    selector_cell = sheet.cell(column=select_column, row=row)
    selector_cell.value = "Base"

    dv.add(selector_cell)
    sheet.bb.scenario_selector = selector_cell.coordinate

    # Make label cells and drop-down selector
    CellStyles.format_scenario_selector_cells(sheet,
                                               column,
                                               select_column,
                                               row)


def calculate_formulas(filename):
    """


    update_formulas -> None

    --``filename`` must be the string path for the file to check, real path

    This function opens an Excel file, saves, and closes the file.
    """
    _write_run_temp_vbs_file(filename, _OPEN_SAVE_CLOSE_VBS_FILE)


def check_alignment(line):
    """


    check_alignment() -> bool

    --``line`` must be a BB LineItem object

    Function checks that ``line`` is aligned with itself from prior periods.
    Function prints misalignments to screen and returns False when detected.
    """
    cell = line.xl_data.cell
    sheet = cell.parent

    try:
        check_data, xl_format = sheet.bb.line_directory[line.id.bbid]
    except KeyError:
        # sheet.bb.line_directory[line.id.bbid] = line.xl_data
        # check_data = sheet.bb.line_directory[line.id.bbid]
        check_data = None

    result = True
    if check_data:
        if check_data.cell.row != cell.row:
            print(line)
            print(cell)
            print("Correct row: %s" % check_data.cell.row)
            print("Current row: %s" % cell.row)

            c = 'Misalignment!  This line should not be written to ' \
                'the current row.'
            print(c)
            result = False

    return result


def check_filename_ext(filename, ext):
    """


    check_filename_ext() -> str

    --``filename`` proposed filename for workbook
    --``ext`` correct file extension for workbook

    Function ensures filename complies with specified extension and corrects
    if not.  Function returns correct file path.
    """
    temp = filename.strip().split('.')
    check_ext = temp[-1]

    if check_ext == ext:
        result = filename
    else:
        base = filename[0:-len(check_ext)]
        result = base + ext

    return result


def close_excel_by_force(excel):
    """


    close_excel_by_force() -> None

    --``excel`` must be an Excel Application instance

    Method closes all Excel instances by brute force. No other way to close
    out all lingering threads.
    Source: http://stackoverflow.com/questions/10221150/cant-close-excel-completely-using-win32com-on-python
    """

    # Get the window's process id's
    hwnd = excel.Hwnd
    t, p = win32process.GetWindowThreadProcessId(hwnd)

    # Ask window nicely to close
    win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)

    # If the application didn't close, force close
    try:
        handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, p)
        if handle:
            win32api.TerminateProcess(handle, 0)
            win32api.CloseHandle(handle)
    except:
        pass


def collapse_groups(filename):
    """


    collapse_groups -> None

    --``filename`` must be the string path for the file to check

    This function opens an Excel file, collapses all columns and rows, saves,
    and closes the file.
    """
    _write_run_temp_vbs_file(filename, _COLLAPSE_GROUPS_VBS_FILE)


def group_lines(sheet, row=None):
    """


    group_lines() -> None


    --``sheet`` must be an instance of openpyxl Worksheet

    Group lines in sheet to outline level specified by sheet data
    """
    if not row:
        row = sheet.bb.current_row

    sheet.row_dimensions[row].outline_level = sheet.bb.outline_level


def is_close(a, b, rel_tol=1e-09, abs_tol=0.0):
    """


    is_close -> bool

    --``a``, ``b`` are numeric values to compare
    --``rel_tol`` is the relative allowable tolerance, taken as fraction of the
        larger of a and b
    --``abs_tol`` is the absolute allowable tolerance

    This method provides a means for comparing two numbers by "fuzzy equals".
    If the absolute difference between ``a`` and ``b`` is less than or equal to
    the maximum of the specified relative tolerance and absolute tolerance,
    method returns True.  This method was copied from documentation for
    math.isclose() method in Python 3.5.

    """
    try:
        res = abs(a - b) <= max(rel_tol * max(abs(a), abs(b)), abs_tol)
    except TypeError:
        res = False

    return res


def rows_to_coordinates(*pargs, lookup, column):
    """


    LineChef._rows_to_coordinates -> dict

    --``lookup`` is a lookup table
    --``column`` must be a column number reference

    """

    result = dict()
    alpha_column = get_column_letter(column)

    for k in lookup.by_name:
        row = lookup.get_position(k)
        result[k] = alpha_column + str(row)

    return result


def set_label(*pargs, label, sheet, row, column=None,
              overwrite=False,
              formatter=None):
    """


    LineChef._set_label() -> Worksheet

    --``label`` must be value to set as a cell label
    --``sheet`` must be an instance of openpyxl Worksheet
    --``row`` must be a row index where ``label`` should be written
    --``column`` column index or None where ``label`` should be written
    --``overwrite`` must be a boolean; True overwrites existing value,
       if any; default is False
    --``formatter`` method in cell_styles to be called on the label cell

    Set (column, row) cell value to label. Throw exception if cell already
    has a different label, unless ``overwrite`` is True.

    If ``column`` is None, method attempts to locate the labels column in
    the sheet.bb.parameters area.
    """
    if column is None:
        if getattr(sheet.bb, FieldNames.PARAMETERS, None):
            param_area = getattr(sheet.bb, FieldNames.PARAMETERS)
            cols = param_area.columns
            column = cols.get_position(FieldNames.LABELS)
        else:
            column = 2

    label_cell = sheet.cell(column=column, row=row)
    existing_label = label_cell.value

    if overwrite or not existing_label:
        label_cell.value = label
    else:
        if existing_label != label:
            c = """
                Something is wrong with our alignment. We are trying to
                write a parameter to an existing row with a different label.
                """
            print(c)
            print("Old Label:", existing_label)
            print("New Label:", label)
            print("Sheet:", sheet.title)
            print("Row:", row)
            print("Col:", column)

            # raise ExcelPrepError(c)
            # Check to make sure we are writing to the right row; if the
            # label doesn't match, we are in trouble.

    if formatter:
        formatter(label_cell)

    return sheet


def set_param_rows(line, sheet):
    """

    set_param_rows() -> None

    --``line`` is an instance of LineItem
    --``sheet`` is an instance of Openpyxl worksheet

    Function determines the SSOT parameters that the line's drivers will need
    in perpetuity.  Template parameters are saved to the sheet's line directory
    """

    try:
        template_xl, xl_format = sheet.bb.line_directory[line.id.bbid]
    except KeyError:
        template_xl = None

    if template_xl:
        for check_data, driver_data in zip(template_xl.derived.calculations,
                                           line.xl_data.derived.calculations):
            if check_data.name != driver_data.name:
                c = "Formulas are not consistent for line over time!"
                raise ExcelPrepError(c)

            driver_data.rows = check_data.rows
    elif FILTER_PARAMETERS:
        # get params to keep
        for driver_data in line.xl_data.derived.calculations:
            params_keep = []
            for step in driver_data.formula.values():
                temp_step = [m.start() for m in re.finditer('parameters\[*',
                                                            step)]

                for idx in temp_step:
                    jnk = step[idx:]
                    idx_end = jnk.find(']') + idx
                    params_keep.append(step[idx + 11:idx_end])

        # clean driver_data
        new_rows = []
        for item in driver_data.rows:
            if item[FieldNames.LABELS] in params_keep:
                new_rows.append(item)

            try:
                temp = driver_data.conversion_map[item[FieldNames.LABELS]]
            except KeyError:
                pass
            else:
                if temp in params_keep:
                    new_rows.append(item)

        driver_data.rows = new_rows

    #*************************************************************************#
    #                          NON-PUBLIC METHODS                             #
    #*************************************************************************#

def _write_run_temp_vbs_file(filename, vbs_file):
    """


    _write_run_temp_vbs_file -> None

    --``filename`` must be the string path for the file being checked

    Function updates template VBScript with file being checked, opens file,
    saves file, and closes file so that Excel will calculate formula values.
    """

    # open and read original VBS file
    orig_path = os.path.join(_VBS_PATH, vbs_file)
    orig_file = open(orig_path, mode='r')
    orig_lines = orig_file.readlines()
    orig_file.close()

    # write temporary VBS file with correct filepath
    i = 0
    while True:
        temp_fnam = vbs_file[:-4] + "_temp_%s.vbs" % i
        temp_path = os.path.join(_VBS_PATH, temp_fnam)
        if os.path.isfile(temp_path):
            i += 1
        else:
            break

    temp_file = open(temp_path, mode='w')

    for line in orig_lines:
        if _VBS_FILENAME_BOOKMARK in line:
            line = line.replace(_VBS_FILENAME_BOOKMARK, filename)

        temp_file.write(line)

    temp_file.close()

    # run the VBS file
    run_path = '"' + temp_path + '"'
    os.system(run_path)

    # delete the temporary VBS file
    os.remove(temp_path)
