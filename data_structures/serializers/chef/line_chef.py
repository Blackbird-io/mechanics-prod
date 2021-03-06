# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.line_chef
"""

Module defines LineChef class which converts Statement and LineItem objects
into dynamic Excel format.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
LineChef              class with methods to chop BB statements into dynamic
                      Excel structures
====================  =========================================================
"""




# Imports
import itertools
import openpyxl as xlio

from bb_exceptions import ExcelPrepError, BBAnalyticalError
from chef_settings import (
    COMMENT_FORMULA_NAME, COMMENT_FORMULA_STRING, COMMENT_CUSTOM,
)
from data_structures.modelling.line_item import LineItem
from openpyxl.comments import Comment

from ._chef_tools import (
    group_lines, check_alignment, set_param_rows, rows_to_coordinates
)
from .cell_styles import CellStyles
from .data_types import TypeCodes
from .field_names import FieldNames
from .formulas import FormulaTemplates




# Constants
# n/a

# Module Globals
get_column_letter = xlio.utils.get_column_letter

# Classes
class LineChef:
    """

    Class packages Statement and LineItem objects into dynamic Excel workbook.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    values_only           whether to write full logic to Excel or only values

    FUNCTIONS:
    attempt_reference_resolution() tries to resolve missing line refs in formulas
    chop_line()           writes LineItems to Excel
    chop_statement()      writes Statements to Excel (except Starting Balance)
    ====================  =====================================================
    """
    def __init__(self, values_only=False, include_ids=False):
        self.values_only = values_only
        self.include_ids = include_ids

    @staticmethod
    def attempt_reference_resolution(sheet, calc, materials):
        """

        LineChef.attempt_reference_resolution() -> None

        --``sheet`` must be an instance of openpyxl Worksheet
        --``calc`` is a driver calculation
        --``materials`` is the info dictionary used to assign line references
                        for driver calculation

        Method attempts to reevaluate problem lines and update their
        problematic line references:
            - update line_coordinates dictionary (copy from below)
            - find first line where steps were written
            - reformat formulas using updated materials dict and save
              values
        """

        # Update materials
        new_lines = dict()
        for key, line in calc.references.items():
            try:
                include = line.xl_data.cell.parent is not sheet
                new_lines[key] = line.xl_data.get_coordinates(include_sheet=include)
            except (ExcelPrepError, AttributeError):
                msg = 'Cannot resolve bad line reference.'
                print(msg)
                print("Name:     ", calc.name)
                print("Template: ", calc.formula)
                print(line)

                new_lines[key] = "#REF!"
                # raise ExcelPrepError(msg)

        materials['lines'] = new_lines

        # now reformat all formulas in steps
        for step in calc.formula:
            # see if step in materials['steps'], if not, it was purposely
            # omitted so skip it!
            if step not in materials['steps']:
                continue

            cell_coos = materials['steps'][step]
            cell = sheet[cell_coos]

            formula_string = calc.formula[step]
            try:
                formula_string = formula_string.format(**materials)
            except Exception as X:
                print("Name:     ", calc.name)
                print("Step:     ", step)
                print("Template: ", formula_string)

                raise ExcelPrepError

            cell.set_explicit_value(formula_string,
                                    data_type=TypeCodes.FORMULA)

    def chop_line(
        self, sheet, column, line, row_container,
        set_labels=True, indent=0, check=True, start_bal=False,
    ):
        """


        LineChef.chop_line() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Method walks through LineItems and their details and converts them to
        dynamic links in Excel cells.  Method adds consolidation and derivation
        logic to cells.

        Routines deliver sheet with the current_row pointing to the last filled
        in cell.
        """
        details = line.get_ordered()

        if line.xl_format.blank_row_before and not details:
            # if row_container.groups or not row_container.offset:
            sheet.bb.need_spacer = True

        matter = row_container.add_group(
            line.title, offset=int(sheet.bb.need_spacer), bbid=line.id.bbid.hex
        )

        if sheet.bb.need_spacer:
            row_size = row_container.calc_size()
            r = sheet.row_dimensions[row_container.tip + row_size]
            r.outline_level = sheet.bb.outline_level

        sheet.bb.need_spacer = False

        if details:
            self._add_details(
                sheet=sheet,
                column=column,
                line=line,
                row_container=matter,
                set_labels=set_labels,
                indent=indent,
                check=check,
                start_bal=start_bal,
            )

        elif start_bal:
            if line.xl_data.cell:
                # here just link the current cell to the cell in line.xl_data.cell
                old_cell = line.xl_data.cell
                line.xl_data.reference.source = line
                self._add_reference(
                    sheet=sheet,
                    column=column,
                    line=line,
                    indent=indent,
                    row_container=matter
                )

                CellStyles.format_line(line)

                line.xl_data.reference.source = None
                line.xl_data.reference.cell = None
                line.xl_data.cell = old_cell

        if not start_bal:
            self._add_reference(
                sheet=sheet,
                column=column,
                line=line,
                indent=indent,
                row_container=matter
            )

            self._add_derivation_logic(
                sheet=sheet,
                column=column,
                line=line,
                indent=indent,
                row_container=matter,
                set_labels=set_labels,
            )

            self._add_consolidation_logic(
                sheet=sheet,
                column=column,
                line=line,
                indent=indent,
                row_container=matter,
                set_labels=set_labels,
            )

        if start_bal and (details or line.xl_data.cell):
            run_segments = False
        else:
            run_segments = True

        if run_segments and not line.xl_data.reference.source:
            self._combine_segments(
                sheet=sheet,
                column=column,
                line=line,
                indent=indent,
                row_container=matter,
            )

        CellStyles.format_line(line)

        # for row alignment
        if check:
            check_alignment(line)

        if check and not start_bal:
            if line.id.bbid not in sheet.bb.line_directory.keys():
                sheet.bb.line_directory[line.id.bbid] = (line.xl_data,
                                                         line.xl_format)

        r = sheet.row_dimensions[line.xl_data.cell.row]
        r.outline_level = sheet.bb.outline_level

        if line.xl_format.blank_row_after:
            sheet.bb.need_spacer = True
        else:
            sheet.bb.need_spacer = False

        row_container.calc_size()

        return sheet

    def chop_statement(
        self, sheet, column, statement, row_container=None,
        start_bal=False, title=None, set_labels=False,
    ):
        """


        LineChef.chop_statement() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``statement`` must be an instance of Statement

        Method walks through Statement lines and delegates LineChef.chop_line()
        to add them as dynamic links in Excel.
        """
        matter = row_container.get_group('lines')
        sheet.bb.need_spacer = False

        check = statement.name != 'ending balance sheet'
        for line in statement.get_ordered():
            self.chop_line(
                sheet=sheet,
                column=column,
                line=line,
                row_container=matter,
                check=check,
                start_bal=start_bal,
                set_labels=set_labels,
            )
            row_container.calc_size()

        if not matter.groups:
            matter.size = 1

        return sheet

    # *************************************************************************#
    #                           NON-PUBLIC METHODS                             #
    # *************************************************************************#

    def _validate_consolidation(self, sheet, line):
        """


        LineChef._validate_consolidation() -> None

        --``line`` must be an instance of LineItem

        Throws an error if all are true:
        1. ``line`` has own content: details, drivers, or value.
        2. ``line`` has consolidation sources.
        3. sources have own content.
        """
        if line.has_own_content():
            for sub in line.xl_data.consolidated.sources:
                if sub.has_own_content():
                    c = (
                        'line "{}" on sheet "{}" has content '
                        'but also tries to consolidate line "{} '
                        'which has its own content'
                    ).format(line.title, sheet.title, sub.title)

                    raise BBAnalyticalError(c)

    def _add_consolidation_logic(
        self, sheet, column, line, set_labels=True, indent=0, row_container=None
    ):
        """


        LineChef._add_consolidation_logic() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Expects line.xl_data.consolidated.sources to include full range of pointers
        to source lines on children.

        Always stuffs consolidation into the same number of rows.
        Derive can still cause staircasing if the line picks up details in the
        future that it doesn't have now.

        Returns Worksheet with consolidation logic added as Excel dynamic links
        """
        if line.has_own_content():
            # a line with own content should have no children with own content
            # throw an error if any of consolidation sources have content
            self._validate_consolidation(sheet, line)
        elif line.xl_data.consolidated.sources:
            sources = line.xl_data.consolidated.sources
            labels = line.xl_data.consolidated.labels
            required_rows = sheet.bb.consolidation_size
            n_sources = len(sources)
            links_per_cell = n_sources // required_rows
            links_per_cell = max(1, links_per_cell)
            # Make sure we include at least 1 link per cell.

            link_template = FormulaTemplates.ADD_COORDINATES
            line.xl_data.consolidated.array.clear()

            # the outer iter converts the sorted list into an iterator
            # which is needed to chunk it by group
            paired = iter(sorted(zip(labels, sources)))
            # chunker gives (label, source) pairs in chunks of "links_per_cell".
            # Each hit returns a links_per_cell-sized tuple of tuples.
            chunker = itertools.repeat(paired, links_per_cell)
            cons_rows = row_container.add_group('cons')
            for source_list in itertools.zip_longest(*chunker):
                batch_summation = ""

                for label, source_line in source_list:
                    if label:
                        sub_indent = indent + LineItem.TAB_WIDTH * 2
                        label_line = (sub_indent * " ") + label
                    include = source_line.xl_data.cell.parent is not sheet
                    source_cos = source_line.xl_data.get_coordinates(
                        include_sheet=include
                    )
                    link = link_template.format(coordinates=source_cos)
                    batch_summation += link

                if batch_summation:
                    finish = cons_rows.add_group(
                        label_line, size=1, label=label_line, outline=1,
                        formatter=CellStyles.format_consolidated_label,
                        bbid=line.id.bbid.hex
                    )
                    batch_cell = sheet.cell(column=column, row=finish.number())
                    batch_cell.set_explicit_value(
                        batch_summation,
                        data_type=TypeCodes.FORMULA
                    )
                    line.xl_data.consolidated.array.append(batch_cell)

            row_container.calc_size()

            # Group the cells
            line.xl_data.consolidated.starting = line.xl_data.consolidated.array[0].row
            line.xl_data.consolidated.ending = line.xl_data.consolidated.array[-1].row
            alpha_column = get_column_letter(column)
            summation_params = {
                "starting_row": line.xl_data.consolidated.starting,
                "ending_row": line.xl_data.consolidated.ending,
                "alpha_column": alpha_column
            }

            summation = FormulaTemplates.SUM_RANGE.format(**summation_params)
            line_label = indent * " " + line.title

            finish = row_container.add_group(
                line.title, size=1, label=line_label, bbid=line.id.bbid.hex
            )
            summation_cell = sheet.cell(column=column, row=finish.number())
            summation_cell.set_explicit_value(
                summation, data_type=TypeCodes.FORMULA
            )
            line.xl_data.consolidated.cell = summation_cell
            line.xl_data.cell = summation_cell

        return sheet

    def _add_derivation_logic(
        self, sheet, column, line, set_labels=True, indent=0, row_container=None
    ):
        """

        LineChef._add_derivation_logic() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Delegates to LineChef._add_driver_calculation() to add driver logic
        to dynamic links in Excel cells.
        """

        if not line.xl_data.derived.calculations:
            pass
        else:
            set_param_rows(line, sheet)
            for data_cluster in line.xl_data.derived.calculations:
                self._add_driver_calculation(
                    sheet=sheet,
                    column=column,
                    line=line,
                    driver_data=data_cluster,
                    set_labels=set_labels,
                    indent=indent,
                    row_container=row_container
                )

            # NOTE: No summation at the end of the derive process. Final
            # derived value may OR MAY NOT be the sum of priors. Up to
            # driver to decide.

        return sheet

    def _add_driver_calculation(self, *pargs, sheet, column, line, driver_data,
                                set_labels=True, indent=0, row_container=None):
        """


        LineChef._add_driver_calculation -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``driver_data`` must be a dictionary of driver_data info
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Writes driver logic to Excel sheet.
        """
        private_data = getattr(sheet.bb, FieldNames.PARAMETERS).copy()
        # Set up a private range that's going to include both "shared" period &
        # unit parameters from the column and "private" driver parameters.

        period_column = column
        param_rows = row_container.add_group('params', size=0)

        for row_data in sorted(driver_data.rows,
                               key=lambda x: x[FieldNames.LABELS]):

            private_label = row_data[FieldNames.LABELS]
            private_value = row_data[FieldNames.VALUES]

            line_label = (indent + LineItem.TAB_WIDTH) * " " + private_label

            group_lines(sheet)
            finish = param_rows.add_group(
                line_label, size=1, label=line_label, outline=0,
                bbid=line.id.bbid.hex
            )
            param_cell = sheet.cell(column=period_column, row=finish.number())

            CellStyles.format_parameter(param_cell)
            CellStyles.format_hardcoded(param_cell)

            if isinstance(private_value, (list, set, dict, map)):
                private_value = str(private_value)
                # Capture mutable and multi-dimensional objects here. We need
                # to figure out how to represent them in Excel later.

            param_cell.value = private_value

            relative_position = finish.number() \
                - (private_data.rows.starting or 0)
            private_data.rows.by_name[private_label] = relative_position
            # ... In this particular case, we could map a specific cell
            # (in memory) to the parameter. Unclear whether that's useful
            # though, because we generally look up locations for lines, not
            # parameters. And lines continue to span several rows.

        # Transform the range values from rows to coordinates
        param_coordinates = rows_to_coordinates(
            lookup=private_data.rows,
            column=period_column)

        # Apply param:var conversions so formula can find its expected inputs
        for param_name, var_name in driver_data.conversion_map.items():
            param_coordinates[var_name] = param_coordinates[param_name]

        # Finally, format the formula as necessary
        # (if references are a dict of objects, could map each obj to its
        #  coordinates)
        line_coordinates = dict()
        problem_line = False
        for k, obj in driver_data.references.items():
            try:
                include = obj.xl_data.cell.parent is not sheet
                line_coordinates[k] = obj.xl_data.get_coordinates(
                    include_sheet=include
                )
            except (ExcelPrepError, AttributeError):
                # set flag for saving problem line
                problem_line = True

                # as a temporary measure, introduce a placeholder for line
                # coordinates
                line_coordinates[k] = '#N/A'  # placeholder

        materials = dict()
        materials["lines"] = line_coordinates
        materials['parameters'] = param_coordinates

        try:
            life_coordinates = rows_to_coordinates(
                lookup=sheet.bb.life.rows,
                column=period_column)
        except AttributeError:
            pass
        else:
            materials["life"] = life_coordinates

        try:
            event_coordinates = rows_to_coordinates(
                lookup=sheet.bb.events.rows,
                column=period_column)
        except AttributeError:
            pass
        else:
            materials["events"] = event_coordinates

        try:
            size = getattr(sheet.bb, FieldNames.SIZE)
            size_coordinates = rows_to_coordinates(
                lookup=size.rows, column=period_column
            )
        except AttributeError:
            pass
        else:
            materials["size"] = size_coordinates

        materials["steps"] = dict()

        n_items = len(driver_data.formula.items())
        count = 0
        row_container.calc_size()

        formula_steps = self._get_formula_steps(sheet, line, driver_data)
        for key in formula_steps:
            count += 1
            # all but the last step are indented an extra level
            if count < n_items:
                line_label = (indent + LineItem.TAB_WIDTH) * " " + key
                outline = sheet.bb.outline_level + 1
            else:
                line_label = indent * " " + line.title
                outline = sheet.bb.outline_level
            finish = row_container.add_group(
                key, size=1, label=line_label, outline=outline,
                bbid=line.id.bbid.hex
            )

            try:
                template = driver_data.formula[key]
            except KeyError:
                continue

            if ':' in template:
                new_materials = dict()
                template = template.replace(':', '#')
                for pk in materials:
                    new_materials[pk] = dict()
                    for k in materials[pk]:
                        if ':' in k:
                            new_k = k.replace(':', '#')
                        else:
                            new_k = k

                        new_materials[pk][new_k] = materials[pk][k]

                materials = new_materials

            try:
                formula = template.format(**materials)
            except Exception as X:
                print("Name:     ", driver_data.name)
                print("Template: ", driver_data.formula)
                raise ExcelPrepError

            calc_cell = sheet.cell(column=period_column, row=finish.number())
            calc_cell.set_explicit_value(formula, data_type=TypeCodes.FORMULA)

            # add current step to materials dictionary
            materials["steps"][key] = calc_cell.coordinate

            if set_labels and (COMMENT_FORMULA_NAME or
                               COMMENT_FORMULA_STRING or
                               COMMENT_CUSTOM):
                c = ""

                if COMMENT_FORMULA_NAME:
                    c += "Formula name: " + driver_data.name + "\n"

                if COMMENT_FORMULA_STRING:
                    c += "Formula string: " + template + "\n"

                if COMMENT_CUSTOM:
                    c += "Comment: " + driver_data.comment + "\n"

                a = "LineChef"
                calc_cell.comment = Comment(c, a)

            CellStyles.format_calculation(calc_cell)
            # If formula included a reference to the prior value of the line
            # itself, it's picked up here. Can now change line.xl_data.derived.final

            # we don't want a blank row between the calc cell and the summary
            # cell
            line.xl_data.derived.ending = finish.number()
            line.xl_data.derived.cell = calc_cell
            line.xl_data.cell = calc_cell

            if problem_line:
                # save tuple of line and derivation materials
                info = (driver_data, materials)
                sheet.bb.problem_lines.append(info)

        return sheet

    def _add_reference(
        self, sheet, column, line, row_container=None,
        indent=0, update_cell=True
    ):
        """


        LineChef._add_reference() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Adds a single cell reference to a new cell.
        (e.g. new_cell.value = '=C18')
        """
        tl = line.period.relationships.parent
        last_date = max(tl.keys())
        last_period = line.period is tl.find_period(last_date)
        actual = tl.name == 'actual'
        monitor = line.usage.monitor

        allowed = not self.values_only or (last_period and monitor and actual)
        if line.xl_data.reference.direct_source and allowed:
            line_label = indent * " " + line.title  # + ': ref'
            finish = row_container.add_group(
                line.title, size=1, label=line_label, bbid=line.id.bbid.hex
            )
            cell = sheet.cell(column=column, row=finish.number())

            excel_str = "=" + line.xl_data.reference.direct_source

            cell.set_explicit_value(excel_str, data_type=TypeCodes.FORMULA)

            line.xl_data.ending = finish.number()
            line.xl_data.reference.cell = cell

            if update_cell:
                line.xl_data.cell = cell
        elif line.xl_data.reference.source:
            line_label = indent * " " + line.title  # + ': ref'

            finish = row_container.add_group(
                line.title, size=1, label=line_label, bbid=line.id.bbid.hex
            )
            cell = sheet.cell(column=column, row=finish.number())

            ref_cell = line.xl_data.reference.source.xl_data.cell
            include = ref_cell.parent is not sheet
            source = line.xl_data.reference.source
            excel_str = "=" + source.xl_data.get_coordinates(include_sheet=include)

            cell.set_explicit_value(excel_str, data_type=TypeCodes.FORMULA)

            line.xl_data.ending = finish.number()
            line.xl_data.reference.cell = ref_cell

            if update_cell:
                line.xl_data.cell = cell

        return sheet

    def _combine_segments(
        self, sheet, column, line, row_container=None, indent=0
    ):
        """


        LineChef._combine_segments() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``column`` must be a column number reference
        --``line`` must be an instance of LineItem
        --``set_labels`` must be a boolean; True will set labels for line
        --``indent`` is amount of indent

        Adds the combination to the current row.
        """
        processed = any((
            line.xl_data.consolidated.cell,
            line.xl_data.derived.cell,
            line.xl_data.detailed.cell,
            line.xl_data.reference.cell
        ))

        write_value = not processed

        if write_value:
            line_label = indent * " " + line.title  # + ': segment'
            finish = row_container.add_group(
                line.title, size=1, label=line_label, bbid=line.id.bbid.hex
            )
            cell = sheet.cell(column=column, row=finish.number())

            # Blank or hard-coded line
            cell.value = line.value
            CellStyles.format_hardcoded(cell)

            line.xl_data.ending = finish.number()
            line.xl_data.cell = cell

        return sheet

    def _add_details(
        self, sheet, column, line, row_container=None, indent=0,
        set_labels=False, start_bal=False, check=False,
    ):
        """


        LineChef._add_detail() -> Worksheet

        --``sheet`` must be an instance of openpyxl Worksheet
        --``line`` must be an instance of LineItem
        --``row_container`` coordinate anchor on the row axis
        --``column`` must be a column number reference
        --``indent`` is amount of indent

        Displays detail sources.
        """
        details = line.get_ordered()
        if details:
            sub_indent = indent + LineItem.TAB_WIDTH
            detail_summation = ""

            sheet.bb.outline_level += 1
            for detail in details:
                self.chop_line(
                    sheet=sheet,
                    column=column,
                    line=detail,
                    row_container=row_container,
                    set_labels=set_labels,
                    indent=sub_indent,
                    check=check,
                    start_bal=start_bal,
                )

                link_template = FormulaTemplates.ADD_COORDINATES
                include = detail.xl_data.cell.parent is not sheet

                cos = detail.xl_data.get_coordinates(include_sheet=include)
                link = link_template.format(coordinates=cos)
                detail_summation += link

            sheet.bb.outline_level -= 1
            row_container.calc_size()

            if line.sum_details:
                if line.xl_format.blank_row_before:
                    row_container.add_group('spacer_details', size=1)

                # subtotal row for details
                line_label = indent * " " + line.title
                finish = row_container.add_group(
                    line_label, size=1, label=line_label, bbid=line.id.bbid.hex
                )
                subtotal_cell = sheet.cell(column=column, row=finish.number())
                subtotal_cell.set_explicit_value(
                    detail_summation, data_type=TypeCodes.FORMULA
                )
                line.xl_data.detailed.ending = finish.number()
                line.xl_data.detailed.cell = subtotal_cell
                line.xl_data.cell = subtotal_cell

        return sheet

    @staticmethod
    def _get_formula_steps(sheet, line, driver_data):
        """

        Function gets and returns steps in formula calculation based on the
        template calculation for this line.
        """
        keys = []  # Return a blank list if no formula steps.

        try:
            template_xl, xl_format = sheet.bb.line_directory[line.id.bbid]
        except KeyError:
            template_xl = None

        if template_xl:
            # find template_driver_data where name = driver_data.name
            for check_data in template_xl.derived.calculations:
                if check_data.name == driver_data.name:
                    keys = check_data.formula.keys()
                    break
        else:
            keys = driver_data.formula.keys()

        return keys
