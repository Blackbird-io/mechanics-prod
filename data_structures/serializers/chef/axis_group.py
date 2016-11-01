# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: data_structures.serializers.chef.axis_group
"""

Module defines abstract anchors for row and column axes.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
AxisGroup             abstract anchors for an axis
====================  =========================================================
"""




# Imports
from collections import OrderedDict
from openpyxl.utils import get_column_letter




# Module Globals
# n/a

# Classes
class AxisGroup:
    """

    A placeholder for spreadsheet layout on the R and C axes. Defines the order
    of rows or cols without specifying their exact location at first.
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    by_name               unique names of groups -> index in groups
    extra                 anything else we need to remember
    groups                ordered container for groups
    name                  our tag, unique within parent
    offset                our offset from the predecessor, or from the tip of
                          container if we are at the top of container
    outline               Excel outline level for group
    path                  shows path to ourselves from the very top
    size                  total span from tip to end, fixed or calculated,
                          does not include own offset
    tip                   0-base starting position of this group in container
                          after offset is applied, may be None if unknown

    FUNCTIONS:
    add_group()           recursively adds nested groups or returns existing
    calc_size()           calculates the span of this group and children
    get_corner_address()  converts row and col coordinates into 'A1' for Excel
    get_group()           same as add_group, but will not add
    get_subgroups()       convenience iterator over subgroups
    number()              converts 0 -> 1 for Excel coordinate system
    resolve_cells()       converts DelayedCell to regular cells
    ====================  =====================================================
    """
    def __init__(self, **kargs):
        """

        AxisGroup.__init__() -> None

        Any unparsed (key, value) pairs in kargs will be placed into the
        ``extra`` dict, to facilitate manipulation and formatting.
        """
        self.tip = kargs.get('tip')
        self.name = kargs.get('name')
        self.path = kargs.get('path')
        self.size = kargs.get('size')
        self.offset = kargs.get('offset')
        self.outline = kargs.get('outline', 0)
        self.groups = []
        self.by_name = {}

        # any extra information, e.g. labels
        self.extra = {}
        # cells will be remembered in order added
        self.cells = OrderedDict()
        # everything else sent to constructor is put into extra
        for k, v in kargs.items():
            if k not in self.__dict__:
                self.extra[k] = v

    def add_group(
        self, *path, size=None, offset=None, **kargs
    ):
        """

        AxisGroup.add_group() -> AxisGroup

        Walks the ``path`` adding groups at nested levels, or matching existing
        ones.
        """
        group = self

        for name in path:
            # anyting that formats will do as a tag
            name = format(name)

            if name in group.by_name:
                group_idx = group.by_name[name]
                group = group.groups[group_idx]
            else:
                new_group = self._make_group(
                    name, size, offset, **kargs
                )
                atindex = kargs.pop('index', None)
                if atindex is None:
                    group.by_name[name] = len(group.groups)
                    group.groups.append(new_group)
                else:
                    # main use case: -1, to preserve the position of last item
                    # shift by_name to accommodate the insertion point
                    if atindex < 0:
                        target_index = min(0, len(group.groups) + atindex)
                    else:
                        target_index = min(atindex, len(group.groups))
                    group.by_name[name] = target_index
                    group.groups.insert(target_index, new_group)
                    for n, i in group.by_name.items():
                        if i >= target_index:
                            group.by_name[n] = i + 1
                group = new_group
        return group

    def calc_size(self, render=None):
        """

        AxisGroup.calc_size() -> int

        --``render`` callable to apply to each subgroup

        Recursively calculates and sets the span sizes of all subgroups.
        Only works if own tip is set.
        Group size is the sum of subgroup sizes and their offsets.
        Own offset is not included in own size, but added to parent's size.
        Will adjusts the starting locations of groups based on predecessor's
        sizes.
        """
        mysize = 0
        for group in self.groups:
            group.tip = self.tip + mysize + (group.offset or 0)
            if group.groups:
                group_size = group.calc_size(render=render)
            else:
                # group without subgroups or size is allowed, counts as empty
                group_size = group.size or 0
            # accumulate subgroup sizes into mysize
            mysize += group_size + (group.offset or 0)
            # if a callback is given, apply it to this group
            if render:
                render(group)
        # if all subgroups are empty, turn to self.size
        if not mysize:
            mysize = self.size or 0
        self.size = mysize
        return mysize

    def find_all(self, *path):
        """

        AxisGroup.find_all() -> iter -> AxisGroup

        Convenience iterator over the subgroups at a certain level in the path.
        If None is in the path, all groups are returned at that level.
        To iterate over all terminal groups in the ``path``, conclude ``path``
        with None.
        """
        if path:
            name = path[0]
            for group in self.groups:
                if name is not None and group.name != format(name):
                    continue
                if len(path) > 1:
                    yield from group.find_all(*path[1:])
                else:
                    yield group

    def get_corner_address(self, col_group, row_path=[], col_path=[]):
        """

        AxisGroup.get_corner_address() -> str

        --``col_group`` AxisGroup, column locator

        Excel convenience: finds the address label of our intersection with
        a column locator (top left cell address), e.g. 'C9'.
        Only makes sense if we are a row locator, and the cross-locator is a
        column locator.
        """
        row = self.get_group(*row_path)
        col = col_group.get_group(*col_path)
        if row and col:
            rownum = row.number()
            colnum = col.number()
            letter = get_column_letter(colnum)
            return '{}{}'.format(letter, rownum)

    def get_group(self, *path):
        """

        AxisGroup.add_group() -> None

        Walks the ``path`` to match an existing group, returns None if not
        found (instead of creating one as add_group would do).
        """
        group = self
        for name in path:
            name = format(name)
            if name in group.by_name:
                group_idx = group.by_name[name]
                group = group.groups[group_idx]
            else:
                return None
        return group

    def get_range_address(self, col_group, row_path=[], col_path=[]):
        """

        AxisGroup.get_corner_address() -> str

        --``col_group`` AxisGroup, column locator

        Excel convenience: finds the full range address of our intersection
        with a column locator, e.g. 'C2:E9'. Sizes need to be set.
        Only makes sense if we are a row locator, and the cross-locator is a
        column locator.
        """
        row = self.get_group(*row_path)
        col = col_group.get_group(*col_path)
        if row and col:
            rowtip = row.number()
            rowend = row.number() + row.size - 1
            coltip = col.number()
            colend = col.number() + col.size - 1
            coltip = get_column_letter(coltip)
            colend = get_column_letter(colend)
            return '{}{}:{}{}'.format(coltip, rowtip, colend, rowend)

    def get_span(self, *path, letters=False):
        """

        AxisGroup.get_span() -> (int, int) or (str, str)

        --``letters`` return column letters instead of numbers

        Excel convenience: first and last row/col numbers of the Excel range
        spanned by us. As column letters, optionally.
        Breaks if size has not been set.
        """
        group = self.get_group(*path)
        tip = group.tip + 1
        end = group.tip + group.size
        if letters:
            tip = get_column_letter(tip)
            end = get_column_letter(end)
        return tip, end

    def get_subgroups(self, name):
        """

        AxisGroup.get_subgroups() -> iter -> AxisGroup

        Convenience iterator over the subgroups of a group given by ``name``.
        """
        if name in self.by_name:
            group_idx = self.by_name[name]
            for group in self.groups[group_idx].groups:
                yield group

    def number(self):
        """

        AxisGroup.number() -> int

        Excel convenience: converts own 0-base location into 1-base.
        """
        if self.tip is not None:
            return self.tip + 1

    def resolve_cells(self):
        """

        AxisGroup.resolve_cells() -> None

        Looks for .cells attribute on each of subgroups recursively and
        evaluates their actual locations.
        Breaks if size has not been set.
        """
        for group in self.groups:
            if group.groups:
                group.resolve_cells()
            for path, soft_cell in group.cells.items():
                if not soft_cell.cell:
                    soft_cell.resolve()
            group.cells.clear()

    # *************************************************************************#
    #                          NON-PUBLIC METHODS                              #
    # *************************************************************************#

    def _make_group(
        self, name, size=None, offset=None, **kargs
    ):
        """

        AxisGroup._make_group() -> AxisGroup

        Constructs AxisGroup instance parameters from inputs. Does some of the
        work of .calc_size(), for cases when span can be calculated at
        construction time.
        """
        name = format(name)

        # decorative property, for debugging
        path = name
        if self.path:
            path = self.path + '.' + path

        # child's head position, if we can tell what it is
        tip = None
        if self.groups:
            # from peer predecessor
            end_group = self.groups[-1]
            if end_group.tip is not None and end_group.size is not None:
                tip = end_group.tip + end_group.size
        elif self.tip is not None:
            # from container
            tip = self.tip

        # shift the starting position, creating an implicit spacer
        if tip is not None and offset:
            tip += offset

        new_group = AxisGroup(
            name=name,
            path=path,
            tip=tip,
            size=size,
            offset=offset,
            **kargs
        )
        return new_group
