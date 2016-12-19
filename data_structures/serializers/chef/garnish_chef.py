# PROPRIETARY AND CONFIDENTIAL
# Property of Blackbird Logical Applications, LLC
# Copyright Blackbird Logical Applications, LLC 2016
# NOT TO BE CIRCULATED OR REPRODUCED WITHOUT PRIOR WRITTEN APPROVAL

# Blackbird Environment
# Module: serializers.chef.garnish_chef
"""

Class for creating dynamic Excel representations of Blackbird Engine models.

Chef modules write formulas to cells explicitly, using the set_explicit_value()
method, to make sure Excel interprets the strings correctly.
====================  =========================================================
Attribute             Description
====================  =========================================================

DATA:
n/a

FUNCTIONS:
n/a

CLASSES:
GarnishChef           adds garnishes and builds foundation for model Excel book
====================  =========================================================
"""




# Imports
import chef_settings
import datetime
import openpyxl as xlio
import os

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import WHITE, BLACK

from ._chef_tools import add_scenario_selector
from .bb_workbook import BB_Workbook as Workbook
from .cell_styles import CellStyles
from .data_types import TypeCodes
from .field_names import FieldNames
from .formulas import FormulaTemplates
from .line_chef import LineChef
from .sheet_style import SheetStyle
from .tab_names import TabNames




# Constants
IMAGE_PATH = os.path.join(
    os.path.dirname(__file__), 'static', 'blackbird_engine_2X_410x120.png'
)

# Module Globals
line_chef = LineChef()

get_column_letter = xlio.utils.get_column_letter
bounding_box = xlio.drawing.image.bounding_box


# Classes
class GarnishChef:
    """

    Class initiates construction on model Excel workbook, adding garnishes,
    such as Driver Tab, Timeline Tab, Cover Tab
    ====================  =====================================================
    Attribute             Description
    ====================  =====================================================

    DATA:
    n/a

    FUNCTIONS:
    add_garnishes()       method creates a workbook and adds base information
    ====================  =====================================================
    """

    @staticmethod
    def add_garnishes(model, book=None, report=False, last_date=None):
        """


        GarnishChef.add_garnishes() -> BB_Workbook

        --``model`` is a Blackbird model instance
        --``report`` is a bool, whether working on a report
        --``last_date`` datetime.date; last available report date or last report
                        date in requested time period

        Return a workbook with:
           cover
           scenarios (only if report is False)
        """
        if book is None:
            book = Workbook()
        else:
            if len(book.worksheets) > 1:
                if 'Transcript' in book.sheetnames:
                    idx = 2
                else:
                    idx = 1

                # two worksheets means 1 transcript sheet and 1 spare tab
                otab = book.create_sheet("Original >>", index=idx)
                otab.sheet_properties.tabColor = chef_settings.COVER_TAB_COLOR
                SheetStyle.style_sheet(otab)
                book.original_tab_count = len(book.worksheets) - idx

        book.properties.creator = chef_settings.WORKBOOK_AUTHOR

        GarnishChef._create_cover_tab(book, model, report=report, last_report=last_date)

        GarnishChef._create_scenarios_tab(book, model, report=report)

        return book

    # *************************************************************************#
    #                           NON-PUBLIC METHODS                             #
    # *************************************************************************#
    @staticmethod
    def _create_cover_tab(book, model, report=False, last_report=None):
        """


        GarnishChef._create_cover_tab() -> None

        --``book`` is an instance of B_Workbook
        --``model`` is a Blackbird Engine model
        --``report`` is a bool, whether working on a report
        --``last_report`` datetime.date; last available report date or last report
                        date in requested time period

        Method adds a cover tab to the workbook
        """
        if report:
            book.remove_sheet(book.active)
        else:
            company = model.get_company()

            sheet = book.active
            sheet.title = TabNames.COVER
            SheetStyle.style_sheet(sheet, label_areas=False)

            row = sheet.row_dimensions[1]
            row.height = 9

            for r in range(8, 24):
                row = sheet.row_dimensions[r]
                row.height = 21.75

            row = sheet.row_dimensions[19]
            row.height = 9

            column = sheet.column_dimensions['A']
            column.width = 10.71

            column = sheet.column_dimensions['D']
            column.width = 14

            column = sheet.column_dimensions['E']
            column.width = 31.71

            column = sheet.column_dimensions['F']
            column.width = 14

            if chef_settings.INCLUDE_LOGO:
                img = xlio.drawing.image.Image(IMAGE_PATH,
                                               size=(250, None))
                img.anchor(sheet.cell('B2'))

                sheet.add_image(img)

            CellStyles.format_border_group(sheet, 2, 8, 9, 23,
                                            border_style='double')

            cell = sheet.cell('E11')

            title = company.name.title()

            cell.value = title
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=18, bold=True, underline='single')

            cell = sheet.cell('D14')
            cell.value = chef_settings.DATE_LABEL
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(size=12, bold=True)

            date_label = chef_settings.REF_DATE_LABEL

            cell = sheet.cell('D15')
            cell.value = date_label
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(size=12, bold=True)

            cell = sheet.cell('D16')
            cell.value = chef_settings.QCOUNT_LABEL
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(size=12, bold=True)

            cell = sheet.cell(chef_settings.COVER_DATE_CELL)
            cell.value = datetime.date.today()
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(size=12)

            use_date = model.time_line.ref_date

            cell = sheet.cell('F15')
            cell.value = use_date
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(size=12)

            questions = []
            for i in model.transcript:
                q = i[0]['q_in']
                if q:
                    questions.append(q['prompt'])

            cell = sheet.cell('F16')
            cell.value = len(questions)
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(size=12)

            cell = sheet.cell('C18')
            cell.value = chef_settings.ESTIMATED_LABEL
            cell.font = Font(color=WHITE, size=11, bold=True)
            cell.fill = PatternFill(start_color=BLACK,
                                    end_color=BLACK,
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells('C18:G18')

            cell = sheet.cell('C20')
            cell.value = chef_settings.DISCLAIMER_TEXT
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            sheet.merge_cells('C20:G22')

            sheet.sheet_properties.tabColor = chef_settings.COVER_TAB_COLOR

    @staticmethod
    def _create_scenarios_tab(book, model, report=False):
        """


        GarnishChef._create_scenarios_tab() -> BB_Worksheet


        Return a worksheet that lays out the assumptions used by the model in
        various scenarios.
        """
        my_tab = book.create_sheet(TabNames.SCENARIOS, index=1)

        if my_tab.title != TabNames.SCENARIOS:
            book.drivers_tab_name = my_tab.title

        # Sheet map:
        # A          |  B      | C             | D     | E
        # param names|  blank  | active values | blank | blackbird values

        starting_row = 2

        label_column = 2
        in_effect_column = 12
        custom_column = 4
        base_case_column = 6

        area = my_tab.bb.add_area(FieldNames.PARAMETERS)
        timeline = my_tab.bb.add_area(FieldNames.TIMELINE)

        area.columns.by_name[FieldNames.LABELS] = label_column
        timeline.columns.by_name[FieldNames.LABELS] = label_column
        current_row = starting_row

        if not report:
            area.columns.by_name[FieldNames.VALUES] = in_effect_column
            area.columns.by_name[FieldNames.CUSTOM_CASE] = custom_column
            area.columns.by_name[FieldNames.BASE_CASE] = base_case_column

            book.set_scenario_names(model)
            scenario_columns = book.scenario_names[1:]

            add_scenario_selector(my_tab, label_column, current_row,
                                  book.scenario_names)
            selector_cell = my_tab.cell(row=current_row, column=custom_column)
            selector_cell.value = FieldNames.CUSTOM
            my_tab.bb.general.rows.by_name[FieldNames.SELECTOR] = current_row

            current_row += 3
            # Make scenario label cells
            custom_cell = my_tab.cell(column=custom_column, row=current_row)
            custom_cell.value = FieldNames.CUSTOM
            CellStyles.format_scenario_label(custom_cell)

            for i, s in enumerate(scenario_columns):
                scen_cell = my_tab.cell(
                    column=base_case_column + i, row=current_row
                )
                scen_cell.value = s.title()
                CellStyles.format_scenario_label(scen_cell)
                if i > 0:
                    # add columns for other cases to area
                    area.columns.by_name[s.lower() + "_case"] = base_case_column + i

            active_label_cell = my_tab.cell(column=in_effect_column,
                                            row=current_row)
            active_label_cell.value = FieldNames.IN_EFFECT
            CellStyles.format_scenario_label(active_label_cell)

        title_row = current_row

        # insert blank row
        current_row += 2

        # storing scenario values in model.time_line is obsolete now, transfer
        # existing values to model.scenarios
        time_line = model.get_timeline()
        base = time_line.parameters
        base.update(time_line.current_period.parameters)

        # now loop through the rest of the time periods and grab parameters
        for per in time_line.iter_ordered():
            if per.end < time_line.current_period.end:
                continue

            existing_parms = set(base.keys())
            tp_parms = set(per.parameters.keys())
            new_parms = tp_parms - existing_parms

            for p in new_parms:
                base[p] = None

        ref_row = 3

        all_scenarios = dict()
        all_scenarios[FieldNames.BASE] = base
        all_scenarios.update(model.scenarios)

        for param_name in sorted(base.keys()):
            # Sort to make sure we display the parameters in stable order,
            # otherwise order could vary from chop to chop on the same model.

            label_cell = my_tab.cell(column=label_column, row=current_row)
            label_cell.value = param_name

            area.rows.by_name[param_name] = current_row

            if not report:
                case_cell = my_tab.cell(column=custom_column,
                                        row=current_row)
                case_cell.value = base[param_name]
                CellStyles.format_hardcoded(case_cell)
                CellStyles.format_parameter(case_cell)

                # Loop through scenarios and add values
                for i, s in enumerate(scenario_columns):
                    case_cell = my_tab.cell(column=base_case_column + i,
                                            row=current_row)
                    case_cell.value = all_scenarios[s].get(param_name, '')
                    CellStyles.format_parameter(case_cell)

                start_cos = custom_cell.coordinate
                end_cos = case_cell.coordinate

                link_template = FormulaTemplates.HLOOKUP
                link = link_template.format(ref_coords=selector_cell.coordinate,
                                            start_coords=start_cos,
                                            end_coords=end_cos,
                                            ref_row=ref_row)

                in_effect_cell = my_tab.cell(column=in_effect_column,
                                             row=current_row)
                in_effect_cell.set_explicit_value(
                    link, data_type=TypeCodes.FORMULA
                )
                CellStyles.format_parameter(in_effect_cell)

            current_row += 1
            ref_row += 1

        if not report:
            # Add cell outline formatting for Scenarios cells here
            CellStyles.format_border_group(
                my_tab,
                custom_column,
                custom_column,
                title_row,
                current_row - 1
            )

            CellStyles.format_border_group(
                my_tab,
                base_case_column,
                base_case_column + i,
                title_row,
                current_row - 1
            )

            CellStyles.format_border_group(
                my_tab,
                in_effect_column,
                in_effect_column,
                title_row,
                current_row - 1
            )

        # Now add the timeline area
        if not report:
            active_column = in_effect_column + 2
            tl_start = active_column
            alpha_master_column = get_column_letter(in_effect_column)
        else:
            active_column = custom_column + 2
            tl_start = active_column
            alpha_master_column = get_column_letter(active_column)

        for period in time_line.iter_ordered(open=time_line.current_period.end):
            timeline.columns.by_name[period.end] = active_column

            SheetStyle.set_column_width(my_tab, active_column)
            # Need this to make sure the parameters Area looks as wide as the
            # timeline. Otherwise, other routines may think that the params
            # area is only one column wide.

            header_cell = my_tab.cell(column=active_column, row=title_row)
            header_cell.value = period.end
            CellStyles.format_date(header_cell)
            col = chef_settings.TIMELINE_HEADER_COLOR
            CellStyles.format_header_label(header_cell, font_color=WHITE,
                                            color=col, bold=False)

            timeline.rows.by_name = area.rows.by_name.copy()

            for k in timeline.rows.by_name:
                # May write the column in undefined order
                param_row = timeline.rows.by_name[k]
                param_cell = my_tab.cell(column=active_column, row=param_row)

                master_cell = my_tab.cell(
                    column=in_effect_column, row=param_row
                )

                link2master = True
                if k in period.parameters:
                    param_cell.value = period.parameters[k]

                    if param_cell.value != master_cell.value:
                        CellStyles.format_hardcoded(param_cell)
                        link2master = False

                if report and alpha_master_column == get_column_letter(active_column):
                    if k not in period.parameters and k in time_line.parameters:
                        param_cell.value = time_line.parameters[k]
                        CellStyles.format_hardcoded(param_cell)

                    link2master = False

                if link2master:
                    link_template = FormulaTemplates.ADD_CELL
                    cos = dict(alpha_column=alpha_master_column, row=param_row)
                    link = link_template.format(**cos)

                    param_cell.set_explicit_value(
                        link, data_type=TypeCodes.FORMULA
                    )

                CellStyles.format_parameter(param_cell)

            active_column += 1

        tl_end = active_column - 1
        timeline.rows.by_name[FieldNames.TITLE] = title_row

        if not report:
            for col in range(tl_start, tl_end + 1):
                alpha = get_column_letter(col)
                my_tab.column_dimensions[alpha].outline_level = 1

            for col in range(base_case_column, in_effect_column + 1):
                alpha = get_column_letter(col)
                my_tab.column_dimensions[alpha].outline_level = 1

            for c in range(1, area.columns.ending + 1):
                SheetStyle.set_column_width(my_tab, c)

            SheetStyle.set_column_width(my_tab, custom_column + 1, 12)
            SheetStyle.set_column_width(my_tab, in_effect_column + 1, 12)
            SheetStyle.set_column_width(my_tab, in_effect_column - 1, 12)

            corner_col = area.columns.by_name[FieldNames.BASE_CASE]
        else:
            corner_col = custom_column + 1

        corner_row = title_row + 1
        corner_cell = my_tab.cell(column=corner_col, row=corner_row)
        my_tab.freeze_panes = corner_cell

        SheetStyle.style_sheet(my_tab, label_areas=False)

        if not report:
            my_tab.sheet_properties.tabColor = chef_settings.SCENARIO_TAB_COLOR
        else:
            my_tab.sheet_properties.tabColor = '808080'

        return my_tab
